import pandas as pd
import numpy as np
import scipy.stats as stats
from datetime import datetime
import os
import re

# ==========================================
# 1. CONFIGURATION
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 自動抓資料夾內最新的問卷 xlsx（檔名含「職涯」或「問卷」）
import glob as _glob
_candidates = sorted(
    _glob.glob(os.path.join(BASE_DIR, "*.xlsx")),
    key=os.path.getmtime,
    reverse=True
)
_candidates = [f for f in _candidates if any(k in os.path.basename(f) for k in ['職涯', '問卷', 'questionnaire', 'survey'])]
if not _candidates:
    raise FileNotFoundError(f"找不到問卷 xlsx，請確認已放入：{BASE_DIR}")
EXCEL_FILE = _candidates[0]
print(f"[自動偵測] 使用問卷檔案：{os.path.basename(EXCEL_FILE)}")
OUTPUT_DIR = os.path.join(BASE_DIR, "Master_Pipeline_Output")

MANUAL_OVERRIDES = {
    'jaychen@trendforce.com': '0710588',
    'huang0447@itri.org.tw': '0315587',
    'baoan5669@gmail.com': '1003082',
    'zxcv70103@gmail.com': '0108108',
    'jhenjiahu@gmail.com': '1230016',
    'zxc52040@gmail.com': '0404983'
}

def clean_str(val):
    if pd.isna(val) or val == "":
        return ""
    return str(val).strip().lower()

def clean_key1_t2t3(val):
    s = clean_str(val)
    if s:
        return s.split()[0] # 去掉後面填寫時間
    return ""

def clean_key2_match_id(val):
    if pd.isna(val) or val == "": return ""
    raw = str(val).replace('.0', '').strip()
    match_val = re.sub(r'\D', '', raw)
    if match_val.isdigit():
        while len(match_val) < 7 and len(match_val) > 0:
            match_val = '0' + match_val
    return match_val

def get_scale_cols(df, keys):
    return [c for c in df.columns if any(k in c for k in keys)]

# ==========================================
# 1. DATA CLEANING & MATCHING MODULE
# ==========================================
def process_phase_data(df, phase_name, k1_idx, k2_idx, k3_idx, is_t2t3=False):
    # DEDUPLICATION: Sort by timestamp (col 0) and keep last distinct key3 (Email/Contact)
    if 'Timestamp' not in df.columns:
        df['Timestamp'] = pd.to_datetime(df.iloc[:, 0], errors='coerce')
    df = df.sort_values(by='Timestamp')
    
    # Extract keys
    if is_t2t3:
        df['key1'] = df.iloc[:, k1_idx].apply(clean_key1_t2t3)
    else:
        df['key1'] = df.iloc[:, k1_idx].apply(clean_str)
        
    df['key2'] = df.iloc[:, k2_idx].apply(clean_key2_match_id)
    
    # key3 is contact/email, apply overrides
    def check_override(val):
        v = clean_str(val)
        if v in MANUAL_OVERRIDES:
            return MANUAL_OVERRIDES[v] # Give them the target key directly as key1 override basically
        return v
    
    df['key3'] = df.iloc[:, k3_idx].apply(check_override)
    
    # Filter out empty key3 before drop_duplicates if we want to drop based on email
    # but some might only have key1 or key2. Let's drop duplicates based on whichever key is available.
    # We will prioritize key3 (email), then key1, then key2.
    df['dedup_id'] = np.where((df['key3'] != "") & (df['key3'] != "nan"), df['key3'],
                     np.where((df['key1'] != "") & (df['key1'] != "nan"), df['key1'], 
                              df['key2']))
                              
    # 找出重複的並印出來 (除了第一筆之外的都會被當作重複刪除)
    duplicates = df[df.duplicated(subset=['dedup_id'], keep='first')]
    if not duplicates.empty:
        print(f"\n[{phase_name}] 發現重複填答，保留最初填答，即將刪除以下後來填寫的 {len(duplicates)} 筆資料:")
        for idx, row in duplicates.iterrows():
            print(f"  - 刪除重複識別碼 (Email/Name): {row['dedup_id']}")
                              
    df = df.drop_duplicates(subset=['dedup_id'], keep='first').copy()
    
    # Score Extraction & Reverse Scoring
    # HP: 6 (Rev: 4, 6)
    # JCP: 6 (Rev: 1, 2, 3, 4, 5)
    # PP: 6
    # DP: 5
    # CI: 8 (Excluding attention check)
    
    all_cols = list(df.columns)

    def find_start(keywords):
        """找到第一個包含任一關鍵字的欄位位置"""
        for i, c in enumerate(all_cols):
            if any(kw in c for kw in keywords):
                return i
        return -1

    def seq_cols(start_idx, n, exclude_kw=None):
        """從 start_idx 連續取 n 欄，可排除含特定關鍵字的欄位"""
        result = []
        i = start_idx
        while len(result) < n and i < len(all_cols):
            c = all_cols[i]
            if exclude_kw is None or exclude_kw not in c:
                result.append(c)
            i += 1
        return result

    # 找各量表起始位置（HP 是第一個量表，其他依序在後）
    hp_start  = find_start(['晉升的可能性是有限', '晉升的可能性'])
    jcp_start = find_start(['學習與成長', '挑戰性']) if hp_start == -1 else hp_start + 6
    pp_start  = jcp_start + 6
    dp_start  = pp_start  + 6
    ci_start  = dp_start  + 5

    hp_cols  = seq_cols(hp_start,  6) if hp_start  != -1 else []
    jc_cols  = seq_cols(jcp_start, 6) if jcp_start != -1 else []
    pp_cols  = seq_cols(pp_start,  6) if pp_start  != -1 else []
    dp_cols  = seq_cols(dp_start,  5) if dp_start  != -1 else []
    # CI：連取9欄後排除注意力檢核題（含「這題請選擇」）
    ci_raw   = seq_cols(ci_start, 9) if ci_start != -1 else []
    ci_cols  = [c for c in ci_raw if '這題請選擇' not in c][:8]

    # -------------------------------------------------------
    # 將各題轉為數值並做反向計分，然後重新命名為英文欄位
    # 命名規則：HP1_T1, HP2_T1 ... CI8_T3（與波次對應）
    # 反向計分：HP4, HP6 / JCP1~JCP5（6點量表，反向 = 6 - 原值）
    # -------------------------------------------------------
    def to_num(val): return pd.to_numeric(val, errors='coerce')
    def rev(val):
        v = to_num(val)
        return 6 - v if pd.notnull(v) else np.nan

    sfx = f'_{phase_name}'  # e.g. _T1, _T2, _T3

    # HP（6題，反向：4、6）
    for i, col in enumerate(hp_cols[:6]):
        item_num = i + 1
        new_name = f'HP{item_num}{sfx}'
        df[new_name] = df[col].apply(rev if item_num in [4, 6] else to_num)

    # JCP（6題，反向：1~5）
    for i, col in enumerate(jc_cols[:6]):
        item_num = i + 1
        new_name = f'JCP{item_num}{sfx}'
        df[new_name] = df[col].apply(rev if item_num <= 5 else to_num)

    # PP（6題，無反向）
    for i, col in enumerate(pp_cols[:6]):
        df[f'PP{i+1}{sfx}'] = df[col].apply(to_num)

    # DP（5題，無反向）
    for i, col in enumerate(dp_cols[:5]):
        df[f'DP{i+1}{sfx}'] = df[col].apply(to_num)

    # CI（8題，無反向；CI5 為注意力題已在 ci_cols 過濾）
    for i, col in enumerate(ci_cols[:8]):
        df[f'CI{i+1}{sfx}'] = df[col].apply(to_num)

    # -------------------------------------------------------
    # 控制變數（只在 T1 波次有，T2/T3 不重複收）
    # -------------------------------------------------------
    def encode_val(val, kw_list):
        s = str(val)
        for idx, kw in enumerate(kw_list):
            if kw in s: return idx + 1
        return np.nan

    if phase_name == 'T1':
        col_gender = [c for c in df.columns if '性別' in c]
        col_age    = [c for c in df.columns if '年齡' in c]
        col_edu    = [c for c in df.columns if '教育程度' in c]
        col_mar    = [c for c in df.columns if '婚姻狀況' in c]
        col_pos    = [c for c in df.columns if '工作職級' in c]
        col_ind    = [c for c in df.columns if '產業別' in c]
        col_size   = [c for c in df.columns if '公司規模' in c]
        col_ny     = [c for c in df.columns if '現職年資 (年)' in c]
        col_nm     = [c for c in df.columns if '現職年資 (月)' in c]
        col_ty     = [c for c in df.columns if '工作總年資 (年)' in c]
        col_tm     = [c for c in df.columns if '工作總年資 (月)' in c]

        if col_gender: df['Gender'] = df[col_gender[0]].apply(lambda x: encode_val(x, ["男", "女", "其他"]))
        if col_age:    df['Age']    = pd.to_numeric(df[col_age[0]], errors='coerce')
        if col_edu:    df['Education'] = df[col_edu[0]].apply(lambda x: encode_val(x, ["高中", "專科", "大學", "碩士", "博士"]))
        if col_mar:    df['Marriage']  = df[col_mar[0]].apply(lambda x: encode_val(x, ["未婚", "無子女", "有子女", "其他"]))
        if col_pos:    df['Position']  = df[col_pos[0]].apply(lambda x: encode_val(x, ["一般", "中階", "基層", "高階"]))
        if col_ind:    df['Industry']  = df[col_ind[0]].apply(lambda x: encode_val(x, ["製造", "科技", "金融", "服務", "醫療", "教育", "公部門", "其他"]))
        if col_size:   df['OrgSize']   = df[col_size[0]].apply(lambda x: encode_val(x, ["30人", "31", "101", "501", "1001"]))
        if col_ny and col_nm:
            df['NowJobTenure'] = pd.to_numeric(df[col_ny[0]], errors='coerce').fillna(0) * 12 \
                               + pd.to_numeric(df[col_nm[0]], errors='coerce').fillna(0)
        if col_ty and col_tm:
            df['JobTenure'] = pd.to_numeric(df[col_ty[0]], errors='coerce').fillna(0) * 12 \
                            + pd.to_numeric(df[col_tm[0]], errors='coerce').fillna(0)

    # -------------------------------------------------------
    # 績效考核相關欄位（三波都收，加波次後綴）
    # PM_Has_Tx      : 是否有績效考核 (1=是, 0=否)
    # PM_Supervisor_Tx 等 : 考核形式 各 0/1
    # PM_Result_Tx   : 考核結果 (1=負向, 2=中立/持平, 3=正向)
    # PM_Help_Tx     : 考核對職涯幫助程度 (1-5 量尺)
    # -------------------------------------------------------
    col_pm_has  = [c for c in df.columns if '是否有進行績效考核' in c or '是否已進行績效考核' in c or ('績效考核' in c and '是否' in c)]
    col_pm_form = [c for c in df.columns if '考核」通常包含哪些形式' in c or '考核」包含' in c or '考核」通常包含' in c or ('考核' in c and '包含' in c and '形式' in c)]
    col_pm_res  = [c for c in df.columns if '考核結果/回饋性質' in c or ('考核結果' in c and '回饋' in c) or ('考核' in c and '性質' in c)]
    col_pm_help = [c for c in df.columns if '職涯發展的幫助程度' in c or ('考核' in c and '幫助' in c)]

    # T2/T3 fallback：若關鍵字沒找到，用欄位位置（PM 區固定在量表前 col 3~6）
    if not col_pm_has and phase_name != 'T1':
        scale_start = find_start(['晉升的可能性是有限', '晉升的可能性'])
        if scale_start >= 4:
            col_pm_has  = [all_cols[scale_start - 4]] if scale_start >= 4 else []
            col_pm_form = [all_cols[scale_start - 3]] if scale_start >= 3 else []
            col_pm_res  = [all_cols[scale_start - 2]] if scale_start >= 2 else []
            col_pm_help = [all_cols[scale_start - 1]] if scale_start >= 1 else []

    if col_pm_has:
        df[f'PM_Has{sfx}'] = df[col_pm_has[0]].apply(lambda x: 1 if '是' in str(x) else 0)

    if col_pm_form:
        form_col = df[col_pm_form[0]].fillna('').astype(str)
        df[f'PM_Supervisor{sfx}'] = form_col.apply(lambda x: 1 if '主管' in x else 0)
        df[f'PM_Self{sfx}']       = form_col.apply(lambda x: 1 if '自評' in x or '自我評核' in x else 0)
        df[f'PM_Interview{sfx}']  = form_col.apply(lambda x: 1 if '面談' in x else 0)
        df[f'PM_Other{sfx}']      = form_col.apply(lambda x: 1 if '其他' in x else 0)

    if col_pm_res:
        df[f'PM_Result{sfx}'] = df[col_pm_res[0]].apply(
            lambda x: 3 if '正向' in str(x) else (2 if '中' in str(x) or '持平' in str(x) else (1 if '負向' in str(x) else np.nan))
        )

    if col_pm_help:
        df[f'PM_Help{sfx}'] = pd.to_numeric(df[col_pm_help[0]], errors='coerce')

    return df, hp_cols, jc_cols, pp_cols, dp_cols, ci_cols

def safe_match(val1, val2):
    return (val1 != "" and val1 != "nan" and val2 != "" and val2 != "nan" and val1 == val2)

def perform_matching():
    print("Loading data...")
    t1_raw = pd.read_excel(EXCEL_FILE, sheet_name='第一階段')
    t2_raw = pd.read_excel(EXCEL_FILE, sheet_name='第二階段')
    t3_raw = pd.read_excel(EXCEL_FILE, sheet_name='第三階段')

    tracking = {
        'T1_Raw': len(t1_raw), 'T2_Raw': len(t2_raw), 'T3_Raw': len(t3_raw),
    }

    # Basic T1 filtering
    attn_col_t1 = [c for c in t1_raw.columns if '這題請選擇「4」' in c][0]
    ans_count_col = [c for c in t1_raw.columns if '共需填寫幾次問卷' in c][0]
    job_col = [c for c in t1_raw.columns if '就業狀態' in c][0]
    
    t1_raw = t1_raw[t1_raw[attn_col_t1] == 4]
    tracking['T1_Pass_Attn'] = len(t1_raw)
    t1_raw = t1_raw[t1_raw[ans_count_col].astype(str).str.contains('3次', na=False)]
    tracking['T1_Pass_Freq'] = len(t1_raw)
    invalid_jobs = ["兼職", "待業", "學生", "自由", "自營"]
    t1_raw = t1_raw[~t1_raw[job_col].apply(lambda x: any(k in str(x) for k in invalid_jobs))]
    tracking['T1_Pass_Job'] = len(t1_raw)
    print(f"T1 Basic Filtered: {len(t1_raw)}")
    
    # T2 & T3 Attention Filter
    attn_col_t2 = [c for c in t2_raw.columns if '這題請選擇「2」' in c][0]
    t2_raw = t2_raw[t2_raw[attn_col_t2] == 2]
    tracking['T2_Pass_Attn'] = len(t2_raw)

    attn_col_t3 = [c for c in t3_raw.columns if '這題請選擇「2」' in c][0]
    t3_raw = t3_raw[t3_raw[attn_col_t3] == 2]
    tracking['T3_Pass_Attn'] = len(t3_raw)

    # Process each phase (Extract keys, deduplicate, reverse score, average)
    # Mapping indices based on tests:
    # T1: BF(57)=k1, AO(40)=k2, BA(52)=k3
    # T2: C(2)=k1, AN(39)=k2, AO(40)=k3
    # T3: C(2)=k1, AN(39)=k2, AO(40)=k3
    
    t1, hp_t1, jc_t1, pp_t1, dp_t1, ci_t1 = process_phase_data(t1_raw, 'T1', 57, 40, 52, is_t2t3=False)
    t2, _, _, _, _, _ = process_phase_data(t2_raw, 'T2', 2, 39, 40, is_t2t3=True)
    t3, _, _, _, _, _ = process_phase_data(t3_raw, 'T3', 2, 39, 40, is_t2t3=True)
    
    print(f"After Deduplication: T1={len(t1)}, T2={len(t2)}, T3={len(t3)}")

    # Tri-Matching Logic
    t1['System_ID'] = range(len(t1))
    
    t2['Matched_T1_ID'] = -1
    for idx, row in t2.iterrows():
        match = t1[
                   t1.apply(lambda r: safe_match(r['key1'], row['key1']) or 
                                      safe_match(r['key2'], row['key2']) or 
                                      safe_match(r['key3'], row['key3']) or
                                      safe_match(r['key1'], row['key3']) # Allow cross-matching override
                           , axis=1)
                   ]
        if not match.empty:
            t2.at[idx, 'Matched_T1_ID'] = match.iloc[0]['System_ID']
            
    t3['Matched_T1_ID'] = -1
    for idx, row in t3.iterrows():
        match = t1[
                   t1.apply(lambda r: safe_match(r['key1'], row['key1']) or 
                                      safe_match(r['key2'], row['key2']) or 
                                      safe_match(r['key3'], row['key3']) or
                                      safe_match(r['key1'], row['key3'])
                           , axis=1)
                   ]
        if not match.empty:
            t3.at[idx, 'Matched_T1_ID'] = match.iloc[0]['System_ID']

    # 印出 T2 沒有配對到的名單
    unmatched_t2 = t2[t2['Matched_T1_ID'] == -1]
    if not unmatched_t2.empty:
        print(f"\n[警示] T2 有 {len(unmatched_t2)} 筆資料無法配對回到 T1，這些孤兒名單識別碼為：")
        for _, row in unmatched_t2.iterrows():
            print(f"  - T2 未配對者: {row['key3']} (備用碼: {row['key1']})")

    # 印出 T3 沒有配對到的名單
    unmatched_t3 = t3[t3['Matched_T1_ID'] == -1]
    if not unmatched_t3.empty:
        print(f"\n[警示] T3 有 {len(unmatched_t3)} 筆資料無法配對回到 T1，這些孤兒名單識別碼為：")
        for _, row in unmatched_t3.iterrows():
            print(f"  - T3 未配對者: {row['key3']} (備用碼: {row['key1']})")

    # Merge everything to T1
    merged = t1.copy()
    
    # 定義各波次個別題目欄位名稱（新命名格式）
    def item_cols(prefix, n, wave):
        return [f'{prefix}{i+1}_{wave}' for i in range(n)]

    t1_item_cols = (item_cols('HP',  6, 'T1') + item_cols('JCP', 6, 'T1') +
                    item_cols('PP',  6, 'T1') + item_cols('DP',  5, 'T1') +
                    item_cols('CI',  8, 'T1'))
    t2_item_cols = (item_cols('HP',  6, 'T2') + item_cols('JCP', 6, 'T2') +
                    item_cols('PP',  6, 'T2') + item_cols('DP',  5, 'T2') +
                    item_cols('CI',  8, 'T2'))
    t3_item_cols = (item_cols('HP',  6, 'T3') + item_cols('JCP', 6, 'T3') +
                    item_cols('PP',  6, 'T3') + item_cols('DP',  5, 'T3') +
                    item_cols('CI',  8, 'T3'))

    pm_cols_t1 = ['PM_Has_T1','PM_Supervisor_T1','PM_Self_T1','PM_Interview_T1',
                  'PM_Other_T1','PM_Result_T1','PM_Help_T1']
    pm_cols_t2 = ['PM_Has_T2','PM_Supervisor_T2','PM_Self_T2','PM_Interview_T2',
                  'PM_Other_T2','PM_Result_T2','PM_Help_T2']
    pm_cols_t3 = ['PM_Has_T3','PM_Supervisor_T3','PM_Self_T3','PM_Interview_T3',
                  'PM_Other_T3','PM_Result_T3','PM_Help_T3']
    ctrl_cols  = ['Gender', 'Age', 'Education', 'Marriage',
                  'NowJobTenure', 'JobTenure', 'Position', 'Industry', 'OrgSize']

    t2_keep_cols = (['Matched_T1_ID'] +
                    [c for c in t2_item_cols if c in t2.columns] +
                    [c for c in pm_cols_t2   if c in t2.columns])
    t3_keep_cols = (['Matched_T1_ID'] +
                    [c for c in t3_item_cols if c in t3.columns] +
                    [c for c in pm_cols_t3   if c in t3.columns])

    t2_subset = t2[t2['Matched_T1_ID'] != -1][t2_keep_cols]
    t3_subset = t3[t3['Matched_T1_ID'] != -1][t3_keep_cols]

    tracking['T2_Matched'] = len(t2_subset)
    tracking['T3_Matched'] = len(t3_subset)

    merged = pd.merge(merged, t2_subset, left_on='System_ID', right_on='Matched_T1_ID', how='left')
    merged = pd.merge(merged, t3_subset, left_on='System_ID', right_on='Matched_T1_ID', how='left')

    # Group 欄位：用 T3 是否有資料判斷
    merged['Group'] = np.where(merged[[c for c in t3_item_cols if c in merged.columns]].notna().any(axis=1), 3,
                      np.where(merged[[c for c in t2_item_cols if c in merged.columns]].notna().any(axis=1), 2, 1))

    # 整理最終輸出欄位順序，排除暫時計算欄位（底線開頭）
    final_cols = (['Custom_UID', 'Timestamp', 'Group'] +
                  t1_item_cols +
                  [c for c in ctrl_cols   if c in merged.columns] +
                  [c for c in pm_cols_t1  if c in merged.columns] +
                  t2_item_cols +
                  [c for c in pm_cols_t2  if c in merged.columns] +
                  t3_item_cols +
                  [c for c in pm_cols_t3  if c in merged.columns])
    final_cols = [c for c in final_cols if c in merged.columns and not c.startswith('_')]
    merged = merged[final_cols]

    # Scales for reliability output（使用新英文欄位名稱）
    def ecols(prefix, n, wave='T1'):
        return [f'{prefix}{i+1}_{wave}' for i in range(n)]

    escales = {
        'HP':  ecols('HP',  6, 'T1'),
        'JCP': ecols('JCP', 6, 'T1'),
        'CP':  ecols('HP',  6, 'T1') + ecols('JCP', 6, 'T1'),
        'PP':  ecols('PP',  6, 'T1'),
        'DP':  ecols('DP',  5, 'T1'),
        'CI':  ecols('CI',  8, 'T1'),
    }

    return merged, escales, tracking

# ==========================================
# 2. ANALYSIS MODULE
# ==========================================
def calculate_cronbach_alpha(df):
    df_corr = df.corr()
    N = df.shape[1]
    if N < 2: return np.nan
    rs = np.array([df_corr.iloc[i, j] for i in range(N) for j in range(N) if i > j])
    mean_r = np.mean(rs)
    if pd.isna(mean_r) or mean_r == 0: return np.nan
    alpha = (N * mean_r) / (1 + (N - 1) * mean_r)
    return alpha

def analyze_attrition(merged, track):
    counts = merged['Group'].value_counts()
    
    def anova(var):
        g1 = merged[merged['Group'] == 1][var].dropna()
        g2 = merged[merged['Group'] == 2][var].dropna()
        g3 = merged[merged['Group'] == 3][var].dropna()
        if len(g1)==0 or len(g2)==0 or len(g3)==0:
            return 0,0,0,0,1
        f_stat, p_val = stats.f_oneway(g1, g2, g3)
        return g1.mean(), g2.mean(), g3.mean(), f_stat, p_val

    results_md = "## 1. 樣本流失分析 (Attrition Analysis)\n\n"
    results_md += "### 各階段填答與清理漏斗\n"
    results_md += f"- **T1 (第一階段)**: 原始名單 {track['T1_Raw']} 人 -> 通過注意力檢測 {track['T1_Pass_Attn']} 人 -> 符合填寫條件與任職資格 {track['T1_Pass_Job']} 人 -> **去重複後實際有效樣本 {len(merged)} 人**\n"
    results_md += f"- **T2 (第二階段)**: 原始名單 {track['T2_Raw']} 人 -> 通過注意力檢測 {track['T2_Pass_Attn']} 人 -> **成功配對回 T1 者 {track['T2_Matched']} 人**\n"
    results_md += f"- **T3 (第三階段)**: 原始名單 {track['T3_Raw']} 人 -> 通過注意力檢測 {track['T3_Pass_Attn']} 人 -> **成功配對回 T1 者 {track['T3_Matched']} 人**\n\n"

    results_md += "### ANOVA 各群組人數\n"
    results_md += f"- **只有完成 T1 (Group 1)**: {counts.get(1, 0)} 人\n"
    results_md += f"- **完成 T1, T2 (Group 2)**: {counts.get(2, 0)} 人\n"
    results_md += f"- **完成 T1, T2, T3 (最終有效樣本 Group 3)**: {counts.get(3, 0)} 人\n\n"
    
    # 用新命名欄位（已在 process_phase_data 轉換好）
    # 計算各量表平均供 ANOVA 流失分析用
    def scale_mean(df, prefix, n, wave):
        cols = [f'{prefix}{i+1}_{wave}' for i in range(n)]
        cols = [c for c in cols if c in df.columns]
        return df[cols].mean(axis=1) if cols else np.nan

    merged['_HP_T1']  = scale_mean(merged, 'HP',  6, 'T1')
    merged['_JCP_T1'] = scale_mean(merged, 'JCP', 6, 'T1')
    merged['_CP_T1']  = merged[['_HP_T1','_JCP_T1']].mean(axis=1)
    merged['_PP_T1']  = scale_mean(merged, 'PP',  6, 'T1')
    merged['_DP_T1']  = scale_mean(merged, 'DP',  5, 'T1')
    merged['_CI_T1']  = scale_mean(merged, 'CI',  8, 'T1')
    if 'Edu' not in merged.columns and 'Education' in merged.columns:
        merged['Edu'] = merged['Education']

    anova_stats = {}   # 儲存各變數 p 值供 draft 使用
    results_md += "| 變數 | G1 (僅T1) | G2 (T1+T2) | G3 (T1+T2+T3) | F | p |\n"
    results_md += "|---|---|---|---|---|---|\n"
    for var, name in zip(['Age', '_CP_T1', '_HP_T1', '_JCP_T1', '_DP_T1', '_CI_T1', '_PP_T1'],
                         ['年齡', '整體職涯停滯', '階層停滯', '工作停滯', '決策拖延', '職涯無所作為', '主動型人格']):
        if var in merged.columns:
            m1, m2, m3, f, p = anova(var)
            sig = ' *' if p < .05 else ''
            results_md += f"| **{name}** | {m1:.2f} | {m2:.2f} | {m3:.2f} | {f:.3f} | {p:.3f}{sig} |\n"
            anova_stats[var] = {'m1': m1, 'm2': m2, 'm3': m3, 'F': f, 'p': p}
    results_md += "\n"

    chi_stats = {}
    try:
        chi2_g, p_g, _, _ = stats.chi2_contingency(pd.crosstab(merged['Group'], merged['Gender']))
        chi2_e, p_e, _, _ = stats.chi2_contingency(pd.crosstab(merged['Group'], merged['Edu']))
        results_md += "| 類別變項 | chi² | p |\n"
        results_md += "|---|---|---|\n"
        results_md += f"| **性別比例差異** | {chi2_g:.3f} | {p_g:.3f}{' *' if p_g < .05 else ''} |\n"
        results_md += f"| **教育程度差異** | {chi2_e:.3f} | {p_e:.3f}{' *' if p_e < .05 else ''} |\n"
        results_md += "\n"
        chi_stats = {'gender_chi2': chi2_g, 'gender_p': p_g, 'edu_chi2': chi2_e, 'edu_p': p_e}
    except:
        pass

    return results_md, merged, anova_stats, chi_stats

def run_descriptives_and_correlations(t1, scales):
    results_md = "\n## 2. 敘述性統計與信度分析 (Descriptives & Reliability)\n\n"
    results_md += (
        "> **CP（職涯高原）測量說明**：CP 由兩個次量表組成——"
        "HP（階層停滯，6 題）與 JCP（工作內容停滯，6 題）。"
        "Mplus 中以 HP、JCP 兩個次量表合成分數作為 CP 的兩個指標（parceling）。"
        "信度分別報告 HP（6 題）、JCP（6 題）及合併 CP（12 題）供參考。\n\n"
    )
    results_md += "| 變數 | 說明 | 題數 | 平均數 (M) | 標準差 (SD) | Cronbach's α |\n"
    results_md += "|---|---|---|---|---|---|\n"

    scale_labels = {
        'HP':  'HP 階層停滯（CP 次量表）',
        'JCP': 'JCP 工作內容停滯（CP 次量表）',
        'CP':  'CP 職涯高原（HP+JCP 合併）',
        'PP':  'PP 主動型人格',
        'DP':  'DP 決策拖延',
        'CI':  'CI 職涯無所作為',
    }

    # 用原始題目欄位計算信度，用計算平均供相關矩陣
    scale_means = {}
    for name, cols in scales.items():
        valid_cols = [c for c in cols if c in t1.columns]
        alpha = calculate_cronbach_alpha(t1[valid_cols]) if valid_cols else np.nan
        mean_val = t1[valid_cols].mean(axis=1).mean() if valid_cols else np.nan
        sd_val   = t1[valid_cols].mean(axis=1).std()  if valid_cols else np.nan
        label = scale_labels.get(name, name)
        alpha_str = f"{alpha:.3f}" if not np.isnan(alpha) else "N/A"
        mean_str  = f"{mean_val:.2f}" if not np.isnan(mean_val) else "N/A"
        sd_str    = f"{sd_val:.2f}" if not np.isnan(sd_val) else "N/A"
        results_md += f"| {name} | {label} | {len(valid_cols)} | {mean_str} | {sd_str} | {alpha_str} |\n"
        scale_means[name] = t1[valid_cols].mean(axis=1)

    results_md += "\n## 3. 相關矩陣 (Correlation Matrix)\n\n"
    scale_names = list(scales.keys())

    results_md += "| 變數 | " + " | ".join(scale_names) + " |\n"
    results_md += "|---|" + "|".join(["---"] * len(scale_names)) + "|\n"

    for i, name_r in enumerate(scale_names):
        row_str = f"| **{name_r}** |"
        for j, name_c in enumerate(scale_names):
            if i == j:
                row_str += " 1.00 |"
            elif i < j:
                s1 = scale_means[name_r]
                s2 = scale_means[name_c]
                valid_data = pd.concat([s1, s2], axis=1).dropna()
                if len(valid_data) > 2:
                    r, p = stats.pearsonr(valid_data.iloc[:,0], valid_data.iloc[:,1])
                    star = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else ''
                    row_str += f" {r:.2f}{star} |"
                else:
                    row_str += " - |"
            else:
                row_str += " - |"
        results_md += row_str + "\n"

    # 額外回傳 stats dict 供 draft 使用
    alpha_dict = {}
    corr_dict  = {}
    for name, cols in scales.items():
        valid_cols = [c for c in cols if c in t1.columns]
        alpha_dict[name] = calculate_cronbach_alpha(t1[valid_cols]) if valid_cols else np.nan

    corr_pairs = [('CP','DP'), ('CP','CI'), ('DP','CI'), ('HP','CI'), ('JCP','CI'), ('PP','DP'), ('PP','CI')]
    for a, b in corr_pairs:
        if a in scale_means and b in scale_means:
            valid = pd.concat([scale_means[a], scale_means[b]], axis=1).dropna()
            if len(valid) > 2:
                r, p = stats.pearsonr(valid.iloc[:,0], valid.iloc[:,1])
                corr_dict[f'{a}_{b}'] = {'r': r, 'p': p}

    return results_md, alpha_dict, corr_dict

def generate_r_script(csv_filename, r_script_path=None):
    output_dir = OUTPUT_DIR.replace('\\\\', '/').replace('\\', '/')
    report_path = f"{output_dir}/RICLPM_Report.md"
    log_clpm_path = f"{output_dir}/model_test_log.txt"
    log_ri_path   = f"{output_dir}/model_v2_log.txt"
    csv_full_path = f"{output_dir}/{csv_filename}"

    r_script_content = f"""# ==============================================================================
# RICLPM_Master.R  (自動產生 by pipeline_master.py)
# 整合：資料診斷 → CLPM 比較 → RI-CLPM 精簡比較 →
#       自由 RI-CLPM (FIML) → 限制 RI-CLPM (FIML) → 最佳模型報告
# ==============================================================================

# install.packages(c("lavaan", "readr"))  # 第一次執行時取消註解
library(lavaan)
library(readr)

# ==============================================================================
# STEP 0. 環境設定與資料讀取
# ==============================================================================
DATA_PATH   <- "{csv_full_path}"
OUTPUT_DIR  <- "{output_dir}"
REPORT_PATH <- "{report_path}"

cat("=== Loading data ===\\n")
df <- read_csv(DATA_PATH, show_col_types = FALSE)

VARS <- c("CP_T1","CP_T2","CP_T3","CI_T1","CI_T2","CI_T3","DP_T1","DP_T2","DP_T3")
df_full <- df[complete.cases(df[, VARS]), ]

cat(sprintf("N total (FIML) = %d\\n", nrow(df)))
cat(sprintf("N 三波完整 (listwise) = %d\\n", nrow(df_full)))

# ==============================================================================
# STEP 1. 資料診斷
# ==============================================================================
cat("\\n=== STEP 1: 資料診斷 ===\\n")
cat("Means:\\n");     print(round(colMeans(df_full[VARS], na.rm=TRUE), 3))
cat("SDs:\\n");       print(round(apply(df_full[VARS], 2, sd,  na.rm=TRUE), 3))
cat("Variances:\\n"); print(round(apply(df_full[VARS], 2, var, na.rm=TRUE), 4))
cat("Correlation matrix:\\n"); print(round(cor(df_full[VARS], use="complete.obs"), 3))

# ==============================================================================
# STEP 2. 模型比較：標準 CLPM（lavaan / FIML）
# ==============================================================================
cat("\\n=== STEP 2: 標準 CLPM 模型比較 ===\\n")

LOG_CLPM <- "{log_clpm_path}"
con_clpm <- file(LOG_CLPM, open="wt", encoding="UTF-8")
write_log <- function(text) cat(text, "\\n", file=con_clpm, append=TRUE)
write_log(paste0("N_full=", nrow(df_full), "  N_total=", nrow(df)))

try_lavaan <- function(model_name, syntax, data, missing_method="ml") {{
  write_log(paste0("\\n", strrep("=",50), "\\n", model_name))
  tryCatch({{
    fit  <- lavaan(syntax, data=data, missing=missing_method, estimator="MLR")
    conv <- lavInspect(fit, "converged")
    fmi  <- fitMeasures(fit, c("cfi","rmsea","srmr","chisq","df","pvalue"))
    write_log(paste0("Converged: ", conv))
    write_log(paste0("CFI=",round(fmi["cfi"],3)," RMSEA=",round(fmi["rmsea"],3)," SRMR=",round(fmi["srmr"],3)))
    write_log(paste0("Chi2(",fmi["df"],")=",round(fmi["chisq"],2)," p=",round(fmi["pvalue"],3)))
    if (conv) {{
      params <- parameterEstimates(fit, standardized=TRUE)
      cl <- params[params$op == "~" & !is.na(params$pvalue), ]
      if (nrow(cl) > 0) {{
        write_log("--- paths ---")
        for (i in seq_len(nrow(cl))) {{
          r <- cl[i,]
          write_log(paste0("  ",r$lhs," ~ ",r$rhs,
                           "  Est=",round(r$est,3)," SE=",round(r$se,3),
                           " z=",round(r$z,3)," p=",round(r$pvalue,4),
                           " std=",round(r$std.all,3)))
        }}
      }}
    }}
    return(conv)
  }}, error=function(e) {{ write_log(paste0("ERROR: ", e$message)); FALSE }})
}}

m_std <- '
  CP_T2 ~ CP_T1; CP_T3 ~ CP_T2
  CI_T2 ~ CI_T1; CI_T3 ~ CI_T2
  DP_T2 ~ DP_T1; DP_T3 ~ DP_T2
  CI_T2 ~ CP_T1 + DP_T1; CI_T3 ~ CP_T2 + DP_T2
  CP_T2 ~ CI_T1 + DP_T1; CP_T3 ~ CI_T2 + DP_T2
  DP_T2 ~ CP_T1 + CI_T1; DP_T3 ~ CP_T2 + CI_T2
  CP_T1 ~~ CI_T1; CP_T1 ~~ DP_T1; CI_T1 ~~ DP_T1
'
m_inv <- '
  CP_T2 ~ a1*CP_T1; CP_T3 ~ a1*CP_T2
  CI_T2 ~ a2*CI_T1; CI_T3 ~ a2*CI_T2
  DP_T2 ~ a3*DP_T1; DP_T3 ~ a3*DP_T2
  CI_T2 ~ c1*CP_T1 + c2*DP_T1; CI_T3 ~ c1*CP_T2 + c2*DP_T2
  CP_T2 ~ c3*CI_T1 + c4*DP_T1; CP_T3 ~ c3*CI_T2 + c4*DP_T2
  DP_T2 ~ c5*CP_T1 + c6*CI_T1; DP_T3 ~ c5*CP_T2 + c6*CI_T2
  CP_T1 ~~ CI_T1; CP_T1 ~~ DP_T1; CI_T1 ~~ DP_T1
'

try_lavaan("M1: Standard CLPM (N=168, listwise)",        m_std, df_full, "ml")
try_lavaan("M2: Time-invariant CLPM (N=168, listwise)",  m_inv, df_full, "ml")
try_lavaan("M3: Standard CLPM FIML (N=total)",           m_std, df,      "fiml")
try_lavaan("M4: Time-invariant CLPM FIML (N=total)",     m_inv, df,      "fiml")

write_log("\\n===== ALL CLPM DONE =====")
close(con_clpm)
cat(sprintf("CLPM log: %s\\n", LOG_CLPM))

# ==============================================================================
# STEP 3. 模型比較：RI-CLPM 精簡系列（sem）
# ==============================================================================
cat("\\n=== STEP 3: RI-CLPM 精簡系列比較 ===\\n")

LOG_RI  <- "{log_ri_path}"
con_ri  <- file(LOG_RI, open="wt", encoding="UTF-8")
wl      <- function(...) cat(..., "\\n", file=con_ri, append=TRUE)
wl(paste0("N_full=", nrow(df_full), "  N_all=", nrow(df)))

try_sem <- function(label, syntax, data_in) {{
  wl(paste0("\\n", strrep("=",50), "\\n", label))
  tryCatch({{
    fit  <- sem(syntax, data=data_in, estimator="MLR")
    conv <- lavInspect(fit, "converged")
    fi   <- fitMeasures(fit, c("cfi","rmsea","srmr","chisq","df","pvalue"))
    wl(paste0("Converged=", conv))
    wl(paste0("CFI=",round(fi["cfi"],3)," RMSEA=",round(fi["rmsea"],3)," SRMR=",round(fi["srmr"],3)))
    wl(paste0("Chi2(",fi["df"],")=",round(fi["chisq"],2)," p=",round(fi["pvalue"],3)))
    if (conv) {{
      pe <- parameterEstimates(fit, standardized=TRUE)
      cl <- pe[pe$op == "~" & !is.na(pe$pvalue), ]
      if (nrow(cl) > 0) {{
        wl("--- regression paths ---")
        for (i in seq_len(nrow(cl))) {{
          r <- cl[i,]
          wl(paste0("  ",r$lhs," ~ ",r$rhs,
                    "  b=",round(r$est,3)," SE=",round(r$se,3),
                    " p=",round(r$pvalue,4)," beta=",round(r$std.all,3)))
        }}
      }}
    }}
    return(conv)
  }}, error=function(e) {{ wl(paste0("ERROR: ", e$message)); FALSE }})
}}

m_ri_compact <- "
  RI_CP =~ 1*CP_T1 + 1*CP_T2 + 1*CP_T3
  RI_CI =~ 1*CI_T1 + 1*CI_T2 + 1*CI_T3
  RI_DP =~ 1*DP_T1 + 1*DP_T2 + 1*DP_T3
  wp_CP_T1 =~ 1*CP_T1; wp_CP_T2 =~ 1*CP_T2; wp_CP_T3 =~ 1*CP_T3
  wp_CI_T1 =~ 1*CI_T1; wp_CI_T2 =~ 1*CI_T2; wp_CI_T3 =~ 1*CI_T3
  wp_DP_T1 =~ 1*DP_T1; wp_DP_T2 =~ 1*DP_T2; wp_DP_T3 =~ 1*DP_T3
  CP_T1 ~~ 0*CP_T1; CP_T2 ~~ 0*CP_T2; CP_T3 ~~ 0*CP_T3
  CI_T1 ~~ 0*CI_T1; CI_T2 ~~ 0*CI_T2; CI_T3 ~~ 0*CI_T3
  DP_T1 ~~ 0*DP_T1; DP_T2 ~~ 0*DP_T2; DP_T3 ~~ 0*DP_T3
  wp_CP_T2 ~ a1*wp_CP_T1; wp_CP_T3 ~ a1*wp_CP_T2
  wp_CI_T2 ~ a2*wp_CI_T1; wp_CI_T3 ~ a2*wp_CI_T2
  wp_DP_T2 ~ a3*wp_DP_T1; wp_DP_T3 ~ a3*wp_DP_T2
  wp_CI_T2 ~ c1*wp_CP_T1 + c2*wp_DP_T1
  wp_CI_T3 ~ c1*wp_CP_T2 + c2*wp_DP_T2
  wp_CP_T2 ~ c3*wp_CI_T1 + c4*wp_DP_T1
  wp_CP_T3 ~ c3*wp_CI_T2 + c4*wp_DP_T2
  wp_DP_T2 ~ c5*wp_CP_T1 + c6*wp_CI_T1
  wp_DP_T3 ~ c5*wp_CP_T2 + c6*wp_CI_T2
  RI_CP ~~ RI_CI; RI_CP ~~ RI_DP; RI_CI ~~ RI_DP
  wp_CP_T1 ~~ wp_CI_T1; wp_CP_T1 ~~ wp_DP_T1; wp_CI_T1 ~~ wp_DP_T1
"

try_sem("M_RI: RI-CLPM time-invariant sem() N=168", m_ri_compact, df_full)

wl("\\n===== ALL RI-CLPM COMPACT DONE =====")
close(con_ri)
cat(sprintf("RI-CLPM log: %s\\n", LOG_RI))

# ==============================================================================
# STEP 4. 自由 RI-CLPM（lavaan / FIML）
# ==============================================================================
cat("\\n=== STEP 4: 自由 RI-CLPM (lavaan + FIML) ===\\n")

riclpm_free <- '
  RI_CP =~ 1*CP_T1 + 1*CP_T2 + 1*CP_T3
  RI_CI =~ 1*CI_T1 + 1*CI_T2 + 1*CI_T3
  RI_DP =~ 1*DP_T1 + 1*DP_T2 + 1*DP_T3
  wp_CP_T1 =~ 1*CP_T1; wp_CP_T2 =~ 1*CP_T2; wp_CP_T3 =~ 1*CP_T3
  wp_CI_T1 =~ 1*CI_T1; wp_CI_T2 =~ 1*CI_T2; wp_CI_T3 =~ 1*CI_T3
  wp_DP_T1 =~ 1*DP_T1; wp_DP_T2 =~ 1*DP_T2; wp_DP_T3 =~ 1*DP_T3
  CP_T1 ~~ 0*CP_T1; CP_T2 ~~ 0*CP_T2; CP_T3 ~~ 0*CP_T3
  CI_T1 ~~ 0*CI_T1; CI_T2 ~~ 0*CI_T2; CI_T3 ~~ 0*CI_T3
  DP_T1 ~~ 0*DP_T1; DP_T2 ~~ 0*DP_T2; DP_T3 ~~ 0*DP_T3
  wp_CP_T2 ~ wp_CP_T1; wp_CP_T3 ~ wp_CP_T2
  wp_CI_T2 ~ wp_CI_T1; wp_CI_T3 ~ wp_CI_T2
  wp_DP_T2 ~ wp_DP_T1; wp_DP_T3 ~ wp_DP_T2
  wp_CI_T2 ~ wp_CP_T1 + wp_DP_T1; wp_CI_T3 ~ wp_CP_T2 + wp_DP_T2
  wp_CP_T2 ~ wp_CI_T1 + wp_DP_T1; wp_CP_T3 ~ wp_CI_T2 + wp_DP_T2
  wp_DP_T2 ~ wp_CP_T1 + wp_CI_T1; wp_DP_T3 ~ wp_CP_T2 + wp_CI_T2
  RI_CP ~~ RI_CI; RI_CP ~~ RI_DP; RI_CI ~~ RI_DP
  wp_CP_T1 ~~ wp_CI_T1; wp_CP_T1 ~~ wp_DP_T1; wp_CI_T1 ~~ wp_DP_T1
  wp_CP_T2 ~~ wp_CI_T2; wp_CP_T2 ~~ wp_DP_T2; wp_CI_T2 ~~ wp_DP_T2
  wp_CP_T3 ~~ wp_CI_T3; wp_CP_T3 ~~ wp_DP_T3; wp_CI_T3 ~~ wp_DP_T3
'
fit_free <- tryCatch({{ lavaan(riclpm_free, data=df, missing="fiml", estimator="MLR", bounds=TRUE) }}, error=function(e) NULL)
if(!is.null(fit_free)) print(summary(fit_free, fit.measures=TRUE, standardized=TRUE, rsquare=TRUE)) else cat("M_Free Non-convergence\\n")

# ==============================================================================
# STEP 5. 時間恆定限制 RI-CLPM（lavaan / FIML）
# ==============================================================================
cat("\\n=== STEP 5: 時間恆定限制 RI-CLPM (lavaan + FIML) ===\\n")

riclpm_constrained <- '
  RI_CP =~ 1*CP_T1 + 1*CP_T2 + 1*CP_T3
  RI_CI =~ 1*CI_T1 + 1*CI_T2 + 1*CI_T3
  RI_DP =~ 1*DP_T1 + 1*DP_T2 + 1*DP_T3
  wp_CP_T1 =~ 1*CP_T1; wp_CP_T2 =~ 1*CP_T2; wp_CP_T3 =~ 1*CP_T3
  wp_CI_T1 =~ 1*CI_T1; wp_CI_T2 =~ 1*CI_T2; wp_CI_T3 =~ 1*CI_T3
  wp_DP_T1 =~ 1*DP_T1; wp_DP_T2 =~ 1*DP_T2; wp_DP_T3 =~ 1*DP_T3
  CP_T1 ~~ 0*CP_T1; CP_T2 ~~ 0*CP_T2; CP_T3 ~~ 0*CP_T3
  CI_T1 ~~ 0*CI_T1; CI_T2 ~~ 0*CI_T2; CI_T3 ~~ 0*CI_T3
  DP_T1 ~~ 0*DP_T1; DP_T2 ~~ 0*DP_T2; DP_T3 ~~ 0*DP_T3
  wp_CP_T2 ~ a1*wp_CP_T1; wp_CP_T3 ~ a1*wp_CP_T2
  wp_CI_T2 ~ a2*wp_CI_T1; wp_CI_T3 ~ a2*wp_CI_T2
  wp_DP_T2 ~ a3*wp_DP_T1; wp_DP_T3 ~ a3*wp_DP_T2
  wp_CI_T2 ~ c1*wp_CP_T1 + c2*wp_DP_T1
  wp_CI_T3 ~ c1*wp_CP_T2 + c2*wp_DP_T2
  wp_CP_T2 ~ c3*wp_CI_T1 + c4*wp_DP_T1
  wp_CP_T3 ~ c3*wp_CI_T2 + c4*wp_DP_T2
  wp_DP_T2 ~ c5*wp_CP_T1 + c6*wp_CI_T1
  wp_DP_T3 ~ c5*wp_CP_T2 + c6*wp_CI_T2
  RI_CP ~~ RI_CI; RI_CP ~~ RI_DP; RI_CI ~~ RI_DP
  wp_CP_T1 ~~ wp_CI_T1; wp_CP_T1 ~~ wp_DP_T1; wp_CI_T1 ~~ wp_DP_T1
  wp_CP_T2 ~~ wp_CI_T2; wp_CP_T2 ~~ wp_DP_T2; wp_CI_T2 ~~ wp_DP_T2
  wp_CP_T3 ~~ wp_CI_T3; wp_CP_T3 ~~ wp_DP_T3; wp_CI_T3 ~~ wp_DP_T3
'
fit_constrained <- tryCatch({{ lavaan(riclpm_constrained, data=df, missing="fiml", estimator="MLR", bounds=TRUE) }}, error=function(e) NULL)
if(!is.null(fit_constrained)) print(summary(fit_constrained, fit.measures=TRUE, standardized=TRUE)) else cat("M_Constrained Non-convergence\\n")

# ==============================================================================
# STEP 6. 最佳模型報告 → Markdown
# ==============================================================================
cat("\\n=== STEP 6: 最佳模型報告輸出 ===\\n")

fit_best <- tryCatch({{ sem(m_ri_compact, data=df_full, estimator="MLR", bounds=TRUE) }}, error=function(e) NULL)
con_out  <- file(REPORT_PATH, open="wt", encoding="UTF-8")
w        <- function(...) cat(..., "\\n", file=con_out, append=TRUE)

if(!is.null(fit_best)) {{
  fi   <- fitMeasures(fit_best, c("cfi","tli","rmsea","srmr","chisq","df","pvalue"))
  pe   <- parameterEstimates(fit_best, standardized=TRUE, ci=TRUE)
  conv <- lavInspect(fit_best, "converged")
}} else {{
  fi <- c(cfi=NA, tli=NA, rmsea=NA, srmr=NA, chisq=NA, df=NA, pvalue=NA)
  pe <- data.frame(label=character(0))
  conv <- FALSE
}}

w("# RI-CLPM 分析報告")
w(paste0("**產出時間**: ", format(Sys.time(), "%Y-%m-%d %H:%M")))
w(paste0("**樣本數 (N)**: ", nrow(df_full), " (三波完整填答)"))
w(paste0("**估計方法**: MLR | **模型**: 時間恆定限制 RI-CLPM"))
w(paste0("**是否收斂**: ", ifelse(conv, "✅ 是", "❌ 否")))

w("\\n---\\n## 1. 模型配適指標\\n")
w("| 指標 | 數值 | 建議標準 |"); w("|---|---|---|")
w(paste0("| CFI   | ", round(fi["cfi"],3),   " | >.95 |"))
w(paste0("| TLI   | ", round(fi["tli"],3),   " | >.95 |"))
w(paste0("| RMSEA | ", round(fi["rmsea"],3), " | <.06 |"))
w(paste0("| SRMR  | ", round(fi["srmr"],3),  " | <.08 |"))
w(paste0("| Chi²  | ", round(fi["chisq"],2), " (df=", fi["df"], "), p=", round(fi["pvalue"],3), " |"))

w("\\n---\\n## 2. 自我迴歸效果\\n")
w("| 路徑 | b | SE | p | β |"); w("|---|---|---|---|---|")
ar_rows <- pe[pe$label %in% c("a1","a2","a3") & !duplicated(pe$label), ]
for (i in seq_len(nrow(ar_rows))) {{
  r   <- ar_rows[i,]
  sig <- ifelse(r$pvalue<.001,"***",ifelse(r$pvalue<.01,"**",ifelse(r$pvalue<.05,"*",ifelse(r$pvalue<.10,"†",""))))
  w(paste0("| ",r$lhs," ~ ",r$rhs," | ",round(r$est,3)," | ",round(r$se,3)," | ",round(r$pvalue,4),sig," | ",round(r$std.all,3)," |"))
}}

w("\\n---\\n## 3. 交叉延遲效果\\n")
w("| 路徑 | b | 95% CI | SE | z | p | β |"); w("|---|---|---|---|---|---|---|")
cl_rows <- pe[pe$label %in% paste0("c", 1:6) & !duplicated(pe$label), ]
for (i in seq_len(nrow(cl_rows))) {{
  r      <- cl_rows[i,]
  sig    <- ifelse(r$pvalue<.001,"***",ifelse(r$pvalue<.01,"**",ifelse(r$pvalue<.05,"*",ifelse(r$pvalue<.10,"†",""))))
  ci_str <- paste0("[",round(r$ci.lower,3),", ",round(r$ci.upper,3),"]")
  w(paste0("| ",r$lhs," ~ ",r$rhs," | ",round(r$est,3),sig," | ",ci_str," | ",round(r$se,3)," | ",round(r$z,3)," | ",round(r$pvalue,4)," | ",round(r$std.all,3)," |"))
}}

w("\\n---\\n## 4. 隨機截距共變數\\n")
w("| 路徑 | r (std) | p |"); w("|---|---|---|")
ri_cov <- pe[pe$op=="~~" & grepl("RI_",pe$lhs) & grepl("RI_",pe$rhs), ]
for (i in seq_len(nrow(ri_cov))) {{
  r   <- ri_cov[i,]
  sig <- ifelse(r$pvalue<.001,"***",ifelse(r$pvalue<.01,"**",ifelse(r$pvalue<.05,"*","")))
  w(paste0("| ",r$lhs," ~~ ",r$rhs," | ",round(r$std.all,3),sig," | ",round(r$pvalue,3)," |"))
}}

w("\\n---\\n## 5. 完整 lavaan 輸出\\n```")
sink_tmp <- tempfile()
sink(sink_tmp)
print(summary(fit_best, fit.measures=TRUE, standardized=TRUE, rsquare=TRUE))
sink()
for (line in readLines(sink_tmp, encoding="UTF-8")) w(line)
w("```")

close(con_out)
cat(sprintf("\\n✅ 報告: %s\\n✅ CLPM log: %s\\n✅ RI-CLPM log: %s\\n", REPORT_PATH, LOG_CLPM, LOG_RI))
cat("=== RICLPM_Master.R 完成 ===\\n")
"""
    return r_script_content
        
    results_md += f"✅ **分析樣本數 (N)**: {len(df_clean)}\n\n"
    
    riclpm_syntax = """
    # Random Intercepts
    RI_CP =~ 1*CP_T1 + 1*CP_T2 + 1*CP_T3
    RI_CI =~ 1*CI_T1 + 1*CI_T2 + 1*CI_T3
    RI_DP =~ 1*DP_T1 + 1*DP_T2 + 1*DP_T3
    
    # Within-person latent variables
    wp_CP_T1 =~ 1*CP_T1
    wp_CP_T2 =~ 1*CP_T2
    wp_CP_T3 =~ 1*CP_T3
    
    wp_CI_T1 =~ 1*CI_T1
    wp_CI_T2 =~ 1*CI_T2
    wp_CI_T3 =~ 1*CI_T3
    
    wp_DP_T1 =~ 1*DP_T1
    wp_DP_T2 =~ 1*DP_T2
    wp_DP_T3 =~ 1*DP_T3
    
    CP_T1 ~~ 0.0*CP_T1
    CP_T2 ~~ 0.0*CP_T2
    CP_T3 ~~ 0.0*CP_T3
    CI_T1 ~~ 0.0*CI_T1
    CI_T2 ~~ 0.0*CI_T2
    CI_T3 ~~ 0.0*CI_T3
    DP_T1 ~~ 0.0*DP_T1
    DP_T2 ~~ 0.0*DP_T2
    DP_T3 ~~ 0.0*DP_T3
    
    # Autoregressive
    wp_CP_T2 ~ wp_CP_T1
    wp_CP_T3 ~ wp_CP_T2
    wp_CI_T2 ~ wp_CI_T1
    wp_CI_T3 ~ wp_CI_T2
    wp_DP_T2 ~ wp_DP_T1
    wp_DP_T3 ~ wp_DP_T2
    
    # Cross-lagged (交叉延遲)
    wp_CI_T2 ~ wp_CP_T1 + wp_DP_T1
    wp_CI_T3 ~ wp_CP_T2 + wp_DP_T2
    
    wp_CP_T2 ~ wp_CI_T1 + wp_DP_T1
    wp_CP_T3 ~ wp_CI_T2 + wp_DP_T2
    
    wp_DP_T2 ~ wp_CP_T1 + wp_CI_T1
    wp_DP_T3 ~ wp_CP_T2 + wp_CI_T2
    
    RI_CP ~~ RI_CI
    RI_CP ~~ RI_DP
    RI_CI ~~ RI_DP
    
    wp_CP_T1 ~~ wp_CI_T1
    wp_CP_T1 ~~ wp_DP_T1
    wp_CI_T1 ~~ wp_DP_T1
    
    wp_CP_T2 ~~ wp_CI_T2
    wp_CP_T2 ~~ wp_DP_T2
    wp_CI_T2 ~~ wp_DP_T2
    
    wp_CP_T3 ~~ wp_CI_T3
    wp_CP_T3 ~~ wp_DP_T3
    wp_CI_T3 ~~ wp_DP_T3
    """
    
    try:
        model = semopy.Model(riclpm_syntax)
        model.fit(df_clean)
        ins = model.inspect()
        
        results_md += "### 交叉延遲效果 (Cross-lagged Effects)\n"
        results_md += "| 依變項 (Outcome) | 預測變項 (Predictor) | 估計值 (Estimate) | p-value | 顯著性 |\n"
        results_md += "|---|---|---|---|---|\n"
        
        cl_paths = ins[(ins['op'] == '~') & (ins['lval'].str.contains('wp_')) & (ins['rval'].str.contains('wp_'))]
        for _, row in cl_paths.iterrows():
            target = row['lval']
            predictor = row['rval']
            if target.split('_')[1] != predictor.split('_')[1]:
                p_val = row.get('p-value', 1.0)
                try: p_val_num = float(p_val)
                except: p_val_num = 1.0
                sig = "***" if p_val_num < 0.001 else ("**" if p_val_num < 0.01 else ("*" if p_val_num < 0.05 else ""))
                if "T1" in predictor and "T2" in target: pass
                elif "T2" in predictor and "T3" in target: pass
                else: continue
                
                results_md += f"| {target} | {predictor} | {row.get('Estimate', 0.0):.3f} | {p_val_num:.4f} | {sig} |\n"
                
    except Exception as e:
        results_md += f"\n⚠️ **RI-CLPM 配適失敗**: {e}\n估計過程可能因樣本少而產生 Singular Matrix，建議使用更精簡模型。\n"
        
    return results_md

def generate_r_script_split_cp(csv_filename):
    output_dir = OUTPUT_DIR.replace('\\\\', '/').replace('\\', '/')
    report_path = f"{output_dir}/RICLPM_SplitCP_Report.md"
    log_ri_path   = f"{output_dir}/model_split_v2_log.txt"
    csv_full_path = f"{output_dir}/{csv_filename}"

    r_script_content = f"""# ==============================================================================
# RICLPM_Split_Master.R (HP ＆ JCP 雙構面分拆版)
# ==============================================================================
library(lavaan)
library(readr)

DATA_PATH   <- "{csv_full_path}"
OUTPUT_DIR  <- "{output_dir}"
REPORT_PATH <- "{report_path}"
LOG_RI  <- "{log_ri_path}"

cat("=== Loading data ===\\n")
df <- read_csv(DATA_PATH, show_col_types = FALSE)
VARS <- c("HP_T1","HP_T2","HP_T3","JCP_T1","JCP_T2","JCP_T3","CI_T1","CI_T2","CI_T3","DP_T1","DP_T2","DP_T3")
df_full <- df[complete.cases(df[, VARS]), ]

cat(sprintf("N total (FIML) = %d\\n", nrow(df)))
cat(sprintf("N 三波完整 (listwise) = %d\\n\\n", nrow(df_full)))

cat("=== RI-CLPM 拆分 HP & JCP 構面模型 ===\\n")
m_ri_compact <- "
  RI_HP =~ 1*HP_T1 + 1*HP_T2 + 1*HP_T3
  RI_JCP =~ 1*JCP_T1 + 1*JCP_T2 + 1*JCP_T3
  RI_CI =~ 1*CI_T1 + 1*CI_T2 + 1*CI_T3
  RI_DP =~ 1*DP_T1 + 1*DP_T2 + 1*DP_T3
  
  wp_HP_T1 =~ 1*HP_T1; wp_HP_T2 =~ 1*HP_T2; wp_HP_T3 =~ 1*HP_T3
  wp_JCP_T1 =~ 1*JCP_T1; wp_JCP_T2 =~ 1*JCP_T2; wp_JCP_T3 =~ 1*JCP_T3
  wp_CI_T1 =~ 1*CI_T1; wp_CI_T2 =~ 1*CI_T2; wp_CI_T3 =~ 1*CI_T3
  wp_DP_T1 =~ 1*DP_T1; wp_DP_T2 =~ 1*DP_T2; wp_DP_T3 =~ 1*DP_T3
  
  HP_T1 ~~ 0*HP_T1; HP_T2 ~~ 0*HP_T2; HP_T3 ~~ 0*HP_T3
  JCP_T1 ~~ 0*JCP_T1; JCP_T2 ~~ 0*JCP_T2; JCP_T3 ~~ 0*JCP_T3
  CI_T1 ~~ 0*CI_T1; CI_T2 ~~ 0*CI_T2; CI_T3 ~~ 0*CI_T3
  DP_T1 ~~ 0*DP_T1; DP_T2 ~~ 0*DP_T2; DP_T3 ~~ 0*DP_T3
  
  wp_HP_T2 ~ a1*wp_HP_T1; wp_HP_T3 ~ a1*wp_HP_T2
  wp_JCP_T2 ~ a2*wp_JCP_T1; wp_JCP_T3 ~ a2*wp_JCP_T2
  wp_CI_T2 ~ a3*wp_CI_T1; wp_CI_T3 ~ a3*wp_CI_T2
  wp_DP_T2 ~ a4*wp_DP_T1; wp_DP_T3 ~ a4*wp_DP_T2

  wp_CI_T2 ~ hp_ci*wp_HP_T1 + jcp_ci*wp_JCP_T1 + dp_ci*wp_DP_T1
  wp_CI_T3 ~ hp_ci*wp_HP_T2 + jcp_ci*wp_JCP_T2 + dp_ci*wp_DP_T2
  
  wp_HP_T2 ~ ci_hp*wp_CI_T1 + jcp_hp*wp_JCP_T1 + dp_hp*wp_DP_T1
  wp_HP_T3 ~ ci_hp*wp_CI_T2 + jcp_hp*wp_JCP_T2 + dp_hp*wp_DP_T2
  
  wp_JCP_T2 ~ ci_jcp*wp_CI_T1 + hp_jcp*wp_HP_T1 + dp_jcp*wp_DP_T1
  wp_JCP_T3 ~ ci_jcp*wp_CI_T2 + hp_jcp*wp_HP_T2 + dp_jcp*wp_DP_T2
  
  wp_DP_T2 ~ hp_dp*wp_HP_T1 + jcp_dp*wp_JCP_T1 + ci_dp*wp_CI_T1
  wp_DP_T3 ~ hp_dp*wp_HP_T2 + jcp_dp*wp_JCP_T2 + ci_dp*wp_CI_T2

  RI_HP ~~ RI_JCP; RI_HP ~~ RI_CI; RI_HP ~~ RI_DP
  RI_JCP ~~ RI_CI; RI_JCP ~~ RI_DP; RI_CI ~~ RI_DP
  
  wp_HP_T1 ~~ wp_JCP_T1; wp_HP_T1 ~~ wp_CI_T1; wp_HP_T1 ~~ wp_DP_T1
  wp_JCP_T1 ~~ wp_CI_T1; wp_JCP_T1 ~~ wp_DP_T1; wp_CI_T1 ~~ wp_DP_T1
  
  wp_HP_T2 ~~ wp_JCP_T2; wp_HP_T2 ~~ wp_CI_T2; wp_HP_T2 ~~ wp_DP_T2
  wp_JCP_T2 ~~ wp_CI_T2; wp_JCP_T2 ~~ wp_DP_T2; wp_CI_T2 ~~ wp_DP_T2
  
  wp_HP_T3 ~~ wp_JCP_T3; wp_HP_T3 ~~ wp_CI_T3; wp_HP_T3 ~~ wp_DP_T3
  wp_JCP_T3 ~~ wp_CI_T3; wp_JCP_T3 ~~ wp_DP_T3; wp_CI_T3 ~~ wp_DP_T3
"

tryCatch({{
  fit_best <- lavaan(m_ri_compact, data=df, missing="fiml", estimator="MLR", bounds=TRUE)
  
  sink(LOG_RI)
  print(summary(fit_best, fit.measures=TRUE, standardized=TRUE))
  sink()
  
  cat(sprintf("✅ 拆分版 RI-CLPM 執行完成，報告已儲存至: %s\\n", LOG_RI))
}}, error=function(e) {{
  cat(sprintf("❌ 模型無法收斂或出現錯誤: %s\\n", e$message))
}})

cat("=== 腳本結束 ===\\n")
"""
    return r_script_content

# ==========================================
# 2.5 MPLUS SYNTAX GENERATOR
# ==========================================
def generate_spss_syntax(csv_filename, ts):
    """產出 SPSS 匯入語法 + 變數標籤 + 數值標籤"""
    return f"""* ============================================================.
* SPSS 匯入與變數標籤語法.
* 資料來源：{csv_filename}.
* 產生時間：{ts}.
* ============================================================.

* ---- 步驟 1：匯入 Analysis_Ready CSV ----.
* 注意：每個變數必須緊接自己的格式碼（SPSS 規定）.
GET DATA
  /TYPE = TXT
  /FILE = "{csv_filename}"
  /ENCODING = 'UTF8'
  /DELIMITERS = ","
  /QUALIFIER = '"'
  /ARRANGEMENT = DELIMITED
  /FIRSTCASE = 2
  /VARIABLES =
    Group        F2.0
    HP1_T1  F8.4   HP2_T1  F8.4   HP3_T1  F8.4
    HP4_T1  F8.4   HP5_T1  F8.4   HP6_T1  F8.4
    JCP1_T1 F8.4   JCP2_T1 F8.4   JCP3_T1 F8.4
    JCP4_T1 F8.4   JCP5_T1 F8.4   JCP6_T1 F8.4
    PP1_T1  F8.4   PP2_T1  F8.4   PP3_T1  F8.4
    PP4_T1  F8.4   PP5_T1  F8.4   PP6_T1  F8.4
    DP1_T1  F8.4   DP2_T1  F8.4   DP3_T1  F8.4
    DP4_T1  F8.4   DP5_T1  F8.4
    CI1_T1  F8.4   CI2_T1  F8.4   CI3_T1  F8.4
    CI4_T1  F8.4   CI5_T1  F8.4   CI6_T1  F8.4
    CI7_T1  F8.4   CI8_T1  F8.4
    Gender       F2.0
    Age          F6.2
    Education    F2.0
    Marriage     F2.0
    NowJobTenure F8.2
    JobTenure    F8.2
    Position     F2.0
    Industry     F2.0
    OrgSize      F2.0
    PM_Has_T1      F2.0
    PM_Supervisor_T1 F2.0   PM_Self_T1 F2.0
    PM_Interview_T1  F2.0   PM_Other_T1 F2.0
    PM_Result_T1   F2.0
    PM_Help_T1     F6.2
    HP1_T2  F8.4   HP2_T2  F8.4   HP3_T2  F8.4
    HP4_T2  F8.4   HP5_T2  F8.4   HP6_T2  F8.4
    JCP1_T2 F8.4   JCP2_T2 F8.4   JCP3_T2 F8.4
    JCP4_T2 F8.4   JCP5_T2 F8.4   JCP6_T2 F8.4
    PP1_T2  F8.4   PP2_T2  F8.4   PP3_T2  F8.4
    PP4_T2  F8.4   PP5_T2  F8.4   PP6_T2  F8.4
    DP1_T2  F8.4   DP2_T2  F8.4   DP3_T2  F8.4
    DP4_T2  F8.4   DP5_T2  F8.4
    CI1_T2  F8.4   CI2_T2  F8.4   CI3_T2  F8.4
    CI4_T2  F8.4   CI5_T2  F8.4   CI6_T2  F8.4
    CI7_T2  F8.4   CI8_T2  F8.4
    PM_Has_T2      F2.0
    PM_Supervisor_T2 F2.0   PM_Self_T2 F2.0
    PM_Interview_T2  F2.0   PM_Other_T2 F2.0
    PM_Result_T2   F2.0
    PM_Help_T2     F6.2
    HP1_T3  F8.4   HP2_T3  F8.4   HP3_T3  F8.4
    HP4_T3  F8.4   HP5_T3  F8.4   HP6_T3  F8.4
    JCP1_T3 F8.4   JCP2_T3 F8.4   JCP3_T3 F8.4
    JCP4_T3 F8.4   JCP5_T3 F8.4   JCP6_T3 F8.4
    PP1_T3  F8.4   PP2_T3  F8.4   PP3_T3  F8.4
    PP4_T3  F8.4   PP5_T3  F8.4   PP6_T3  F8.4
    DP1_T3  F8.4   DP2_T3  F8.4   DP3_T3  F8.4
    DP4_T3  F8.4   DP5_T3  F8.4
    CI1_T3  F8.4   CI2_T3  F8.4   CI3_T3  F8.4
    CI4_T3  F8.4   CI5_T3  F8.4   CI6_T3  F8.4
    CI7_T3  F8.4   CI8_T3  F8.4
    PM_Has_T3      F2.0
    PM_Supervisor_T3 F2.0   PM_Self_T3 F2.0
    PM_Interview_T3  F2.0   PM_Other_T3 F2.0
    PM_Result_T3   F2.0
    PM_Help_T3     F6.2.
CACHE.
EXECUTE.

* ---- 步驟 2：變數標籤（Variable Labels）----.
VARIABLE LABELS
  Group       '波次群組 (1=僅T1, 2=T1+T2, 3=三波完整)'
  HP1_T1      'HP題1_T1：晉升可能性有限'
  HP2_T1      'HP題2_T1：職位達頂'
  HP3_T1      'HP題3_T1：不太可能獲更高職位'
  HP4_T1      'HP題4_T1(反)：預期近期可晉升'
  HP5_T1      'HP題5_T1：向上晉升機會有限'
  HP6_T1      'HP題6_T1(反)：預期晉升機會頻繁'
  JCP1_T1     'JCP題1_T1(反)：工作有挑戰性'
  JCP2_T1     'JCP題2_T1(反)：需拓展能力'
  JCP3_T1     'JCP題3_T1(反)：學習成長機會多'
  JCP4_T1     'JCP題4_T1(反)：工作經常有挑戰'
  JCP5_T1     'JCP題5_T1(反)：職責明顯增加'
  JCP6_T1     'JCP題6_T1：工作已成家常便飯'
  PP1_T1      'PP題1_T1：改正看不順眼的事'
  PP2_T1      'PP題2_T1：努力實現所信之事'
  PP3_T1      'PP題3_T1：擁護自己的想法'
  PP4_T1      'PP題4_T1：尋求更好做事方法'
  PP5_T1      'PP題5_T1：落實所信的理念'
  PP6_T1      'PP題6_T1：洞察先機'
  DP1_T1      'DP題1_T1：決定前花時間處理瑣事'
  DP2_T1      'DP題2_T1：決定後仍拖延行動'
  DP3_T1      'DP題3_T1：等很久才思考決定'
  DP4_T1      'DP題4_T1：拖延到為時已晚'
  DP5_T1      'DP題5_T1：拖延做決定'
  CI1_T1      'CI題1_T1：想改變職涯但未積極追求'
  CI2_T1      'CI題2_T1：不知如何開始'
  CI3_T1      'CI題3_T1：不敢放棄現有'
  CI4_T1      'CI題4_T1：未採具體行動'
  CI5_T1      'CI題5_T1：感到無法動彈'
  CI6_T1      'CI題6_T1：行動困難'
  CI7_T1      'CI題7_T1：職涯停滯感'
  CI8_T1      'CI題8_T1：無法實現職涯渴望'
  HP1_T2      'HP題1_T2：晉升可能性有限'
  HP2_T2      'HP題2_T2：職位達頂'
  HP3_T2      'HP題3_T2：不太可能獲更高職位'
  HP4_T2      'HP題4_T2(反)：預期近期可晉升'
  HP5_T2      'HP題5_T2：向上晉升機會有限'
  HP6_T2      'HP題6_T2(反)：預期晉升機會頻繁'
  JCP1_T2     'JCP題1_T2(反)：工作有挑戰性'
  JCP2_T2     'JCP題2_T2(反)：需拓展能力'
  JCP3_T2     'JCP題3_T2(反)：學習成長機會多'
  JCP4_T2     'JCP題4_T2(反)：工作經常有挑戰'
  JCP5_T2     'JCP題5_T2(反)：職責明顯增加'
  JCP6_T2     'JCP題6_T2：工作已成家常便飯'
  PP1_T2      'PP題1_T2：改正看不順眼的事'
  PP2_T2      'PP題2_T2：努力實現所信之事'
  PP3_T2      'PP題3_T2：擁護自己的想法'
  PP4_T2      'PP題4_T2：尋求更好做事方法'
  PP5_T2      'PP題5_T2：落實所信的理念'
  PP6_T2      'PP題6_T2：洞察先機'
  DP1_T2      'DP題1_T2：決定前花時間處理瑣事'
  DP2_T2      'DP題2_T2：決定後仍拖延行動'
  DP3_T2      'DP題3_T2：等很久才思考決定'
  DP4_T2      'DP題4_T2：拖延到為時已晚'
  DP5_T2      'DP題5_T2：拖延做決定'
  CI1_T2      'CI題1_T2：想改變職涯但未積極追求'
  CI2_T2      'CI題2_T2：不知如何開始'
  CI3_T2      'CI題3_T2：不敢放棄現有'
  CI4_T2      'CI題4_T2：未採具體行動'
  CI5_T2      'CI題5_T2：感到無法動彈'
  CI6_T2      'CI題6_T2：行動困難'
  CI7_T2      'CI題7_T2：職涯停滯感'
  CI8_T2      'CI題8_T2：無法實現職涯渴望'
  HP1_T3      'HP題1_T3：晉升可能性有限'
  HP2_T3      'HP題2_T3：職位達頂'
  HP3_T3      'HP題3_T3：不太可能獲更高職位'
  HP4_T3      'HP題4_T3(反)：預期近期可晉升'
  HP5_T3      'HP題5_T3：向上晉升機會有限'
  HP6_T3      'HP題6_T3(反)：預期晉升機會頻繁'
  JCP1_T3     'JCP題1_T3(反)：工作有挑戰性'
  JCP2_T3     'JCP題2_T3(反)：需拓展能力'
  JCP3_T3     'JCP題3_T3(反)：學習成長機會多'
  JCP4_T3     'JCP題4_T3(反)：工作經常有挑戰'
  JCP5_T3     'JCP題5_T3(反)：職責明顯增加'
  JCP6_T3     'JCP題6_T3：工作已成家常便飯'
  PP1_T3      'PP題1_T3：改正看不順眼的事'
  PP2_T3      'PP題2_T3：努力實現所信之事'
  PP3_T3      'PP題3_T3：擁護自己的想法'
  PP4_T3      'PP題4_T3：尋求更好做事方法'
  PP5_T3      'PP題5_T3：落實所信的理念'
  PP6_T3      'PP題6_T3：洞察先機'
  DP1_T3      'DP題1_T3：決定前花時間處理瑣事'
  DP2_T3      'DP題2_T3：決定後仍拖延行動'
  DP3_T3      'DP題3_T3：等很久才思考決定'
  DP4_T3      'DP題4_T3：拖延到為時已晚'
  DP5_T3      'DP題5_T3：拖延做決定'
  CI1_T3      'CI題1_T3：想改變職涯但未積極追求'
  CI2_T3      'CI題2_T3：不知如何開始'
  CI3_T3      'CI題3_T3：不敢放棄現有'
  CI4_T3      'CI題4_T3：未採具體行動'
  CI5_T3      'CI題5_T3：感到無法動彈'
  CI6_T3      'CI題6_T3：行動困難'
  CI7_T3      'CI題7_T3：職涯停滯感'
  CI8_T3      'CI題8_T3：無法實現職涯渴望'
  Gender      '性別'
  Age         '年齡（實歲）'
  Education   '教育程度'
  Marriage    '婚姻狀況'
  NowJobTenure '現職年資（月）'
  JobTenure   '工作總年資（月）'
  Position    '工作職級'
  Industry    '產業別'
  OrgSize     '公司規模'
  PM_Has_T1      'T1 是否有績效考核'
  PM_Supervisor_T1 'T1 考核形式：主管評核'
  PM_Self_T1     'T1 考核形式：自我評核'
  PM_Interview_T1 'T1 考核形式：績效面談'
  PM_Other_T1    'T1 考核形式：其他'
  PM_Result_T1   'T1 考核結果性質'
  PM_Help_T1     'T1 考核對職涯幫助程度（1-5）'
  PM_Has_T2      'T2 是否有績效考核'
  PM_Supervisor_T2 'T2 考核形式：主管評核'
  PM_Self_T2     'T2 考核形式：自我評核'
  PM_Interview_T2 'T2 考核形式：績效面談'
  PM_Other_T2    'T2 考核形式：其他'
  PM_Result_T2   'T2 考核結果性質'
  PM_Help_T2     'T2 考核對職涯幫助程度（1-5）'
  PM_Has_T3      'T3 是否有績效考核'
  PM_Supervisor_T3 'T3 考核形式：主管評核'
  PM_Self_T3     'T3 考核形式：自我評核'
  PM_Interview_T3 'T3 考核形式：績效面談'
  PM_Other_T3    'T3 考核形式：其他'
  PM_Result_T3   'T3 考核結果性質'
  PM_Help_T3     'T3 考核對職涯幫助程度（1-5）'.

* ---- 步驟 3：數值標籤（Value Labels）----.
VALUE LABELS
  Group
    1 '僅完成T1'
    2 '完成T1+T2'
    3 '三波完整' /
  Gender
    1 '男'
    2 '女'
    3 '其他' /
  Education
    1 '高中職'
    2 '專科'
    3 '大學'
    4 '碩士'
    5 '博士' /
  Marriage
    1 '未婚'
    2 '已婚無子女'
    3 '已婚有子女'
    4 '其他' /
  Position
    1 '一般員工'
    2 '基層主管'
    3 '中階主管'
    4 '高階主管' /
  Industry
    1 '製造業'
    2 '科技/資訊業'
    3 '金融/保險業'
    4 '服務業'
    5 '醫療/健康業'
    6 '教育業'
    7 '公部門'
    8 '其他' /
  OrgSize
    1 '30人以下'
    2 '31-100人'
    3 '101-500人'
    4 '501-1000人'
    5 '1001人以上' /
  PM_Has_T1 PM_Has_T2 PM_Has_T3
    0 '無績效考核'
    1 '有績效考核' /
  PM_Supervisor_T1 PM_Self_T1 PM_Interview_T1 PM_Other_T1
  PM_Supervisor_T2 PM_Self_T2 PM_Interview_T2 PM_Other_T2
  PM_Supervisor_T3 PM_Self_T3 PM_Interview_T3 PM_Other_T3
    0 '否'
    1 '是' /
  PM_Result_T1 PM_Result_T2 PM_Result_T3
    1 '負向回饋'
    2 '中性/持平'
    3 '正向回饋'.

* ---- 步驟 4：遺漏值設定 ----.
* 注意：SYSMIS（系統遺漏值）由 SPSS 自動處理，無需額外宣告。
* 若問卷使用特定代碼代表遺漏（如 99、-9），可在此宣告，例如：
* MISSING VALUES HP1_T1 TO CI8_T3 (99).
* 本資料已於 Python 清理階段處理遺漏值，無需另行設定。

* ---- 步驟 5：只保留三波完整樣本（Group=3）----.
* SELECT IF (Group = 3).
* EXECUTE.
* 如需全樣本分析，請移除上兩行前面的星號（*）.

* ---- 步驟 6：描述統計 ----.
DESCRIPTIVES VARIABLES = HP1_T1 TO CI8_T1
  /STATISTICS = MEAN STDDEV MIN MAX.
"""


def _mplus_header(dat_filename):
    return f"""ANALYSIS:
  ESTIMATOR = MLR;
  ITERATIONS = 10000;
  CONVERGENCE = 0.000001;
"""

def _mplus_cp_measurement():
    """CP 測量模型（HP+JCP 跨波等同）"""
    return """  ! CP: HP + JCP 跨波負荷等同 + 截距等同
  CP1 BY HP_T1  (l_hp)
          JCP_T1 (l_jcp);
  CP2 BY HP_T2  (l_hp)
          JCP_T2 (l_jcp);
  CP3 BY HP_T3  (l_hp)
          JCP_T3 (l_jcp);
  [HP_T1]  (int_hp);  [HP_T2]  (int_hp);  [HP_T3]  (int_hp);
  [JCP_T1] (int_jcp); [JCP_T2] (int_jcp); [JCP_T3] (int_jcp);
  CP1@0;  CP2@0;  CP3@0;
"""

def _mplus_single_indicator(var):
    """單一指標構念（負荷=1，殘差=0）"""
    v = var.upper()
    return (
        f"  {v}1 BY {v}_T1@1;  {v}2 BY {v}_T2@1;  {v}3 BY {v}_T3@1;\n"
        f"  {v}_T1@0;  {v}_T2@0;  {v}_T3@0;\n"
        f"  {v}1@0;  {v}2@0;  {v}3@0;\n"
    )

def _mplus_ri_clpm_core(vars_list):
    """隨機截距 + Within-person 殘差 + 自回歸（通用）"""
    ri_lines = "  ! 隨機截距\n"
    w_lines  = "  ! Within-person 殘差\n"
    ar_lines = "  ! 自回歸（跨波等同）\n"
    t1_cov   = "  ! T1 Within-person 共變\n"
    t2_cov   = "  ! T2 殘差共變\n"
    t3_cov   = "  ! T3 殘差共變\n"

    for v in vars_list:
        ri_lines += f"  RI_{v} BY {v}1@1 {v}2@1 {v}3@1;\n"
        w_lines  += f"  W{v}1 BY {v}1@1;  W{v}2 BY {v}2@1;  W{v}3 BY {v}3@1;\n"
        ar_lines += f"  W{v}2 ON W{v}1 (ar_{v.lower()});  W{v}3 ON W{v}2 (ar_{v.lower()});\n"

    w_vars = [f"W{v}1" for v in vars_list]
    t1_cov += "  " + " WITH ".join(w_vars) + ";\n" if len(w_vars) > 1 else ""
    for i in range(len(vars_list)):
        rest = [f"W{v}2" for v in vars_list[i+1:]]
        if rest:
            t2_cov += f"  W{vars_list[i]}2 WITH {' '.join(rest)};\n"
        rest3 = [f"W{v}3" for v in vars_list[i+1:]]
        if rest3:
            t3_cov += f"  W{vars_list[i]}3 WITH {' '.join(rest3)};\n"

    ri_cov = "  ! 隨機截距共變\n"
    for i in range(len(vars_list)):
        rest = [f"RI_{v}" for v in vars_list[i+1:]]
        if rest:
            ri_cov += f"  RI_{vars_list[i]} WITH {' '.join(rest)};\n"

    return ri_lines + w_lines + ar_lines + t1_cov + t2_cov + t3_cov + ri_cov


# ----------------------------------------------------------
# Step 1：主路徑 CP → DP → CI（確認信效度用）
# ----------------------------------------------------------
def generate_mplus_step1(dat_filename, ts):
    return f"""TITLE:
  [Step1] RI-CLPM 主路徑: CP -> DP -> CI
  Generated: {ts}
  用途: 確認主路徑信效度，不含 PP 與控制變數

DATA:
  FILE = "{dat_filename}";

VARIABLE:
  NAMES =
    HP_T1  JCP_T1  PP_T1  DP_T1  CI_T1
    HP_T2  JCP_T2  PP_T2  DP_T2  CI_T2
    HP_T3  JCP_T3  PP_T3  DP_T3  CI_T3;
  USEVARIABLES =
    HP_T1  JCP_T1  DP_T1  CI_T1
    HP_T2  JCP_T2  DP_T2  CI_T2
    HP_T3  JCP_T3  DP_T3  CI_T3;
  MISSING = ALL(-999);

{_mplus_header(dat_filename)}
MODEL:

{_mplus_cp_measurement()}
{_mplus_single_indicator('DP')}
{_mplus_single_indicator('CI')}
{_mplus_ri_clpm_core(['CP','DP','CI'])}
  ! 交叉延遲：CP -> DP [正向]
  WDP2 ON WCP1 (cl_cp_dp);  WDP3 ON WCP2 (cl_cp_dp);
  ! 交叉延遲：DP -> CI [正向]
  WCI2 ON WDP1 (cl_dp_ci);  WCI3 ON WDP2 (cl_dp_ci);

OUTPUT:
  SAMPSTAT;  STDYX;  MODINDICES(10);  CINTERVAL;
"""


# ----------------------------------------------------------
# Step 2：加入 PP（H8：PP 隨機截距與 DP/CI 隨機截距的個體間關聯）
# 修正說明：PP 為跨期穩定特質，within-person 變異接近 0，
# 不適合當 within-person 交叉延遲的前因。正確做法：
# (1) RI_PP WITH RI_DP/RI_CI（已由 _mplus_ri_clpm_core 自動產生）
# (2) H8 調節效果採高/低 PP 分組模型（C1/C2/D1/D2）
# ----------------------------------------------------------
def generate_mplus_step2(dat_filename, ts):
    return f"""TITLE:
  [Step2] RI-CLPM 加入 PP（PP 隨機截距共變）
  Generated: {ts}
  H8: PP 穩定水準（RI_PP）與 DP/CI 穩定水準的個體間關聯
  注意：PP 為跨期穩定特質，使用 RI 共變而非 within-person 路徑

DATA:
  FILE = "{dat_filename}";

VARIABLE:
  NAMES =
    HP_T1  JCP_T1  PP_T1  DP_T1  CI_T1
    HP_T2  JCP_T2  PP_T2  DP_T2  CI_T2
    HP_T3  JCP_T3  PP_T3  DP_T3  CI_T3;
  USEVARIABLES = ALL;
  MISSING = ALL(-999);

{_mplus_header(dat_filename)}
MODEL:

{_mplus_cp_measurement()}
{_mplus_single_indicator('PP')}
{_mplus_single_indicator('DP')}
{_mplus_single_indicator('CI')}
{_mplus_ri_clpm_core(['CP','PP','DP','CI'])}
  ! 交叉延遲：CP -> DP [正向]
  WDP2 ON WCP1 (cl_cp_dp);  WDP3 ON WCP2 (cl_cp_dp);
  ! 交叉延遲：DP -> CI [正向]
  WCI2 ON WDP1 (cl_dp_ci);  WCI3 ON WDP2 (cl_dp_ci);
  ! H8（between-person）：RI_PP 與 RI_DP/RI_CI 隨機截距共變
  ! 已在 _mplus_ri_clpm_core 產出的 RI WITH 區段中包含
  ! 可從 OUTPUT STDYX 的 Correlations Among Latent Variables 讀取

OUTPUT:
  SAMPSTAT;  STDYX;  MODINDICES(10);  CINTERVAL;
"""


# ----------------------------------------------------------
# Step 3：加入控制變數（性別、年齡、年資、職級）
# ----------------------------------------------------------
def generate_mplus_step3(dat_filename, ts):
    return f"""TITLE:
  [Step3] RI-CLPM 完整版: CP/PP/DP/CI + 控制變數
  Generated: {ts}
  控制變數回歸至隨機截距：Gender, Age, Tenure, Position
  H8 調節效果見分組模型（C1/C2 JCP 組、D1/D2 HP 組）

DATA:
  FILE = "{dat_filename}";

VARIABLE:
  NAMES =
    HP_T1  JCP_T1  PP_T1  DP_T1  CI_T1
    HP_T2  JCP_T2  PP_T2  DP_T2  CI_T2
    HP_T3  JCP_T3  PP_T3  DP_T3  CI_T3
    Gender Tenure Position Age;
  USEVARIABLES = ALL;
  MISSING = ALL(-999);

{_mplus_header(dat_filename)}
MODEL:

{_mplus_cp_measurement()}
{_mplus_single_indicator('PP')}
{_mplus_single_indicator('DP')}
{_mplus_single_indicator('CI')}
{_mplus_ri_clpm_core(['CP','PP','DP','CI'])}
  ! 交叉延遲：CP -> DP
  WDP2 ON WCP1 (cl_cp_dp);  WDP3 ON WCP2 (cl_cp_dp);
  ! 交叉延遲：DP -> CI
  WCI2 ON WDP1 (cl_dp_ci);  WCI3 ON WDP2 (cl_dp_ci);
  ! H8（between-person）：RI_PP 與 RI_DP/RI_CI 隨機截距共變
  ! 已在 _mplus_ri_clpm_core 產出的 RI WITH 區段中包含

  ! 控制變數 -> 隨機截距（排除個體間差異）
  RI_CP ON Gender Age Tenure Position;
  RI_PP ON Gender Age Tenure Position;
  RI_DP ON Gender Age Tenure Position;
  RI_CI ON Gender Age Tenure Position;

OUTPUT:
  SAMPSTAT;  STDYX;  MODINDICES(10);  CINTERVAL;
"""


# ==========================================
# 信度分析 SPSS 語法
# ==========================================
def generate_spss_reliability_syntax(analysis_path, ts):
    return f"""\
* ============================================================.
* 信度分析語法（Cronbach's Alpha）— 自動產生 {ts}.
* 使用方式：先用 SPSS_Syntax_{ts}.sps 匯入資料，再執行本語法.
* 資料來源：{analysis_path}.
* ============================================================.

* === HP 階層停滯（T1，6 題，HP4/HP6 已反向計分）===.
RELIABILITY
  /VARIABLES = HP1_T1 HP2_T1 HP3_T1 HP4_T1 HP5_T1 HP6_T1
  /SCALE('HP 階層停滯 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* === JCP 工作內容停滯（T1，6 題，JCP1~5 已反向計分）===.
RELIABILITY
  /VARIABLES = JCP1_T1 JCP2_T1 JCP3_T1 JCP4_T1 JCP5_T1 JCP6_T1
  /SCALE('JCP 工作內容停滯 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* === CP 職涯高原合併（T1，HP+JCP 共 12 題）===.
RELIABILITY
  /VARIABLES = HP1_T1 HP2_T1 HP3_T1 HP4_T1 HP5_T1 HP6_T1
               JCP1_T1 JCP2_T1 JCP3_T1 JCP4_T1 JCP5_T1 JCP6_T1
  /SCALE('CP 職涯高原合併 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* === PP 主動型人格（T1，6 題）===.
RELIABILITY
  /VARIABLES = PP1_T1 PP2_T1 PP3_T1 PP4_T1 PP5_T1 PP6_T1
  /SCALE('PP 主動型人格 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* === DP 決策拖延（T1，5 題）===.
RELIABILITY
  /VARIABLES = DP1_T1 DP2_T1 DP3_T1 DP4_T1 DP5_T1
  /SCALE('DP 決策拖延 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* === CI 職涯無所作為（T1，8 題）===.
RELIABILITY
  /VARIABLES = CI1_T1 CI2_T1 CI3_T1 CI4_T1 CI5_T1 CI6_T1 CI7_T1 CI8_T1
  /SCALE('CI 職涯無所作為 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* ============================================================.
* 各波信度（T2 / T3）——若需跨波比較請執行以下語法.
* ============================================================.
* HP T2.
RELIABILITY
  /VARIABLES = HP1_T2 HP2_T2 HP3_T2 HP4_T2 HP5_T2 HP6_T2
  /SCALE('HP T2') ALL /MODEL = ALPHA /SUMMARY = TOTAL.
RELIABILITY
  /VARIABLES = HP1_T3 HP2_T3 HP3_T3 HP4_T3 HP5_T3 HP6_T3
  /SCALE('HP T3') ALL /MODEL = ALPHA /SUMMARY = TOTAL.

RELIABILITY
  /VARIABLES = DP1_T2 DP2_T2 DP3_T2 DP4_T2 DP5_T2
  /SCALE('DP T2') ALL /MODEL = ALPHA /SUMMARY = TOTAL.
RELIABILITY
  /VARIABLES = DP1_T3 DP2_T3 DP3_T3 DP4_T3 DP5_T3
  /SCALE('DP T3') ALL /MODEL = ALPHA /SUMMARY = TOTAL.

RELIABILITY
  /VARIABLES = CI1_T2 CI2_T2 CI3_T2 CI4_T2 CI5_T2 CI6_T2 CI7_T2 CI8_T2
  /SCALE('CI T2') ALL /MODEL = ALPHA /SUMMARY = TOTAL.
RELIABILITY
  /VARIABLES = CI1_T3 CI2_T3 CI3_T3 CI4_T3 CI5_T3 CI6_T3 CI7_T3 CI8_T3
  /SCALE('CI T3') ALL /MODEL = ALPHA /SUMMARY = TOTAL.
"""


def generate_spss_analysis_syntax(analysis_path, ts, pp_median=3.5, n_total=340, exclude=None):
    """
    產出完整 SPSS 分析語法：
      步驟 0：篩選三波完整樣本 (Group=3)
      步驟 1：計算合成分數（各量表 T1/T2/T3 平均）
      步驟 2：PP 中位數分群（PP_group：0=低PP, 1=高PP）
      步驟 3：人口統計頻率分佈（FREQUENCIES）
      步驟 4：量表描述統計（DESCRIPTIVES，各波合成分數）
      步驟 5：Pearson 相關矩陣（T1 合成分數）
      步驟 6：Harman 共同方法偏差單因子檢定（CMV）
      步驟 7：常態性檢定（EXAMINE）
    exclude: list of item names to omit from composites and reliability (e.g. ['JCP6', 'DP1'])
    """
    excl = set(exclude or [])

    def _m(prefix, total, wave):
        items = [f'{prefix}{i}_{wave}' for i in range(1, total+1) if f'{prefix}{i}' not in excl]
        return f"MEAN({', '.join(items)})"

    def _r(prefix, total, wave):
        items = [f'{prefix}{i}_{wave}' for i in range(1, total+1) if f'{prefix}{i}' not in excl]
        return ' '.join(items)

    jcp_m1=_m('JCP',6,'T1'); jcp_m2=_m('JCP',6,'T2'); jcp_m3=_m('JCP',6,'T3')
    dp_m1 =_m('DP', 5,'T1'); dp_m2 =_m('DP', 5,'T2'); dp_m3 =_m('DP', 5,'T3')
    jcp_r1=_r('JCP',6,'T1'); jcp_r2=_r('JCP',6,'T2'); jcp_r3=_r('JCP',6,'T3')
    dp_r1 =_r('DP', 5,'T1'); dp_r2 =_r('DP', 5,'T2'); dp_r3 =_r('DP', 5,'T3')
    excl_note = f'（排除題目：{", ".join(sorted(excl))}）' if excl else ''

    return f"""\
* ============================================================.
* 完整 SPSS 分析語法 — 自動產生 {ts}{excl_note}.
* 執行前請先開啟並執行 SPSS_Syntax_{ts}.sps（匯入資料）.
* 資料來源：{analysis_path}.
* N（三波完整）= {n_total}.
* ============================================================.

* ============================================================.
* 步驟 0：篩選三波完整樣本（N = {n_total}）.
* ============================================================.
SELECT IF (Group = 3).
EXECUTE.


* ============================================================.
* 步驟 1：計算合成分數（各量表各波次題目平均值）.
* 注意：反向題已於資料清理時處理，無需再反向.
* ============================================================.

* --- T1 合成分數 ---.
COMPUTE HP_T1  = MEAN(HP1_T1,  HP2_T1,  HP3_T1,  HP4_T1,  HP5_T1,  HP6_T1).
COMPUTE JCP_T1 = {jcp_m1}.
COMPUTE PP_T1  = MEAN(PP1_T1,  PP2_T1,  PP3_T1,  PP4_T1,  PP5_T1,  PP6_T1).
COMPUTE DP_T1  = {dp_m1}.
COMPUTE CI_T1  = MEAN(CI1_T1,  CI2_T1,  CI3_T1,  CI4_T1,  CI5_T1,
                       CI6_T1,  CI7_T1,  CI8_T1).
VARIABLE LABELS
  HP_T1  'HP 階層停滯 T1（合成分數）'
  JCP_T1 'JCP 工作內容停滯 T1（合成分數）'
  PP_T1  'PP 主動型人格 T1（合成分數）'
  DP_T1  'DP 決策拖延 T1（合成分數）'
  CI_T1  'CI 職涯無所作為 T1（合成分數）'.

* --- T2 合成分數 ---.
COMPUTE HP_T2  = MEAN(HP1_T2,  HP2_T2,  HP3_T2,  HP4_T2,  HP5_T2,  HP6_T2).
COMPUTE JCP_T2 = {jcp_m2}.
COMPUTE PP_T2  = MEAN(PP1_T2,  PP2_T2,  PP3_T2,  PP4_T2,  PP5_T2,  PP6_T2).
COMPUTE DP_T2  = {dp_m2}.
COMPUTE CI_T2  = MEAN(CI1_T2,  CI2_T2,  CI3_T2,  CI4_T2,  CI5_T2,
                       CI6_T2,  CI7_T2,  CI8_T2).
VARIABLE LABELS
  HP_T2  'HP 階層停滯 T2（合成分數）'
  JCP_T2 'JCP 工作內容停滯 T2（合成分數）'
  PP_T2  'PP 主動型人格 T2（合成分數）'
  DP_T2  'DP 決策拖延 T2（合成分數）'
  CI_T2  'CI 職涯無所作為 T2（合成分數）'.

* --- T3 合成分數 ---.
COMPUTE HP_T3  = MEAN(HP1_T3,  HP2_T3,  HP3_T3,  HP4_T3,  HP5_T3,  HP6_T3).
COMPUTE JCP_T3 = {jcp_m3}.
COMPUTE PP_T3  = MEAN(PP1_T3,  PP2_T3,  PP3_T3,  PP4_T3,  PP5_T3,  PP6_T3).
COMPUTE DP_T3  = {dp_m3}.
COMPUTE CI_T3  = MEAN(CI1_T3,  CI2_T3,  CI3_T3,  CI4_T3,  CI5_T3,
                       CI6_T3,  CI7_T3,  CI8_T3).
VARIABLE LABELS
  HP_T3  'HP 階層停滯 T3（合成分數）'
  JCP_T3 'JCP 工作內容停滯 T3（合成分數）'
  PP_T3  'PP 主動型人格 T3（合成分數）'
  DP_T3  'DP 決策拖延 T3（合成分數）'
  CI_T3  'CI 職涯無所作為 T3（合成分數）'.

EXECUTE.


* ============================================================.
* 步驟 2：PP 中位數分群（H8 調節效果子群）.
* 中位數 = {pp_median}（由三波完整樣本 T1 PP 中位數計算）.
* PP_group：0 = 低PP（≤中位數），1 = 高PP（>中位數）.
* ============================================================.
COMPUTE PP_group = 0.
IF (PP_T1 > {pp_median}) PP_group = 1.
EXECUTE.
VARIABLE LABELS PP_group 'PP 中位數分群（0=低PP, 1=高PP）'.
VALUE LABELS PP_group
  0 '低PP（≤{pp_median}）'
  1 '高PP（>{pp_median}）'.
FREQUENCIES VARIABLES = PP_group.


* ============================================================.
* 步驟 3：人口統計背景變項頻率分佈.
* ============================================================.
FREQUENCIES VARIABLES = Gender Education Marriage Position Industry OrgSize
  /STATISTICS = MINIMUM MAXIMUM
  /ORDER = ANALYSIS.

* 年齡描述統計（連續變數）.
DESCRIPTIVES VARIABLES = Age NowJobTenure JobTenure
  /STATISTICS = MEAN STDDEV MIN MAX.

* 績效考核現況（三波次）.
FREQUENCIES VARIABLES = PM_Has_T1 PM_Result_T1 PM_Help_T1
                        PM_Has_T2 PM_Result_T2 PM_Help_T2
                        PM_Has_T3 PM_Result_T3 PM_Help_T3.


* ============================================================.
* 步驟 4：各量表合成分數描述統計（三波次）.
* 輸出：平均數、標準差、偏態、峰態.
* ============================================================.
DESCRIPTIVES
  VARIABLES = HP_T1  JCP_T1  PP_T1  DP_T1  CI_T1
              HP_T2  JCP_T2  PP_T2  DP_T2  CI_T2
              HP_T3  JCP_T3  PP_T3  DP_T3  CI_T3
  /STATISTICS = MEAN STDDEV MIN MAX SKEWNESS KURTOSIS.


* ============================================================.
* 步驟 5：Pearson 相關矩陣（T1 合成分數）.
* 論文表格：Table 3.
* ============================================================.
CORRELATIONS
  /VARIABLES = HP_T1 JCP_T1 PP_T1 DP_T1 CI_T1
  /PRINT = TWOTAIL SIG
  /MISSING = PAIRWISE.

* 跨波相關（三波合成分數完整矩陣）.
CORRELATIONS
  /VARIABLES = HP_T1 JCP_T1 PP_T1 DP_T1 CI_T1
               HP_T2 JCP_T2 PP_T2 DP_T2 CI_T2
               HP_T3 JCP_T3 PP_T3 DP_T3 CI_T3
  /PRINT = TWOTAIL SIG
  /MISSING = PAIRWISE.


* ============================================================.
* 步驟 6：Harman 共同方法偏差（CMV）單因子檢定.
* 方法：對所有 T1 題目做 EFA，強制單因子，.
*   看第一個因子解釋的變異量是否 < 50%.
* 建議報告：「強制單因子所解釋的總變異量為 XX%，.
*            低於 50% 標準，共同方法偏差問題不嚴重。」.
* ============================================================.
FACTOR
  /VARIABLES = HP1_T1 HP2_T1 HP3_T1 HP4_T1 HP5_T1 HP6_T1
               JCP1_T1 JCP2_T1 JCP3_T1 JCP4_T1 JCP5_T1 JCP6_T1
               PP1_T1 PP2_T1 PP3_T1 PP4_T1 PP5_T1 PP6_T1
               DP1_T1 DP2_T1 DP3_T1 DP4_T1 DP5_T1
               CI1_T1 CI2_T1 CI3_T1 CI4_T1 CI5_T1 CI6_T1 CI7_T1 CI8_T1
  /MISSING = LISTWISE
  /ANALYSIS = HP1_T1 TO CI8_T1
  /PRINT = INITIAL KMO EXTRACTION
  /CRITERIA = FACTORS(1)
  /EXTRACTION = PC
  /ROTATION = NOROTATE
  /METHOD = CORRELATION.
* 重點看：「Total Variance Explained」表格第一行的 % of Variance 欄位.
* 若 < 50%，CMV 不嚴重（Harman's Single Factor Test 通過）.


* ============================================================.
* 步驟 7：常態性檢定（Shapiro-Wilk / K-S）.
* 注意：SPSS EXAMINE 在 N>50 時自動只顯示 K-S，N≤50 才顯示 Shapiro-Wilk.
* 建議：若偏態 < |2| 且峰態 < |7| 即符合近似常態（步驟4已可看）.
* ============================================================.
EXAMINE VARIABLES = HP_T1 JCP_T1 PP_T1 DP_T1 CI_T1
  /PLOT NONE
  /STATISTICS DESCRIPTIVES
  /CINTERVAL 95
  /MISSING LISTWISE
  /NOTOTAL.


* ============================================================.
* 步驟 8：信度分析（CITC + Alpha-if-Deleted + 題目間相關矩陣）.
* 關鍵輸出：.
*   Item-Total Statistics：CITC（< .30 考慮刪題）、Alpha-if-Deleted（> 整體α 考慮刪題）.
*   Inter-Item Correlation Matrix：題目間相關矩陣.
*   Item Statistics：各題平均數與標準差.
*   Summary Item Statistics：各題平均數與變異數摘要.
* ============================================================.

* --- HP 階層停滯（T1）---.
RELIABILITY
  /VARIABLES = HP1_T1 HP2_T1 HP3_T1 HP4_T1 HP5_T1 HP6_T1
  /SCALE('HP 階層停滯 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* --- JCP 工作內容停滯（T1）---.
RELIABILITY
  /VARIABLES = {jcp_r1}
  /SCALE('JCP 工作內容停滯 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* --- PP 主動型人格（T1）---.
RELIABILITY
  /VARIABLES = PP1_T1 PP2_T1 PP3_T1 PP4_T1 PP5_T1 PP6_T1
  /SCALE('PP 主動型人格 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* --- DP 決策拖延（T1）---.
RELIABILITY
  /VARIABLES = {dp_r1}
  /SCALE('DP 決策拖延 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* --- CI 職涯無所作為（T1）---.
RELIABILITY
  /VARIABLES = CI1_T1 CI2_T1 CI3_T1 CI4_T1 CI5_T1 CI6_T1 CI7_T1 CI8_T1
  /SCALE('CI 職涯無所作為 T1') ALL
  /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR
  /SUMMARY = TOTAL MEANS VARIANCE.

* --- T2 信度 ---.
RELIABILITY
  /VARIABLES = HP1_T2 HP2_T2 HP3_T2 HP4_T2 HP5_T2 HP6_T2
  /SCALE('HP 階層停滯 T2') ALL /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR /SUMMARY = TOTAL MEANS VARIANCE.
RELIABILITY
  /VARIABLES = {jcp_r2}
  /SCALE('JCP 工作內容停滯 T2') ALL /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR /SUMMARY = TOTAL MEANS VARIANCE.
RELIABILITY
  /VARIABLES = PP1_T2 PP2_T2 PP3_T2 PP4_T2 PP5_T2 PP6_T2
  /SCALE('PP 主動型人格 T2') ALL /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR /SUMMARY = TOTAL MEANS VARIANCE.
RELIABILITY
  /VARIABLES = {dp_r2}
  /SCALE('DP 決策拖延 T2') ALL /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR /SUMMARY = TOTAL MEANS VARIANCE.
RELIABILITY
  /VARIABLES = CI1_T2 CI2_T2 CI3_T2 CI4_T2 CI5_T2 CI6_T2 CI7_T2 CI8_T2
  /SCALE('CI 職涯無所作為 T2') ALL /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR /SUMMARY = TOTAL MEANS VARIANCE.

* --- T3 信度 ---.
RELIABILITY
  /VARIABLES = HP1_T3 HP2_T3 HP3_T3 HP4_T3 HP5_T3 HP6_T3
  /SCALE('HP 階層停滯 T3') ALL /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR /SUMMARY = TOTAL MEANS VARIANCE.
RELIABILITY
  /VARIABLES = {jcp_r3}
  /SCALE('JCP 工作內容停滯 T3') ALL /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR /SUMMARY = TOTAL MEANS VARIANCE.
RELIABILITY
  /VARIABLES = PP1_T3 PP2_T3 PP3_T3 PP4_T3 PP5_T3 PP6_T3
  /SCALE('PP 主動型人格 T3') ALL /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR /SUMMARY = TOTAL MEANS VARIANCE.
RELIABILITY
  /VARIABLES = {dp_r3}
  /SCALE('DP 決策拖延 T3') ALL /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR /SUMMARY = TOTAL MEANS VARIANCE.
RELIABILITY
  /VARIABLES = CI1_T3 CI2_T3 CI3_T3 CI4_T3 CI5_T3 CI6_T3 CI7_T3 CI8_T3
  /SCALE('CI 職涯無所作為 T3') ALL /MODEL = ALPHA
  /STATISTICS = DESCRIPTIVE SCALE CORR /SUMMARY = TOTAL MEANS VARIANCE.


* ============================================================.
* 步驟 9：高/低 PP 群組比較（獨立樣本 t 檢定）.
* 目的：確認 PP 中位數分群的效度 / H8 輔助描述.
* ============================================================.
T-TEST GROUPS = PP_group(0 1)
  /MISSING = ANALYSIS
  /VARIABLES = HP_T1 JCP_T1 DP_T1 CI_T1
  /CRITERIA = CI(.95).


* ============================================================.
* 結束語.
* ============================================================.
* 以上完成所有 SPSS 可執行分析.
* CFA 量測模型與測量不變性請使用 Mplus 語法（.inp）執行.
* 詳見同目錄下的 Mplus .inp 語法檔.
"""


# ==========================================
# CFA 用 dat 檔（原始題目，T1）
# ==========================================
def generate_cfa_dat(df, output_dir, ts, exclude=None):
    """匯出 T1 原始題目供五因子 CFA 使用（無標題，空白分隔）。
    exclude: list of item names to drop, e.g. ['JCP6', 'DP1']
    """
    excl = set(exclude or [])
    cfa_cols = (
        [f'HP{i+1}_T1'  for i in range(6)] +
        [f'JCP{i+1}_T1' for i in range(6)] +
        [f'PP{i+1}_T1'  for i in range(6)] +
        [f'DP{i+1}_T1'  for i in range(5)] +
        [f'CI{i+1}_T1'  for i in range(8)]
    )
    # 排除指定題目（如 JCP6_T1）並過濾不存在的欄位
    cfa_cols = [c for c in cfa_cols
                if c.rsplit('_', 1)[0] not in excl and c in df.columns]
    cfa_df = df[cfa_cols].fillna(-999)
    tag = ('_excl' + '_'.join(sorted(excl))) if excl else ''
    dat_filename = f"CFA_Data_T1{tag}_{ts}.dat"
    dat_path = os.path.join(output_dir, dat_filename)
    cfa_df.to_csv(dat_path, sep=' ', index=False, header=False, float_format='%.4f')
    return dat_path, dat_filename


def generate_cfa_h_dat(df, output_dir, ts):
    """跨波次 CFA-H 資料：HP/JCP/PP 取 T1，DP 取 T2，CI 取 T3（同一份 .dat，短變數名）"""
    col_map = (
        [(f'HP{i+1}_T1',  f'HP{i+1}')  for i in range(6)] +
        [(f'JCP{i+1}_T1', f'JCP{i+1}') for i in range(6)] +
        [(f'PP{i+1}_T1',  f'PP{i+1}')  for i in range(6)] +
        [(f'DP{i+1}_T2',  f'DP{i+1}')  for i in range(5)] +
        [(f'CI{i+1}_T3',  f'CI{i+1}')  for i in range(8)]
    )
    col_map = [(src, dst) for src, dst in col_map if src in df.columns]
    src_cols = [src for src, _ in col_map]
    dst_cols = [dst for _, dst in col_map]
    cfa_h_df = df[src_cols].rename(columns=dict(col_map)).reindex(columns=dst_cols).fillna(-999)
    dat_filename = f"CFA_H_CrossWave_{ts}.dat"
    dat_path = os.path.join(output_dir, dat_filename)
    cfa_h_df.to_csv(dat_path, sep=' ', index=False, header=False, float_format='%.4f')
    return dat_path, dat_filename


# ==========================================
# Mplus CFA 語法（五因子量測模型）
# ==========================================
def generate_mplus_cfa_five_factor(dat_filename, ts):
    return f"""\
TITLE:
  五因子 CFA — HP / JCP / PP / DP / CI（T1）
  Generated: {ts}

DATA:
  FILE = "{dat_filename}";

VARIABLE:
  NAMES =
    HP1 HP2 HP3 HP4 HP5 HP6
    JCP1 JCP2 JCP3 JCP4 JCP5 JCP6
    PP1 PP2 PP3 PP4 PP5 PP6
    DP1 DP2 DP3 DP4 DP5
    CI1 CI2 CI3 CI4 CI5 CI6 CI7 CI8;
  USEVARIABLES = ALL;
  MISSING = ALL(-999);

ANALYSIS:
  ESTIMATOR = MLR;

MODEL:
  ! 因子負荷量自由估計（* 符號），以固定因子變異數 @1 識別
  HP  BY HP1*  HP2  HP3  HP4  HP5  HP6;   HP@1;
  JCP BY JCP1* JCP2 JCP3 JCP4 JCP5 JCP6;  JCP@1;
  PP  BY PP1*  PP2  PP3  PP4  PP5  PP6;   PP@1;
  DP  BY DP1*  DP2  DP3  DP4  DP5;        DP@1;
  CI  BY CI1*  CI2  CI3  CI4  CI5  CI6  CI7  CI8;  CI@1;

  ! HP 與 JCP 同屬 CP 構念，允許共變
  HP WITH JCP;

OUTPUT:
  STDYX;           ! 標準化負荷量（報告用）
  MODINDICES(10);  ! 修正指標（≥ 10 才考慮修正）
  CINTERVAL;       ! 95% 信賴區間
"""


# ==========================================
# Mplus 測量恆等性語法（Configural / Metric / Scalar）
# ==========================================
def generate_mplus_cfa_three_factor(dat_filename, ts):
    """三因子 CFA：CP合併（12題）/ DP / CI — 僅主路徑構念"""
    return f"""\
TITLE:
  三因子 CFA（主路徑構念）— CP / DP / CI（T1）
  CP = HP1~HP6 + JCP1~JCP6 共 12 題視為單一構念
  不含 PP（主動型人格）
  Generated: {ts}
  【用途】驗證主路徑三構念之區別效度，與四因子/五因子比較

DATA:
  FILE = "{dat_filename}";

VARIABLE:
  NAMES =
    HP1 HP2 HP3 HP4 HP5 HP6
    JCP1 JCP2 JCP3 JCP4 JCP5 JCP6
    PP1 PP2 PP3 PP4 PP5 PP6
    DP1 DP2 DP3 DP4 DP5
    CI1 CI2 CI3 CI4 CI5 CI6 CI7 CI8;
  USEVARIABLES =
    HP1 HP2 HP3 HP4 HP5 HP6
    JCP1 JCP2 JCP3 JCP4 JCP5 JCP6
    DP1 DP2 DP3 DP4 DP5
    CI1 CI2 CI3 CI4 CI5 CI6 CI7 CI8;
  MISSING = ALL(-999);

ANALYSIS:
  ESTIMATOR = MLR;

MODEL:
  ! CP = HP + JCP 合併為單一因子（12 題）
  CP BY HP1* HP2 HP3 HP4 HP5 HP6
         JCP1 JCP2 JCP3 JCP4 JCP5 JCP6;  CP@1;

  DP BY DP1* DP2 DP3 DP4 DP5;  DP@1;
  CI BY CI1* CI2 CI3 CI4 CI5 CI6 CI7 CI8;  CI@1;

OUTPUT:
  STDYX;
  MODINDICES(10);
  CINTERVAL;
"""


def generate_mplus_cfa_four_factor(dat_filename, ts):
    """四因子 CFA：CP 合併（12 題）/ PP / DP / CI — 與五因子對照比較"""
    return f"""\
TITLE:
  四因子 CFA（對照模型）— CP合併 / PP / DP / CI（T1）
  CP = HP1~HP6 + JCP1~JCP6 共 12 題視為單一構念
  Generated: {ts}
  【用途】與五因子模型比較 CFI/RMSEA，驗證 HP/JCP 是否需要分開

DATA:
  FILE = "{dat_filename}";

VARIABLE:
  NAMES =
    HP1 HP2 HP3 HP4 HP5 HP6
    JCP1 JCP2 JCP3 JCP4 JCP5 JCP6
    PP1 PP2 PP3 PP4 PP5 PP6
    DP1 DP2 DP3 DP4 DP5
    CI1 CI2 CI3 CI4 CI5 CI6 CI7 CI8;
  USEVARIABLES = ALL;
  MISSING = ALL(-999);

ANALYSIS:
  ESTIMATOR = MLR;

MODEL:
  ! CP = HP + JCP 合併為單一因子（12 題）
  CP BY HP1* HP2 HP3 HP4 HP5 HP6
         JCP1 JCP2 JCP3 JCP4 JCP5 JCP6;  CP@1;

  PP BY PP1* PP2 PP3 PP4 PP5 PP6;  PP@1;
  DP BY DP1* DP2 DP3 DP4 DP5;     DP@1;
  CI BY CI1* CI2 CI3 CI4 CI5 CI6 CI7 CI8;  CI@1;

OUTPUT:
  STDYX;
  MODINDICES(10);
  CINTERVAL;
"""


def generate_mi_inp_files(df, run_dir, ts):
    """
    產出測量不變性（MI）所需資料與 .inp 語法（Model A/B × Step 1/2/3，共 8 檔）。
    Model A：JCP（6題）+ DP（5題）+ CI（8題），三波 = 57 個欄位
    Model B：HP（6題）+ DP（5題）+ CI（8題），三波 = 57 個欄位
    """

    def _mi_dat(df, prefixes_n, run_dir, ts, tag):
        """寫出 MI 用 dat 檔，回傳 (dat_path, dat_filename)"""
        cols_ordered = []
        for wave in ['T1', 'T2', 'T3']:
            for prefix, n in prefixes_n:
                for i in range(1, n + 1):
                    cols_ordered.append(f'{prefix}{i}_{wave}')
        valid_cols = [c for c in cols_ordered if c in df.columns]
        mi_df = df[valid_cols].apply(pd.to_numeric, errors='coerce').fillna(-999)
        dat_fn = f"MI_{tag}_Data_{ts}.dat"
        dat_path = os.path.join(run_dir, dat_fn)
        mi_df.to_csv(dat_path, sep=' ', index=False, header=False, float_format='%.4f')
        return dat_path, dat_fn

    def _var_block(prefixes_n, wave_suffix=''):
        """Generate NAMES / USEVARIABLES lines for a wave or all waves."""
        lines = []
        for wave in ['T1', 'T2', 'T3']:
            for prefix, n in prefixes_n:
                items = ' '.join(f'{prefix}{i}{wave_suffix}{wave}' for i in range(1, n + 1))
                lines.append(f'    {items}')
        return '\n'.join(lines)

    def _wrap_by(factor, items_with_labels, max_len=85):
        """把 BY 語法行拆成不超過 max_len 字元的多行（Mplus 7.4 上限 90）。"""
        prefix = f'  {factor} BY '
        indent = ' ' * len(prefix)
        lines, cur = [], prefix
        for token in items_with_labels:
            candidate = cur + token + ' '
            if len(candidate) > max_len and cur != prefix:
                lines.append(cur.rstrip())
                cur = indent + token + ' '
            else:
                cur = candidate
        lines.append(cur.rstrip() + ';')
        return '\n'.join(lines)

    def _resid_cov_block(prefixes_n):
        """
        Cross-wave residual covariances for same indicator across T1/T2/T3.
        Standard practice in longitudinal MI (Little 2013; Widaman 1985):
        item-specific variance is stable over time and must be freed.
        Without this, the model cannot converge because it cannot explain
        the extra correlation between the same item measured at different waves.
        """
        lines = ['  ! Cross-wave residual covariances (same item, different waves)']
        for prefix, n in prefixes_n:
            for i in range(1, n + 1):
                v = f'{prefix}{i}'
                lines.append(f'  {v}_T1 WITH {v}_T2 {v}_T3;')
                lines.append(f'  {v}_T2 WITH {v}_T3;')
        return lines

    def _configural(dat_fn, prefixes_n, tag):
        var_names = []
        for wave in ['T1', 'T2', 'T3']:
            for prefix, n in prefixes_n:
                var_names += [f'{prefix}{i}_{wave}' for i in range(1, n + 1)]
        names_str = '\n    '.join(
            ' '.join(var_names[i:i+10]) for i in range(0, len(var_names), 10))
        model_lines = []
        for wi, wave in enumerate(['T1', 'T2', 'T3'], 1):
            for prefix, n in prefixes_n:
                items = ' '.join(f'{prefix}{i}_{wave}' for i in range(1, n + 1))
                factor = f'F_{prefix}{wi}'
                model_lines.append(f'  {factor} BY {items}* ({prefix.lower()}_l{wi}_1);')
                model_lines.append(f'  {factor}@1;')
            model_lines.append('')
        model_lines.append('')
        model_lines.extend(_resid_cov_block(prefixes_n))
        return f"""\
TITLE:
  Measurement Invariance - Configural Model
  Model {tag}, Three-wave longitudinal
  Generated: {ts}

DATA:
  FILE = "{dat_fn}";

VARIABLE:
  NAMES =
    {names_str};
  USEVARIABLES = ALL;
  MISSING = ALL(-999);

ANALYSIS:
  ESTIMATOR = MLR;
  ITERATIONS = 10000;

MODEL:
{''.join(l + chr(10) for l in model_lines)}
OUTPUT:
  STDYX;
  MODINDICES(10);
  CINTERVAL;
"""

    def _metric(dat_fn, prefixes_n, tag):
        var_names = []
        for wave in ['T1', 'T2', 'T3']:
            for prefix, n in prefixes_n:
                var_names += [f'{prefix}{i}_{wave}' for i in range(1, n + 1)]
        names_str = '\n    '.join(
            ' '.join(var_names[i:i+10]) for i in range(0, len(var_names), 10))
        model_lines = []
        for prefix, n in prefixes_n:
            for wi, wave in enumerate(['T1', 'T2', 'T3'], 1):
                items_constraints = ' '.join(
                    f'{prefix}{i}_{wave} ({prefix.lower()}_l{i})' for i in range(1, n + 1))
                factor = f'F_{prefix}{wi}'
                # First item gets * to free estimate, rest constrained equal across waves
                items_tokens = [f'{prefix}1_{wave}* ({prefix.lower()}_l1)'] + [
                    f'{prefix}{i}_{wave} ({prefix.lower()}_l{i})' for i in range(2, n + 1)]
                model_lines.append(_wrap_by(f'F_{prefix}{wi}', items_tokens))
                model_lines.append(f'  F_{prefix}{wi}@1;')
            model_lines.append('')
        model_lines.append('')
        model_lines.extend(_resid_cov_block(prefixes_n))
        return f"""\
TITLE:
  Measurement Invariance - Metric Model (equal loadings)
  Model {tag}, Three-wave longitudinal
  Generated: {ts}

DATA:
  FILE = "{dat_fn}";

VARIABLE:
  NAMES =
    {names_str};
  USEVARIABLES = ALL;
  MISSING = ALL(-999);

ANALYSIS:
  ESTIMATOR = MLR;
  ITERATIONS = 10000;

MODEL:
{''.join(l + chr(10) for l in model_lines)}
OUTPUT:
  STDYX;
  MODINDICES(10);
  CINTERVAL;
"""

    def _scalar(dat_fn, prefixes_n, tag):
        var_names = []
        for wave in ['T1', 'T2', 'T3']:
            for prefix, n in prefixes_n:
                var_names += [f'{prefix}{i}_{wave}' for i in range(1, n + 1)]
        names_str = '\n    '.join(
            ' '.join(var_names[i:i+10]) for i in range(0, len(var_names), 10))
        model_lines = []
        for prefix, n in prefixes_n:
            for wi, wave in enumerate(['T1', 'T2', 'T3'], 1):
                items_first = f'{prefix}1_{wave}* ({prefix.lower()}_l1)'
                items_tokens = [f'{prefix}1_{wave}* ({prefix.lower()}_l1)'] + [
                    f'{prefix}{i}_{wave} ({prefix.lower()}_l{i})' for i in range(2, n + 1)]
                model_lines.append(_wrap_by(f'F_{prefix}{wi}', items_tokens))
                model_lines.append(f'  F_{prefix}{wi}@1;')
            # intercept constraints across waves
            for i in range(1, n + 1):
                intercepts = ' '.join(
                    f'[{prefix}{i}_{wave}]' for wave in ['T1', 'T2', 'T3'])
                model_lines.append(f'  {intercepts} ({prefix.lower()}_int{i});')
            model_lines.append('')
        model_lines.append('')
        model_lines.extend(_resid_cov_block(prefixes_n))
        return f"""\
TITLE:
  Measurement Invariance - Scalar Model (equal intercepts)
  Model {tag}, Three-wave longitudinal
  Generated: {ts}

DATA:
  FILE = "{dat_fn}";

VARIABLE:
  NAMES =
    {names_str};
  USEVARIABLES = ALL;
  MISSING = ALL(-999);

ANALYSIS:
  ESTIMATOR = MLR;
  ITERATIONS = 10000;

MODEL:
{''.join(l + chr(10) for l in model_lines)}
OUTPUT:
  STDYX;
  MODINDICES(10);
  CINTERVAL;
"""

    # Model A: JCP + DP + CI
    a_prefixes = [('JCP', 6), ('DP', 5), ('CI', 8)]
    _, a_dat_fn = _mi_dat(df, a_prefixes, run_dir, ts, 'A_JCP_DP_CI')

    # Model B: HP + DP + CI
    b_prefixes = [('HP', 6), ('DP', 5), ('CI', 8)]
    _, b_dat_fn = _mi_dat(df, b_prefixes, run_dir, ts, 'B_HP_DP_CI')

    files_generated = []
    for tag, dat_fn, prefixes in [('A_JCP_DP_CI', a_dat_fn, a_prefixes),
                                   ('B_HP_DP_CI',  b_dat_fn, b_prefixes)]:
        for step_label, gen_fn in [
            ('Step1_Configural', _configural),
            ('Step2_Metric',     _metric),
            ('Step3_Scalar',     _scalar),
        ]:
            content = gen_fn(dat_fn, prefixes, tag)
            fname   = f"MI_{tag}_{step_label}_{ts}.inp"
            fpath   = os.path.join(run_dir, fname)
            with open(fpath, 'w', encoding='utf-8') as fh:
                fh.write(content)
            files_generated.append(fname)

    print(f"  [MI] 已產出 {len(files_generated)} 個測量不變性 .inp 檔：{', '.join(files_generated)}")
    return files_generated


def generate_mplus_measurement_invariance(dat_filename_ri, ts):
    """
    使用 parcel 合成分數（HP, JCP, PP, DP, CI）跨三波測量恆等性（保留供參考）
    """
    return f"""\
TITLE:
  測量恆等性檢定（Measurement Invariance）
  HP / JCP / PP / DP / CI 三波追蹤
  產生時間 {ts}

! ============================================================
! 注意：以下三個模型請分別儲存為三個 .inp 檔案跑
!   MI_Step1_Configural.inp  （形態模式）
!   MI_Step2_Metric.inp      （因子負荷等同）
!   MI_Step3_Scalar.inp      （截距等同）
! ============================================================

DATA:
  FILE = "{dat_filename_ri}";   ! 使用含三波 parcel 分數的 dat

VARIABLE:
  NAMES =
    HP_T1 JCP_T1 PP_T1 DP_T1 CI_T1
    HP_T2 JCP_T2 PP_T2 DP_T2 CI_T2
    HP_T3 JCP_T3 PP_T3 DP_T3 CI_T3
    Gender Age Tenure Position;
  USEVARIABLES =
    HP_T1 JCP_T1 PP_T1 DP_T1 CI_T1
    HP_T2 JCP_T2 PP_T2 DP_T2 CI_T2
    HP_T3 JCP_T3 PP_T3 DP_T3 CI_T3;
  MISSING = ALL(-999);

ANALYSIS:
  ESTIMATOR = MLR;

! ===========================================================
! 【Step 1 — Configural Model（形態模式）：因子結構相同即可】
! ===========================================================
MODEL:
  ! T1 因子（自由估計負荷量）
  F_HP1  BY HP_T1*;   F_HP1@1;
  F_JCP1 BY JCP_T1*;  F_JCP1@1;
  F_PP1  BY PP_T1*;   F_PP1@1;
  F_DP1  BY DP_T1*;   F_DP1@1;
  F_CI1  BY CI_T1*;   F_CI1@1;
  ! T2
  F_HP2  BY HP_T2*;   F_HP2@1;
  F_JCP2 BY JCP_T2*;  F_JCP2@1;
  F_PP2  BY PP_T2*;   F_PP2@1;
  F_DP2  BY DP_T2*;   F_DP2@1;
  F_CI2  BY CI_T2*;   F_CI2@1;
  ! T3
  F_HP3  BY HP_T3*;   F_HP3@1;
  F_JCP3 BY JCP_T3*;  F_JCP3@1;
  F_PP3  BY PP_T3*;   F_PP3@1;
  F_DP3  BY DP_T3*;   F_DP3@1;
  F_CI3  BY CI_T3*;   F_CI3@1;

! ===========================================================
! 【Step 2 — Metric Model（因子負荷等同）：跨波加等同約束】
! 請複製此段改為以下語法再另存 MI_Step2_Metric.inp
! ===========================================================
! MODEL:
!   F_HP1  BY HP_T1*  (l_hp);
!   F_HP2  BY HP_T2   (l_hp);
!   F_HP3  BY HP_T3   (l_hp);
!   F_HP1@1; F_HP2@1; F_HP3@1;
!   (... 其餘 JCP / PP / DP / CI 同樣加 (l_xxx) 標籤)

! ===========================================================
! 【Step 3 — Scalar Model（截距等同）：再加截距約束】
! 請另存 MI_Step3_Scalar.inp，在 Metric 基礎上加：
! ===========================================================
! MODEL:
!   [HP_T1] (int_hp);  [HP_T2] (int_hp);  [HP_T3] (int_hp);
!   [JCP_T1](int_jcp); [JCP_T2](int_jcp); [JCP_T3](int_jcp);
!   (... 其餘同樣加截距等同標籤)

OUTPUT:
  STDYX;
  MODINDICES(10);
  CINTERVAL;
"""


def generate_mplus_dat(df, output_dir, ts, exclude=None):
    """
    從個別題目欄位計算 parcel 平均，產出 Mplus 用的 .dat 檔
    HP_T1 = mean(HP1_T1..HP6_T1), JCP_T1 = mean(JCP1_T1..JCP6_T1), etc.
    exclude: list of item names to skip when computing parcel means, e.g. ['JCP6', 'DP1']
    """
    excl = set(exclude or [])

    def pmean(data, prefix, n, wave):
        cols = [f'{prefix}{i+1}_{wave}' for i in range(n)
                if f'{prefix}{i+1}' not in excl]
        cols = [c for c in cols if c in data.columns]
        return data[cols].mean(axis=1) if cols else pd.Series([-999]*len(data), index=data.index)

    mplus_df = pd.DataFrame(index=df.index)
    for wave in ['T1', 'T2', 'T3']:
        mplus_df[f'HP_{wave}']  = pmean(df, 'HP',  6, wave)
        mplus_df[f'JCP_{wave}'] = pmean(df, 'JCP', 6, wave)
        mplus_df[f'PP_{wave}']  = pmean(df, 'PP',  6, wave)
        mplus_df[f'DP_{wave}']  = pmean(df, 'DP',  5, wave)
        mplus_df[f'CI_{wave}']  = pmean(df, 'CI',  8, wave)

    # 控制變數（Step3 用）
    for col, src in [('Gender','Gender'), ('Age','Age'),
                     ('Tenure','NowJobTenure'), ('Position','Position')]:
        mplus_df[col] = df[src] if src in df.columns else -999

    # PP 中位數切割：PP_group = 0(低PP) / 1(高PP)，依 T1 PP 分數
    pp_median = mplus_df['PP_T1'].replace(-999, np.nan).median()
    mplus_df['PP_group'] = mplus_df['PP_T1'].apply(
        lambda x: -999 if x == -999 else (1 if x > pp_median else 0))
    print(f"[PP分群] T1 PP 中位數 = {pp_median:.3f}  "
          f"低PP組(0): {(mplus_df['PP_group']==0).sum()}人  "
          f"高PP組(1): {(mplus_df['PP_group']==1).sum()}人")

    # 職涯階段虛擬變數 + 效果編碼（deviation coding：各階段 vs. 其他兩階段平均）
    _age = mplus_df['Age'].replace(-999, np.nan)
    mplus_df['EXP']   = ((_age >= 21) & (_age <= 30)).astype(int)
    mplus_df['MAINT'] = (_age >= 41).astype(int)
    mplus_df.loc[mplus_df['Age'] == -999, ['EXP', 'MAINT']] = -999
    # 效果編碼：focal stage = +1，其他兩階段各 = -0.5
    mplus_df['EXP_C']   = np.where(mplus_df['Age'] == -999, -999.0,
                           np.where((_age >= 21) & (_age <= 30),  1.0, -0.5))
    mplus_df['MAINT_C'] = np.where(mplus_df['Age'] == -999, -999.0,
                           np.where(_age >= 41,                    1.0, -0.5))
    _n_exp   = (mplus_df['EXP']   == 1).sum()
    _n_maint = (mplus_df['MAINT'] == 1).sum()
    _n_estab = ((mplus_df['EXP'] == 0) & (mplus_df['MAINT'] == 0) & (mplus_df['Age'] != -999)).sum()
    print(f"[職涯階段] 探索期(EXP,21-30): {_n_exp}人  "
          f"建立期(31-40): {_n_estab}人  "
          f"維持期(MAINT,41+): {_n_maint}人  "
          f"（效果編碼：各階段 vs. 另外兩階段平均）")

    mplus_df = mplus_df.fillna(-999)
    tag = ('_excl' + '_'.join(sorted(excl))) if excl else ''
    dat_filename = f"Mplus_Data{tag}_{ts}.dat"
    dat_path = os.path.join(output_dir, dat_filename)
    mplus_df.to_csv(dat_path, sep=' ', index=False, header=False,
                    float_format='%.4f')
    return dat_path, dat_filename


# ==========================================
# MODULE A: P值格式化工具
# ==========================================
def fmt_p(p):
    """
    格式化 p 值，回傳 (星號str, p值str)
    星號：*** p<.001 / ** p<.01 / * p<.05 / '' 不顯著
    p值：p < .001 / p = .032（三位小數）
    """
    if p is None or (isinstance(p, float) and np.isnan(p)):
        return '', 'N/A'
    star  = '***' if p < .001 else '**' if p < .01 else '*' if p < .05 else ''
    p_str = 'p < .001' if p < .001 else f'p = {p:.3f}'
    return star, p_str


def fmt_beta(b, p, decimals=3):
    """
    格式化路徑係數：β值 + 星號（如 .234**），同時回傳 p 值字串
    用於 Excel 表格同一格顯示係數+顯著性，另一欄顯示 p 值
    """
    if b is None or (isinstance(b, float) and np.isnan(b)):
        return 'N/A', 'N/A'
    star, p_str = fmt_p(p)
    fmt = f'.{decimals}f'
    b_str = f'{b:{fmt}}{star}'
    return b_str, p_str


# ==========================================
# MODULE B: Mplus .inp 雙編碼儲存
# ==========================================
def save_inp_dual_encoding(content, run_dir, base_fname):
    """
    Save Mplus .inp syntax as UTF-8. Returns (utf8_path, None).
    Big5 version no longer generated since all comments are in English.
    """
    utf8_path = os.path.join(run_dir, base_fname + '.inp')

    with open(utf8_path, 'w', encoding='utf-8') as f:
        f.write(content)

    return utf8_path, None


# ==========================================
# MODULE C: Mplus 自動執行模組
# ==========================================
import subprocess

MPLUS_EXE_CANDIDATES = [
    r"C:\Program Files\Mplus\mplus.exe",
    r"C:\Program Files (x86)\Mplus\mplus.exe",
    r"C:\Mplus\mplus.exe",
    r"D:\Program Files\Mplus\mplus.exe",
]

def find_mplus_exe():
    """自動搜尋 Mplus 執行檔；找不到回傳 None"""
    for path in MPLUS_EXE_CANDIDATES:
        if os.path.isfile(path):
            return path
    return None


def run_mplus_single(inp_path, mplus_exe=None, timeout=300):
    """
    執行單一 Mplus .inp 檔。
    回傳: (success: bool, out_path: str|None, error_msg: str)
    """
    if mplus_exe is None:
        mplus_exe = find_mplus_exe()
    if mplus_exe is None:
        return False, None, "找不到 Mplus 執行檔，請確認安裝路徑"

    inp_dir  = os.path.dirname(os.path.abspath(inp_path))
    inp_file = os.path.basename(inp_path)
    out_file = inp_file.lower().replace('.inp', '.out')
    out_path = os.path.join(inp_dir, out_file)

    try:
        subprocess.run(
            [mplus_exe, inp_file],
            cwd=inp_dir,
            capture_output=True,
            text=True,
            timeout=timeout,
            encoding='utf-8',
            errors='replace'
        )
        success = os.path.isfile(out_path)
        return success, out_path if success else None, ''
    except subprocess.TimeoutExpired:
        return False, None, f'執行逾時（>{timeout}s）'
    except Exception as e:
        return False, None, str(e)


def run_all_mplus(inp_list, mplus_exe=None, timeout=300):
    """
    批次執行多個 Mplus .inp 檔。
    inp_list: [(label, inp_path), ...]
    回傳: [(label, success, out_path, err_msg), ...]
    """
    if mplus_exe is None:
        mplus_exe = find_mplus_exe()
    if mplus_exe is None:
        print("  [警告] 找不到 Mplus，跳過自動執行。請手動執行 .inp 檔後再繼續。")
        return [(label, False, None, '未找到 Mplus') for label, _ in inp_list]

    results = []
    for label, inp_path in inp_list:
        print(f"  [Mplus] {label}...", end=' ', flush=True)
        ok, out, err = run_mplus_single(inp_path, mplus_exe, timeout)
        print('[OK]' if ok else f'[FAIL] ({err})')
        results.append((label, ok, out, err))
    return results


# ==========================================
# MODULE D: Mplus .out 結果解析
# ==========================================
def parse_mplus_fit(out_path):
    """
    從 Mplus .out 擷取模型適配指數。
    回傳 dict 含：chi2, df, p_chi2, scaling_cf,
                  cfi, tli, rmsea, rmsea_lo, rmsea_hi, p_rmsea, srmr
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    try:
        with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
    except Exception:
        return {}

    fit = {}
    # chi-square
    m = re.search(r'Chi-Square Test of Model Fit\s+Value\s+([\d.]+)\*?', text)
    if m: fit['chi2'] = float(m.group(1))
    m = re.search(r'Degrees of Freedom\s+(\d+)', text)
    if m: fit['df'] = int(m.group(1))
    m = re.search(r'P-Value\s+([\d.]+)', text)
    if m: fit['p_chi2'] = float(m.group(1))
    m = re.search(r'Scaling Correction Factor\s+([\d.]+)', text)
    if m: fit['scaling_cf'] = float(m.group(1))
    # RMSEA
    m = re.search(r'RMSEA.*?Estimate\s+([\d.]+)\s+90 Percent C\.I\.\s+([\d.]+)\s+([\d.]+)',
                  text, re.DOTALL)
    if m:
        fit['rmsea']    = float(m.group(1))
        fit['rmsea_lo'] = float(m.group(2))
        fit['rmsea_hi'] = float(m.group(3))
    m = re.search(r'Probability RMSEA <= \.05\s+([\d.]+)', text)
    if m: fit['p_rmsea'] = float(m.group(1))
    # CFI / TLI
    m = re.search(r'\bCFI\s+([\d.]+)', text)
    if m: fit['cfi'] = float(m.group(1))
    m = re.search(r'\bTLI\s+([\d.]+)', text)
    if m: fit['tli'] = float(m.group(1))
    # SRMR
    m = re.search(r'SRMR \(Standardized Root Mean Square Residual\)\s+Value\s+([\d.]+)', text)
    if m: fit['srmr'] = float(m.group(1))

    return fit


def parse_mplus_stdyx(out_path, path_map):
    """
    從 Mplus .out 的 STDYX STANDARDIZATION 區段擷取路徑係數。

    path_map: dict，格式為：
        { '顯示標籤': ('結果變數', '預測變數') }
    例：
        { 'JCP→DP': ('WDP2', 'WJCP1'),
          'DP→CI':  ('WCI2', 'WDP1'),
          'PP→DP':  ('WDP2', 'WPP1') }

    回傳 dict：{ '標籤': {'est': float, 'se': float, 'z': float, 'p': float} }
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    try:
        with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
    except Exception:
        return {}

    # 找 STDYX 區段（到 R-SQUARE 或下一個大標題為止）
    m = re.search(r'STDYX Standardization\s+(.*?)(?=\nR-SQUARE|\nSTD |\Z)',
                  text, re.DOTALL)
    if not m:
        return {}
    stdyx = m.group(1)

    results = {}
    for label, (outcome, predictor) in path_map.items():
        # 找 outcome ON ... 區段
        on_m = re.search(
            rf'{re.escape(outcome)}\s+ON\s+(.*?)(?=\n\s*\n\s*\w|\n\s*\n\s*$)',
            stdyx, re.DOTALL | re.IGNORECASE
        )
        if not on_m:
            continue
        on_text = on_m.group(1)
        row = re.search(
            rf'\b{re.escape(predictor)}\s+([-\d.]+)\s+([\d.]+)\s+([-\d.]+)\s+([\d.]+)',
            on_text, re.IGNORECASE
        )
        if row:
            results[label] = {
                'est': float(row.group(1)),
                'se':  float(row.group(2)),
                'z':   float(row.group(3)),
                'p':   float(row.group(4))
            }

    # 從 CONFIDENCE INTERVALS OF STANDARDIZED MODEL RESULTS 補 95% CI
    # 注意：用 \Z 直接到檔末，避免 [A-Z]{3} 誤判預測變項名（如 JCP, HP）為終止符
    ci_sec_m = re.search(
        r'CONFIDENCE INTERVALS OF STANDARDIZED MODEL RESULTS.*?STDYX Standardization\s+(.*)',
        text, re.DOTALL
    )
    if ci_sec_m:
        ci_text = ci_sec_m.group(1)
        for label, (outcome, predictor) in path_map.items():
            if label not in results:
                continue
            ci_on_m = re.search(
                rf'{re.escape(outcome)}\s+ON\s+(.*?)(?=\n\s*\n\s*\w|\n\s*\n\s*$)',
                ci_text, re.DOTALL | re.IGNORECASE
            )
            if not ci_on_m:
                continue
            ci_on_text = ci_on_m.group(1)
            # 7 columns: lo.5% lo2.5% lo5% est hi5% hi2.5% hi.5%
            ci_row = re.search(
                rf'\b{re.escape(predictor)}\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)',
                ci_on_text, re.IGNORECASE
            )
            if ci_row:
                results[label]['ci_lo'] = float(ci_row.group(2))  # Lower 2.5%
                results[label]['ci_hi'] = float(ci_row.group(6))  # Upper 2.5%

    return results


def parse_mplus_ri_corr(out_path, ri_pairs):
    """
    從 Mplus .out 的 CONFIDENCE INTERVALS OF STANDARDIZED MODEL RESULTS 區段
    擷取隨機截距相關係數（STDYX 標準化）。

    ri_pairs: list of (label, ri_x, ri_y)
    例: [('RI_DP↔RI_CI', 'RI_DP', 'RI_CI'), ...]

    回傳: { label: {'est': float, 'ci_lo': float, 'ci_hi': float, 'sig': bool} }
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    try:
        with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
    except Exception:
        return {}

    # 從 CI OF STANDARDIZED 區段找 RI WITH 值（4 欄：lo95, lo90, lo68, est, hi68, hi90, hi95）
    ci_m = re.search(
        r'CONFIDENCE INTERVALS OF STANDARDIZED MODEL RESULTS.*?STDYX Standardization\s+(.*?)(?=\nTECHNICAL|\Z)',
        text, re.DOTALL
    )
    if not ci_m:
        return {}
    ci_text = ci_m.group(1)

    results = {}
    for label, ri_x, ri_y in ri_pairs:
        # 找 RI_X WITH ... 的 RI_Y 列
        with_m = re.search(
            rf'{re.escape(ri_x)}\s+WITH\s+(.*?)(?=\n\s*\n|\n\s*\w+\s+WITH|\Z)',
            ci_text, re.DOTALL | re.IGNORECASE
        )
        if not with_m:
            continue
        with_text = with_m.group(1)
        row = re.search(
            rf'\b{re.escape(ri_y)}\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)',
            with_text, re.IGNORECASE
        )
        if row:
            lo95 = float(row.group(2))   # Lower 2.5%
            est  = float(row.group(4))   # Estimate (中間值)
            hi95 = float(row.group(6))   # Upper 2.5%
            sig  = not (lo95 <= 0 <= hi95)   # 95% CI 不含 0 → 顯著
            results[label] = {
                'est': est,
                'ci_lo': lo95,
                'ci_hi': hi95,
                'sig': sig
            }
    return results


def parse_mplus_indirect(out_path, indirect_specs):
    """
    從 Mplus .out 解析 MODEL INDIRECT 間接效果（STDYX 標準化）。

    Mplus 7.4 輸出格式：
      STANDARDIZED TOTAL...INDIRECT...
      STDYX Standardization
        ...
        Effects from X to Y
          Indirect    Estimate  SE  Est./S.E.  P-Value
      CONFIDENCE INTERVALS OF STANDARDIZED TOTAL...
      STDYX Standardization
        ...
        Effects from X to Y
          Indirect    Lo.5%  Lo2.5%  Lo5%  Est  Hi5%  Hi2.5%  Hi.5%

    indirect_specs: list of (label, x_var, m_var, y_var)
      m_var 在此格式中不出現在獨立行，只要 x_var 和 y_var 對即可。
      例: [('H7a: JCP→DP→CI', 'WJCP1', 'WDP2', 'WCI3'), ...]

    回傳: { label: {'est': float, 'se': float, 'z': float, 'p': float,
                    'ci_lo': float, 'ci_hi': float, 'sig': bool} }
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    try:
        with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
    except Exception:
        return {}

    results = {}

    # ── 1. STDYX 標準化間接效果（Estimate SE z p）──
    std_m = re.search(
        r'STANDARDIZED TOTAL.*?INDIRECT.*?STDYX Standardization\s+(.*?)'
        r'(?=CONFIDENCE INTERVALS|\nTECHNICAL|\Z)',
        text, re.DOTALL | re.IGNORECASE
    )
    std_text = std_m.group(1) if std_m else ''

    # ── 2. STDYX CI (7 欄: lo.5% lo2.5% lo5% est hi5% hi2.5% hi.5%) ──
    ci_m = re.search(
        r'CONFIDENCE INTERVALS OF STANDARDIZED TOTAL.*?INDIRECT.*?STDYX Standardization\s+(.*?)'
        r'(?=\nTECHNICAL|\Z)',
        text, re.DOTALL | re.IGNORECASE
    )
    ci_text = ci_m.group(1) if ci_m else ''

    for label, x_var, m_var, y_var in indirect_specs:
        # 在 STDYX 區段找 "Effects from X to Y" 並取其下的 "Indirect" 行
        eff_m = re.search(
            rf'Effects from\s+{re.escape(x_var)}\s+to\s+{re.escape(y_var)}\s+(.*?)'
            rf'(?=Effects from|\Z)',
            std_text, re.DOTALL | re.IGNORECASE
        )
        if not eff_m:
            continue

        # "  Indirect    est  se  z  p" 行
        row = re.search(
            r'^\s*Indirect\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)',
            eff_m.group(1), re.MULTILINE | re.IGNORECASE
        )
        if not row:
            continue
        est = float(row.group(1))
        se  = float(row.group(2))
        z   = float(row.group(3))
        p   = float(row.group(4))

        # CI 區段
        ci_lo, ci_hi = np.nan, np.nan
        ci_eff = re.search(
            rf'Effects from\s+{re.escape(x_var)}\s+to\s+{re.escape(y_var)}\s+(.*?)'
            rf'(?=Effects from|\Z)',
            ci_text, re.DOTALL | re.IGNORECASE
        )
        if ci_eff:
            ci_row = re.search(
                r'^\s*Indirect\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)'
                r'\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)',
                ci_eff.group(1), re.MULTILINE | re.IGNORECASE
            )
            if ci_row:
                ci_lo = float(ci_row.group(2))   # Lower 2.5%
                ci_hi = float(ci_row.group(6))   # Upper 2.5%

        sig = (not (np.isnan(ci_lo) or np.isnan(ci_hi)) and not (ci_lo <= 0 <= ci_hi)) \
              if not (np.isnan(ci_lo) or np.isnan(ci_hi)) else (p < .05)
        results[label] = {'est': est, 'se': se, 'z': z, 'p': p,
                          'ci_lo': ci_lo, 'ci_hi': ci_hi, 'sig': sig}
    return results


def parse_mplus_modconstraint(out_path, param_names):
    """
    Parse MODEL CONSTRAINT NEW(...) parameters from Mplus .out.
    Looks for "New/Additional Parameters" in MODEL RESULTS and BC Bootstrap CI sections.
    param_names: list of uppercase parameter names (e.g. ['IND_HI_JCP', 'IND_LO_JCP']).
    Returns: {name: {'est', 'se', 'z', 'p', 'ci_lo', 'ci_hi', 'sig'}}
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    try:
        with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
    except Exception:
        return {}

    results = {}

    # --- Unstandardized estimates from MODEL RESULTS ---
    est_m = re.search(
        r'MODEL RESULTS\s.*?New/Additional Parameters\s+(.*?)'
        r'(?=\n\s*\n\s*[A-Z]|\nR-SQUARE|\Z)',
        text, re.DOTALL | re.IGNORECASE
    )
    est_text = est_m.group(1) if est_m else ''

    for pname in param_names:
        row = re.search(
            rf'^\s*{re.escape(pname)}\s+([-\d.]+)\s+([\d.]+)\s+([-\d.]+)\s+([\d.]+)',
            est_text, re.MULTILINE | re.IGNORECASE
        )
        if row:
            results[pname] = {
                'est': float(row.group(1)),
                'se':  float(row.group(2)),
                'z':   float(row.group(3)),
                'p':   float(row.group(4)),
                'ci_lo': np.nan,
                'ci_hi': np.nan,
                'sig': float(row.group(4)) < .05
            }

    # --- BC Bootstrap CI from CONFIDENCE INTERVALS OF MODEL RESULTS ---
    ci_m = re.search(
        r'CONFIDENCE INTERVALS OF MODEL RESULTS.*?New/Additional Parameters\s+(.*?)'
        r'(?=\nTECHNICAL|\Z)',
        text, re.DOTALL | re.IGNORECASE
    )
    ci_text = ci_m.group(1) if ci_m else ''

    for pname in param_names:
        ci_row = re.search(
            rf'^\s*{re.escape(pname)}\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)'
            r'\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)',
            ci_text, re.MULTILINE | re.IGNORECASE
        )
        if ci_row:
            ci_lo = float(ci_row.group(2))  # Lower 2.5%
            ci_hi = float(ci_row.group(6))  # Upper 2.5%
            sig = not (ci_lo <= 0 <= ci_hi)
            if pname in results:
                results[pname].update({'ci_lo': ci_lo, 'ci_hi': ci_hi, 'sig': sig})
            else:
                results[pname] = {
                    'est': float(ci_row.group(4)),
                    'se': np.nan, 'z': np.nan, 'p': np.nan,
                    'ci_lo': ci_lo, 'ci_hi': ci_hi, 'sig': sig
                }

    return results


def parse_mplus_bayes_paths(out_path, path_map):
    """
    Parse Bayesian Mplus STDYX output for path coefficients.
    Bayesian format per row: predictor  est  sd  p_onetail  ci_lo  ci_hi
    Significance = 95% HPD CI excludes 0.
    Returns {label: {'est', 'sd', 'p', 'ci_lo', 'ci_hi', 'sig'}}
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    try:
        with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
    except Exception:
        return {}

    m = re.search(
        r'STDYX Standardization\s+(.*?)(?=\nR-SQUARE|\nSTD |\nTECHNICAL|\Z)',
        text, re.DOTALL
    )
    if not m:
        return {}
    stdyx = m.group(1)

    results = {}
    for label, (outcome, predictor) in path_map.items():
        on_m = re.search(
            rf'{re.escape(outcome)}\s+ON\s+(.*?)(?=\n\s*\n\s*\w|\n\s*\n\s*$)',
            stdyx, re.DOTALL | re.IGNORECASE
        )
        if not on_m:
            continue
        row = re.search(
            rf'\b{re.escape(predictor)}\s+([-\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([-\d.]+)\s+([-\d.]+)',
            on_m.group(1), re.IGNORECASE
        )
        if row:
            est, sd, p = float(row.group(1)), float(row.group(2)), float(row.group(3))
            ci_lo, ci_hi = float(row.group(4)), float(row.group(5))
            results[label] = {
                'est': est, 'sd': sd, 'p': p,
                'ci_lo': ci_lo, 'ci_hi': ci_hi,
                'sig': not (ci_lo <= 0 <= ci_hi)
            }
    return results


def parse_mplus_bayes_modconstraint(out_path, param_names):
    """
    Parse Bayesian Mplus MODEL CONSTRAINT NEW parameters from STDYX section.
    Bayesian format: param  est  sd  p_onetail  ci_lo  ci_hi
    Significance = 95% HPD CI excludes 0.
    Returns {name: {'est', 'sd', 'p', 'ci_lo', 'ci_hi', 'sig'}}
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    try:
        with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
    except Exception:
        return {}

    # 先找第一個 New/Additional Parameters（MODEL RESULTS 段，5欄 Bayesian 格式）
    # 必須在 STANDARDIZED MODEL RESULTS 之前，避免讀到 STDYX 段的 7欄格式
    for pattern in [
        r'New/Additional Parameters\s+(.*?)(?=\nSTANDARDIZED|\nTECHNICAL|\Z)',
        r'STDYX Standardization.*?New/Additional Parameters\s+(.*?)(?=\nTECHNICAL|\Z)',
    ]:
        m = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if m:
            block = m.group(1)
            break
    else:
        return {}

    results = {}
    for pname in param_names:
        row = re.search(
            rf'^\s*{re.escape(pname)}\s+([-\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([-\d.]+)\s+([-\d.]+)',
            block, re.MULTILINE | re.IGNORECASE
        )
        if row:
            est, sd, p = float(row.group(1)), float(row.group(2)), float(row.group(3))
            ci_lo, ci_hi = float(row.group(4)), float(row.group(5))
            results[pname] = {
                'est': est, 'sd': sd, 'p': p,
                'ci_lo': ci_lo, 'ci_hi': ci_hi,
                'sig': not (ci_lo <= 0 <= ci_hi)
            }
    return results


def parse_mplus_cfa_loadings(out_path):
    """
    從 Mplus .out 的 STDYX Standardization 區段擷取 CFA 因素負荷量。

    回傳：
        list of dict，每筆含：
            factor (str)   — 因子名稱，如 'HP'
            item   (str)   — 題目名稱，如 'HP1'
            beta   (float) — STDYX 標準化負荷量
            se     (float) — 標準誤
            z      (float) — z 值
            p      (float) — p 值
    """
    if not out_path or not os.path.isfile(out_path):
        return []
    try:
        with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
    except Exception:
        return []

    # 找第一個 STDYX Standardization 區段（BY 語句）
    m = re.search(
        r'STDYX Standardization\s+'
        r'Two-Tailed\s+Estimate\s+S\.E\.\s+Est\./S\.E\.\s+P-Value\s+(.*?)'
        r'(?=\nSTDYX Standardization|\nCONFIDENCE INTERVALS|\nTECHNICAL|\Z)',
        text, re.DOTALL
    )
    if not m:
        return []

    block = m.group(1)
    rows = []
    current_factor = None

    for line in block.splitlines():
        # WITH 行出現代表 BY 區段已結束，停止解析
        if re.match(r'^\s*\w+\s+WITH\s*$', line):
            break
        # 因子行：" HP       BY" 或 "HP       BY"（第一行可能無前導空白）
        factor_m = re.match(r'^\s*(\w+)\s+BY\s*$', line)
        if factor_m:
            current_factor = factor_m.group(1)
            continue
        # 題目行：" HP1  0.710  0.041  17.383  0.000"
        if current_factor:
            item_m = re.match(
                r'^\s+(\w+)\s+([-\d.]+)\s+([\d.]+)\s+([-\d.]+)\s+([\d.]+)\s*$',
                line
            )
            if item_m:
                rows.append({
                    'factor': current_factor,
                    'item':   item_m.group(1),
                    'beta':   float(item_m.group(2)),
                    'se':     float(item_m.group(3)),
                    'z':      float(item_m.group(4)),
                    'p':      float(item_m.group(5)),
                })
    return rows


def run_and_parse_mi(run_dir, ts, mplus_exe=None):
    """
    找到測量不變性（MI）Step1/2/3 .inp 檔，自動執行並解析適配指數。
    計算 ΔCFI 與 ΔRMSEA（Step2-Step1, Step3-Step2）。

    回傳：
        dict，鍵為 'A (JCP)' 和 'B (HP)'，值為各步驟的適配與差異值。
    """
    mi_results = {}

    mi_specs = [
        ('A (JCP路徑)', 'A_JCP_DP_CI'),
        ('B (HP路徑)',  'B_HP_DP_CI'),
    ]
    step_labels = [
        ('Step1', 'Step1_Configural', '組態不變性'),
        ('Step2', 'Step2_Metric',     '因子負荷不變性'),
        ('Step3', 'Step3_Scalar',     '截距不變性'),
    ]

    for mi_label, tag in mi_specs:
        step_fits = {}
        inp_list  = []
        for step_key, step_fname_part, step_desc in step_labels:
            fname  = f"MI_{tag}_{step_fname_part}_{ts}.inp"
            fpath  = os.path.join(run_dir, fname)
            if os.path.isfile(fpath):
                inp_list.append((step_key, fpath, step_desc))

        if inp_list:
            print(f"[MI] 執行測量不變性 {mi_label}（{len(inp_list)} 步驟）...")
            for step_key, fpath, step_desc in inp_list:
                run_results = run_all_mplus([(step_key, fpath)], mplus_exe)
                for lbl, ok, out_path, err in run_results:
                    if ok:
                        step_fits[step_key] = parse_mplus_fit(out_path)
                    else:
                        step_fits[step_key] = {}

        # 計算 ΔCFI, ΔRMSEA
        mi_step_rows = []
        prev_fit = None
        for step_key, _, step_desc in step_labels:
            fit = step_fits.get(step_key, {})
            cfi_v   = fit.get('cfi')
            tli_v   = fit.get('tli')
            rmsea_v = fit.get('rmsea')
            srmr_v  = fit.get('srmr')
            chi2_v  = fit.get('chi2')
            df_v    = fit.get('df')
            if prev_fit is not None:
                prev_cfi   = prev_fit.get('cfi')
                prev_rmsea = prev_fit.get('rmsea')
                d_cfi   = (cfi_v   - prev_cfi)   if isinstance(cfi_v,   float) and isinstance(prev_cfi,   float) else None
                d_rmsea = (rmsea_v - prev_rmsea) if isinstance(rmsea_v, float) and isinstance(prev_rmsea, float) else None
            else:
                d_cfi, d_rmsea = None, None
            # 不變性判斷：|ΔCFI| < .01 且 |ΔRMSEA| < .015
            if d_cfi is not None and d_rmsea is not None:
                invariant = (abs(d_cfi) < .01 and abs(d_rmsea) < .015)
            else:
                invariant = None
            mi_step_rows.append({
                'step':      step_key,
                'desc':      step_desc,
                'chi2':      chi2_v,
                'df':        df_v,
                'cfi':       cfi_v,
                'tli':       tli_v,
                'rmsea':     rmsea_v,
                'srmr':      srmr_v,
                'd_cfi':     d_cfi,
                'd_rmsea':   d_rmsea,
                'invariant': invariant,
            })
            prev_fit = fit

        mi_results[mi_label] = mi_step_rows

    return mi_results


def calculate_ave_cr(out_path):
    """
    從 Mplus .out 檔案的 STDYX 區段擷取因素負荷量，
    計算各因子的 AVE（Average Variance Extracted）與 CR（Composite Reliability）。

    AVE = Σ(λ²) / n  （各因子負荷量平方的平均值）
    CR  = (Σλ)² / [(Σλ)² + Σ(1-λ²)]

    回傳：{factor: {'loadings': [...], 'n_items': int, 'AVE': float, 'CR': float}}
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    try:
        with open(out_path, 'r', encoding='utf-8', errors='replace') as fh:
            content = fh.read()
    except Exception:
        return {}

    # 定位 STDYX Standardization 區段
    stdyx_match = re.search(r'STDYX Standardization', content, re.IGNORECASE)
    if not stdyx_match:
        return {}
    stdyx_text = content[stdyx_match.start():]

    # 解析因子 BY 區塊（因素負荷量）
    # 格式: factor BY 區塊 -> 每行 item  β  SE  z  p
    factor_data = {}
    current_factor = None
    # 找到所有 "FACTOR BY" 區塊：先切分至下一個大段落
    by_sections = re.findall(
        r'([A-Z][A-Z0-9_]*)\s+BY\s*\n((?:\s+\S+\s+[\d\.\-]+\s+[\d\.\-]+\s+[\d\.\-]+\s+[\d\.\-]+\s*\n)+)',
        stdyx_text
    )
    for fac_name, rows_text in by_sections:
        fac_name_upper = fac_name.upper()
        # 只保留我們關心的因子
        if fac_name_upper not in ('HP', 'JCP', 'PP', 'DP', 'CI'):
            continue
        lambdas = []
        for row_m in re.finditer(
            r'\s+(\S+)\s+([\d\.\-]+)\s+([\d\.\-]+)\s+([\d\.\-]+)\s+([\d\.\-]+)',
            rows_text
        ):
            try:
                lam = float(row_m.group(2))
                lambdas.append(lam)
            except ValueError:
                pass
        if lambdas:
            factor_data[fac_name_upper] = lambdas

    if not factor_data:
        return {}

    result = {}
    for fac, lambdas in factor_data.items():
        n = len(lambdas)
        lam_sq  = [l ** 2 for l in lambdas]
        ave     = sum(lam_sq) / n if n > 0 else float('nan')
        sum_lam = sum(lambdas)
        sum_err = sum(1 - l2 for l2 in lam_sq)
        cr_denom = (sum_lam ** 2) + sum_err
        cr = (sum_lam ** 2) / cr_denom if cr_denom > 0 else float('nan')
        result[fac] = {
            'loadings': lambdas,
            'n_items': n,
            'AVE': round(ave, 3),
            'CR':  round(cr, 3),
        }
    return result


# ==========================================
# MODULE E: 整合執行所有 Mplus 模型並收集結果
# ==========================================
def run_and_parse_all_models(run_dir, mplus_dat_filename, cfa_dat_filename, ts,
                              mplus_exe=None, exclude=None, phases=None,
                              cfa_h_dat_filename=None):
    """
    生成 CFA/RI-CLPM .inp → 自動執行 → 解析結果
    exclude: list of item names to remove from CFA (e.g. ['JCP6', 'DP1'])
    回傳 all_results dict 供 Excel/Word/PPT 使用
    """
    if phases is None:
        phases = ['cfa', 'riclpm']
    excl = set(exclude or [])
    all_results = {}

    # ---- 動態建立各因子題目清單 ----
    def _items(prefix, total):
        return [f'{prefix}{i}' for i in range(1, total + 1) if f'{prefix}{i}' not in excl]

    jcp_items = _items('JCP', 6)
    hp_items  = _items('HP',  6)
    pp_items  = _items('PP',  6)
    dp_items  = _items('DP',  5)
    ci_items  = _items('CI',  8)

    def _vline(items):
        return '    ' + '  '.join(items)

    def _by_line(factor, items):
        if not items:
            return ''
        parts = [items[0] + '*'] + items[1:]
        return f'  {factor} BY {" ".join(parts)};  {factor}@1;\n'

    _v_jcp = _vline(jcp_items)
    _v_hp  = _vline(hp_items)
    _v_pp  = _vline(pp_items)
    _v_dp  = _vline(dp_items)
    _v_ci  = _vline(ci_items)

    # NAMES 區塊只列 CFA dat 實際包含的題目（排除後）
    all_items_ordered = (
        [f'HP{i}'  for i in range(1, 7)  if f'HP{i}'  not in excl] +
        [f'JCP{i}' for i in range(1, 7)  if f'JCP{i}' not in excl] +
        [f'PP{i}'  for i in range(1, 7)  if f'PP{i}'  not in excl] +
        [f'DP{i}'  for i in range(1, 6)  if f'DP{i}'  not in excl] +
        [f'CI{i}'  for i in range(1, 9)  if f'CI{i}'  not in excl]
    )
    names_block = '\n'.join(
        '    ' + '  '.join(all_items_ordered[i:i+8])
        for i in range(0, len(all_items_ordered), 8)
    )

    # ---- CFA 模型 A-G ----
    cfa_models = {
        'CFA-A (JCP+DP+CI)': {
            'fname': f'CFA_A_JCP_DP_CI_{ts}',
            'vars_lines': f'{_v_jcp}\n{_v_dp}\n{_v_ci}',
            'model_lines': (
                _by_line('JCP', jcp_items) +
                _by_line('DP',  dp_items)  +
                _by_line('CI',  ci_items)
            )
        },
        'CFA-B (HP+DP+CI)': {
            'fname': f'CFA_B_HP_DP_CI_{ts}',
            'vars_lines': f'{_v_hp}\n{_v_dp}\n{_v_ci}',
            'model_lines': (
                _by_line('HP', hp_items) +
                _by_line('DP', dp_items) +
                _by_line('CI', ci_items)
            )
        },
        'CFA-C (JCP+PP+DP+CI)': {
            'fname': f'CFA_C_JCP_PP_DP_CI_{ts}',
            'vars_lines': f'{_v_jcp}\n{_v_pp}\n{_v_dp}\n{_v_ci}',
            'model_lines': (
                _by_line('JCP', jcp_items) +
                _by_line('PP',  pp_items)  +
                _by_line('DP',  dp_items)  +
                _by_line('CI',  ci_items)
            )
        },
        'CFA-D (HP+PP+DP+CI)': {
            'fname': f'CFA_D_HP_PP_DP_CI_{ts}',
            'vars_lines': f'{_v_hp}\n{_v_pp}\n{_v_dp}\n{_v_ci}',
            'model_lines': (
                _by_line('HP', hp_items) +
                _by_line('PP', pp_items) +
                _by_line('DP', dp_items) +
                _by_line('CI', ci_items)
            )
        },
        # ---- Discriminant validity comparison models CFA-E/F/G ----
        'CFA-E (HP+JCP+PP+DP+CI, 5F)': {
            'fname': f'CFA_E_FiveFactor_{ts}',
            'vars_lines': f'{_v_hp}\n{_v_jcp}\n{_v_pp}\n{_v_dp}\n{_v_ci}',
            'model_lines': (
                _by_line('HP',  hp_items)  +
                _by_line('JCP', jcp_items) +
                _by_line('PP',  pp_items)  +
                _by_line('DP',  dp_items)  +
                _by_line('CI',  ci_items)  +
                '  HP WITH JCP;\n'
            )
        },
        'CFA-F (CP_merged+PP+DP+CI, 4F)': {
            'fname': f'CFA_F_FourFactor_CP_merged_{ts}',
            'vars_lines': f'{_v_hp}\n{_v_jcp}\n{_v_pp}\n{_v_dp}\n{_v_ci}',
            'model_lines': (
                '  CP  BY ' + '  '.join([i + ('*' if j == 0 else '') for j, i in enumerate(hp_items)]) +
                '\n          ' + '  '.join(jcp_items) + ';  CP@1;\n' +
                _by_line('PP', pp_items) +
                _by_line('DP', dp_items) +
                _by_line('CI', ci_items)
            )
        },
        'CFA-G (CP_merged+DP+CI, 3F)': {
            'fname': f'CFA_G_ThreeFactor_CP_DP_CI_{ts}',
            'vars_lines': f'{_v_hp}\n{_v_jcp}\n{_v_dp}\n{_v_ci}',
            'model_lines': (
                '  CP  BY ' + '  '.join([i + ('*' if j == 0 else '') for j, i in enumerate(hp_items)]) +
                '\n          ' + '  '.join(jcp_items) + ';  CP@1;\n' +
                _by_line('DP', dp_items) +
                _by_line('CI', ci_items)
            )
        },
    }

    cfa_inp_list = []
    for label, cfg in cfa_models.items():
        vl = cfg['vars_lines']
        content = (
            f'TITLE:\n  {label} CFA (T1)\n  Generated: {ts}\n\n'
            f'DATA:\n  FILE = "{cfa_dat_filename}";\n\n'
            f'VARIABLE:\n  NAMES =\n{names_block};\n'
            f'  USEVARIABLES =\n{vl};\n'
            f'  MISSING =\n{vl} (-999);\n\n'
            f'ANALYSIS:\n  ESTIMATOR = MLR;\n\n'
            f'MODEL:\n{cfg["model_lines"]}\n'
            f'OUTPUT:\n  STDYX;\n  MODINDICES(10);\n  CINTERVAL;\n'
        )
        utf8_path, b5_path = save_inp_dual_encoding(content, run_dir, cfg['fname'])
        cfa_inp_list.append((label, utf8_path))

    # ---- CFA-H：跨波次五因子 CFA（HP/JCP/PP@T1，DP@T2，CI@T3）----
    if cfa_h_dat_filename:
        _h_vars = f'{_v_hp}\n{_v_jcp}\n{_v_pp}\n{_v_dp}\n{_v_ci}'
        _h_model = (
            _by_line('HP',  hp_items)  +
            _by_line('JCP', jcp_items) +
            _by_line('PP',  pp_items)  +
            _by_line('DP',  dp_items)  +
            _by_line('CI',  ci_items)  +
            '  HP WITH JCP;\n'
        )
        _h_content = (
            f'TITLE:\n  CFA-H Cross-Wave 五因子 CFA\n'
            f'  HP/JCP/PP 取 T1；DP 取 T2；CI 取 T3\n'
            f'  Generated: {ts}\n\n'
            f'DATA:\n  FILE = "{cfa_h_dat_filename}";\n\n'
            f'VARIABLE:\n  NAMES =\n{names_block};\n'
            f'  USEVARIABLES =\n{_h_vars};\n'
            f'  MISSING =\n{_h_vars} (-999);\n\n'
            f'ANALYSIS:\n  ESTIMATOR = MLR;\n\n'
            f'MODEL:\n{_h_model}\n'
            f'OUTPUT:\n  STDYX;\n  MODINDICES(10);\n  CINTERVAL;\n'
        )
        _h_utf8, _ = save_inp_dual_encoding(_h_content, run_dir, f'CFA_H_CrossWave_{ts}')
        cfa_inp_list.append(('CFA-H (Cross-Wave 5F)', _h_utf8))

        # ---- CFA-I：跨波次四因子 CFA（HP/JCP@T1，DP@T2，CI@T3，不含 PP）----
        _i_vars = f'{_v_hp}\n{_v_jcp}\n{_v_dp}\n{_v_ci}'
        _i_model = (
            _by_line('HP',  hp_items)  +
            _by_line('JCP', jcp_items) +
            _by_line('DP',  dp_items)  +
            _by_line('CI',  ci_items)  +
            '  HP WITH JCP;\n'
        )
        _i_content = (
            f'TITLE:\n  CFA-I Cross-Wave 四因子 CFA（不含 PP）\n'
            f'  HP/JCP 取 T1；DP 取 T2；CI 取 T3\n'
            f'  Generated: {ts}\n\n'
            f'DATA:\n  FILE = "{cfa_h_dat_filename}";\n\n'
            f'VARIABLE:\n  NAMES =\n{names_block};\n'
            f'  USEVARIABLES =\n{_i_vars};\n'
            f'  MISSING =\n{_i_vars} (-999);\n\n'
            f'ANALYSIS:\n  ESTIMATOR = MLR;\n\n'
            f'MODEL:\n{_i_model}\n'
            f'OUTPUT:\n  STDYX;\n  MODINDICES(10);\n  CINTERVAL;\n'
        )
        _i_utf8, _ = save_inp_dual_encoding(_i_content, run_dir, f'CFA_I_CrossWave_4F_{ts}')
        cfa_inp_list.append(('CFA-I (Cross-Wave 4F, no PP)', _i_utf8))

    # ---- 調節中介路徑模型：PP 調節 JCP/HP(T1) → DP(T2) → CI(T3) ----
    # NAMES 欄位順序（含 PP_group）
    all_var_names = ('HP_T1  JCP_T1  PP_T1  DP_T1  CI_T1\n'
                     '    HP_T2  JCP_T2  PP_T2  DP_T2  CI_T2\n'
                     '    HP_T3  JCP_T3  PP_T3  DP_T3  CI_T3\n'
                     '    Gender Age Tenure Position PP_group EXP MAINT EXP_C MAINT_C')

    def make_path_model(ts, mplus_dat_filename):
        """
        完整調節中介路徑分析（Full Moderated Mediation）：
          PP 同時調節 a-path（JCP/HP->DP）與 b-path（DP->CI）
          H1a/b: a-path 主效果；H2a/b: a-path 調節；H3: b-path 主效果；
          H4: b-path 調節；H5a/b: 直接效果；H6a/b: 條件間接效果 (PP ±1SD)
        ML + Bootstrap 5000; MODEL CONSTRAINT for conditional indirect effects.
        """
        return (
            f'TITLE:\n  Full Moderated Mediation Path Model\n'
            f'  PP(T1) moderates a-path (JCP/HP->DP) and b-path (DP->CI)\n'
            f'  Generated: {ts}\n\n'
            f'DATA:\n  FILE = "{mplus_dat_filename}";\n\n'
            f'VARIABLE:\n  NAMES =\n    {all_var_names};\n'
            f'  USEVARIABLES =\n    JCP_T1  HP_T1  PP_T1\n'
            f'    DP_T2  CI_T3\n'
            f'    JCP_PP  HP_PP  DP_PP;\n'
            f'  MISSING = ALL(-999);\n\n'
            f'DEFINE:\n'
            f'  CENTER JCP_T1 HP_T1 PP_T1 DP_T2 (GRANDMEAN);\n'
            f'  JCP_PP = JCP_T1 * PP_T1;\n'
            f'  HP_PP  = HP_T1  * PP_T1;\n'
            f'  DP_PP  = DP_T2  * PP_T1;\n\n'
            f'ANALYSIS:\n  ESTIMATOR = ML;\n  BOOTSTRAP = 5000;\n\n'
            f'MODEL:\n'
            f'  ! a-paths: H1a/H1b (at mean PP) + H2a/H2b (PP moderation)\n'
            f'  DP_T2 ON JCP_T1 (a_jcp)\n'
            f'           HP_T1  (a_hp)\n'
            f'           PP_T1  (a_pp)\n'
            f'           JCP_PP (a_jcp_pp)\n'
            f'           HP_PP  (a_hp_pp);\n\n'
            f'  ! b-path: H3 (at mean PP) + H4 (PP moderation of b-path)\n'
            f'  CI_T3 ON DP_T2  (b_dp)\n'
            f'           DP_PP  (b_dp_pp);\n'
            f'  ! direct c-paths: H5a/H5b (at mean PP) + H6a/H6b (PP moderation of c-path)\n'
            f'  CI_T3 ON JCP_T1 (c_jcp)\n'
            f'           HP_T1  (c_hp)\n'
            f'           PP_T1  (c_pp)\n'
            f'           JCP_PP (c_jcp_pp)\n'
            f'           HP_PP  (c_hp_pp);\n\n'
            f'  ! T1 predictor covariances\n'
            f'  JCP_T1 WITH HP_T1 PP_T1;\n'
            f'  HP_T1  WITH PP_T1;\n\n'
            f'MODEL CONSTRAINT:\n'
            f'  NEW(ind_hi_j ind_lo_j ind_hi_h ind_lo_h\n'
            f'      dir_hi_j dir_lo_j dir_hi_h dir_lo_h);\n'
            f'  ! H7a: JCP -> DP -> CI conditional indirect at PP = +1SD / -1SD\n'
            f'  ind_hi_j = (a_jcp + a_jcp_pp * 1) * (b_dp + b_dp_pp * 1);\n'
            f'  ind_lo_j = (a_jcp - a_jcp_pp * 1) * (b_dp - b_dp_pp * 1);\n'
            f'  ! H7b: HP  -> DP -> CI conditional indirect at PP = +1SD / -1SD\n'
            f'  ind_hi_h = (a_hp  + a_hp_pp  * 1) * (b_dp + b_dp_pp * 1);\n'
            f'  ind_lo_h = (a_hp  - a_hp_pp  * 1) * (b_dp - b_dp_pp * 1);\n'
            f'  ! H6a/b supplement: conditional direct effect at PP = +1SD / -1SD\n'
            f'  dir_hi_j = c_jcp + c_jcp_pp * 1;\n'
            f'  dir_lo_j = c_jcp - c_jcp_pp * 1;\n'
            f'  dir_hi_h = c_hp  + c_hp_pp  * 1;\n'
            f'  dir_lo_h = c_hp  - c_hp_pp  * 1;\n\n'
            f'OUTPUT:\n  SAMPSTAT;  STDYX;  CINTERVAL(BCBOOTSTRAP);\n'
        )

    def make_baseline_mediation_model(ts, mplus_dat_filename):
        """
        Baseline mediation (no moderation): JCP/HP(T1) -> DP(T2) -> CI(T3)
        PP as control only. ML + Bootstrap 5000.
        """
        return (
            f'TITLE:\n  Baseline Mediation Model (no moderation)\n'
            f'  JCP/HP(T1) -> DP(T2) -> CI(T3), PP as control\n'
            f'  Generated: {ts}\n\n'
            f'DATA:\n  FILE = "{mplus_dat_filename}";\n\n'
            f'VARIABLE:\n  NAMES =\n    {all_var_names};\n'
            f'  USEVARIABLES =\n    JCP_T1  HP_T1  PP_T1  DP_T2  CI_T3;\n'
            f'  MISSING = ALL(-999);\n\n'
            f'DEFINE:\n'
            f'  CENTER JCP_T1 HP_T1 PP_T1 DP_T2 (GRANDMEAN);\n\n'
            f'ANALYSIS:\n  ESTIMATOR = ML;\n  BOOTSTRAP = 5000;\n\n'
            f'MODEL:\n'
            f'  DP_T2 ON JCP_T1 (a_jcp)\n'
            f'           HP_T1  (a_hp)\n'
            f'           PP_T1;\n\n'
            f'  CI_T3 ON DP_T2  (b_dp)\n'
            f'           JCP_T1 (c_jcp)\n'
            f'           HP_T1  (c_hp)\n'
            f'           PP_T1;\n\n'
            f'  JCP_T1 WITH HP_T1 PP_T1;\n'
            f'  HP_T1  WITH PP_T1;\n\n'
            f'MODEL CONSTRAINT:\n'
            f'  NEW(ind_jcp ind_hp);\n'
            f'  ind_jcp = a_jcp * b_dp;\n'
            f'  ind_hp  = a_hp  * b_dp;\n\n'
            f'OUTPUT:\n  SAMPSTAT;  STDYX;  CINTERVAL(BCBOOTSTRAP);\n'
        )

    def make_nopp_mediation_model(ts, mplus_dat_filename):
        """純中介路徑（不含 PP）：JCP/HP(T1) → DP(T2) → CI(T3)，Bootstrap 間接效果"""
        return (
            f'TITLE:\n  Pure Mediation Model (No PP)\n'
            f'  JCP/HP(T1) -> DP(T2) -> CI(T3), PP excluded entirely\n'
            f'  Generated: {ts}\n\n'
            f'DATA:\n  FILE = "{mplus_dat_filename}";\n\n'
            f'VARIABLE:\n  NAMES =\n    {all_var_names};\n'
            f'  USEVARIABLES =\n    JCP_T1  HP_T1  DP_T2  CI_T3;\n'
            f'  MISSING = ALL(-999);\n\n'
            f'DEFINE:\n'
            f'  CENTER JCP_T1 HP_T1 DP_T2 (GRANDMEAN);\n\n'
            f'ANALYSIS:\n  ESTIMATOR = ML;\n  BOOTSTRAP = 5000;\n\n'
            f'MODEL:\n'
            f'  DP_T2 ON JCP_T1 (a_jcp)\n'
            f'           HP_T1  (a_hp);\n\n'
            f'  CI_T3 ON DP_T2  (b_dp)\n'
            f'           JCP_T1 (c_jcp)\n'
            f'           HP_T1  (c_hp);\n\n'
            f'  JCP_T1 WITH HP_T1;\n\n'
            f'MODEL CONSTRAINT:\n'
            f'  NEW(ind_jcp ind_hp);\n'
            f'  ind_jcp = a_jcp * b_dp;\n'
            f'  ind_hp  = a_hp  * b_dp;\n\n'
            f'OUTPUT:\n  SAMPSTAT;  STDYX;  CINTERVAL(BCBOOTSTRAP);\n'
        )

    def make_jcp_only_model(ts, mplus_dat_filename):
        """
        JCP-only mediation (without HP): test whether JCP has independent effect
        when HP is removed from the model. ML + Bootstrap 5000.
        """
        return (
            f'TITLE:\n  JCP-only Mediation Model (HP excluded, multicollinearity check)\n'
            f'  JCP(T1) -> DP(T2) -> CI(T3), PP as control\n'
            f'  Generated: {ts}\n\n'
            f'DATA:\n  FILE = "{mplus_dat_filename}";\n\n'
            f'VARIABLE:\n  NAMES =\n    {all_var_names};\n'
            f'  USEVARIABLES =\n    JCP_T1  PP_T1  DP_T2  CI_T3;\n'
            f'  MISSING = ALL(-999);\n\n'
            f'DEFINE:\n'
            f'  CENTER JCP_T1 PP_T1 DP_T2 (GRANDMEAN);\n\n'
            f'ANALYSIS:\n  ESTIMATOR = ML;\n  BOOTSTRAP = 5000;\n\n'
            f'MODEL:\n'
            f'  DP_T2 ON JCP_T1 (a_jcp)\n'
            f'           PP_T1;\n\n'
            f'  CI_T3 ON DP_T2  (b_dp)\n'
            f'           JCP_T1 (c_jcp)\n'
            f'           PP_T1;\n\n'
            f'  JCP_T1 WITH PP_T1;\n\n'
            f'MODEL CONSTRAINT:\n'
            f'  NEW(ind_jcp);\n'
            f'  ind_jcp = a_jcp * b_dp;\n\n'
            f'OUTPUT:\n  SAMPSTAT;  STDYX;  CINTERVAL(BCBOOTSTRAP);\n'
        )

    def make_jcp_pp_model(ts, mplus_dat_filename):
        """
        JCP-only moderated mediation: PP moderates both a-path (JCP->DP) and
        b-path (DP->CI), without HP in the model.
        Confirms whether the low-PP conditional indirect (JCP->DP->CI) holds
        independently of HP. ML + Bootstrap 5000.
        """
        return (
            f'TITLE:\n  JCP-only Moderated Mediation (PP moderates a and b paths, no HP)\n'
            f'  JCP(T1)->DP(T2)->CI(T3), PP(T1) moderator\n'
            f'  Generated: {ts}\n\n'
            f'DATA:\n  FILE = "{mplus_dat_filename}";\n\n'
            f'VARIABLE:\n  NAMES =\n    {all_var_names};\n'
            f'  USEVARIABLES =\n    JCP_T1  PP_T1  DP_T2  CI_T3\n'
            f'    JCP_PP  DP_PP;\n'
            f'  MISSING = ALL(-999);\n\n'
            f'DEFINE:\n'
            f'  CENTER JCP_T1 PP_T1 DP_T2 (GRANDMEAN);\n'
            f'  JCP_PP = JCP_T1 * PP_T1;\n'
            f'  DP_PP  = DP_T2  * PP_T1;\n\n'
            f'ANALYSIS:\n  ESTIMATOR = ML;\n  BOOTSTRAP = 5000;\n\n'
            f'MODEL:\n'
            f'  ! a-path: JCP->DP, moderated by PP\n'
            f'  DP_T2 ON JCP_T1 (a_jcp)\n'
            f'           PP_T1\n'
            f'           JCP_PP  (a_pp);\n\n'
            f'  ! b-path + direct c-path: PP also moderates b-path\n'
            f'  CI_T3 ON DP_T2  (b_dp)\n'
            f'           JCP_T1 (c_jcp)\n'
            f'           PP_T1\n'
            f'           DP_PP  (b_pp);\n\n'
            f'  JCP_T1 WITH PP_T1;\n\n'
            f'MODEL CONSTRAINT:\n'
            f'  NEW(ind_hi ind_lo);\n'
            f'  ! Conditional indirect at PP +1SD and -1SD\n'
            f'  ind_hi = (a_jcp + a_pp*1)    * (b_dp + b_pp*1);\n'
            f'  ind_lo = (a_jcp + a_pp*(-1)) * (b_dp + b_pp*(-1));\n\n'
            f'OUTPUT:\n  STDYX;  CINTERVAL(BCBOOTSTRAP);\n'
        )

    def make_career_stage_model(ts, mplus_dat_filename):
        """
        Bayesian Path Analysis: Career Stage Moderation of a-path (JCP/HP -> DP)
        Effects (deviation) coding: EXP_C (+1 for EXP, -0.5 for ESTAB/MAINT)
                                    MAINT_C (+1 for MAINT, -0.5 for EXP/ESTAB)
        Each coefficient = that stage vs. the mean of the other two stages.
        PP retained as control variable only (not moderator).
        """
        return (
            f'TITLE:\n  Career Stage Moderated Path Analysis (Bayesian, effects coding)\n'
            f'  EXP_C: +1=exploration(21-30) -0.5=others\n'
            f'  MAINT_C: +1=maintenance(41+) -0.5=others\n'
            f'  Generated: {ts}\n\n'
            f'DATA:\n  FILE = "{mplus_dat_filename}";\n\n'
            f'VARIABLE:\n  NAMES =\n    {all_var_names};\n'
            f'  USEVARIABLES =\n'
            f'    JCP_T1  HP_T1  PP_T1  DP_T2  CI_T3\n'
            f'    EXP_C  MAINT_C\n'
            f'    JCP_EXC JCP_MNC HP_EXC HP_MNC;\n'
            f'  MISSING = ALL(-999);\n\n'
            f'DEFINE:\n'
            f'  CENTER JCP_T1 HP_T1 DP_T2 (GRANDMEAN);\n'
            f'  JCP_EXC = JCP_T1 * EXP_C;\n'
            f'  JCP_MNC = JCP_T1 * MAINT_C;\n'
            f'  HP_EXC  = HP_T1  * EXP_C;\n'
            f'  HP_MNC  = HP_T1  * MAINT_C;\n\n'
            f'ANALYSIS:\n'
            f'  ESTIMATOR = BAYES;\n'
            f'  BITERATIONS = (10000);\n'
            f'  CHAINS = 2;\n'
            f'  BCONVERGENCE = .05;\n\n'
            f'MODEL:\n'
            f'  ! a-path: JCP/HP -> DP, career stage (effects-coded) moderates\n'
            f'  DP_T2 ON JCP_T1  (a_jcp)\n'
            f'           HP_T1   (a_hp)\n'
            f'           PP_T1\n'
            f'           EXP_C  MAINT_C\n'
            f'           JCP_EXC (a_jce)\n'
            f'           JCP_MNC (a_jcm)\n'
            f'           HP_EXC  (a_hpe)\n'
            f'           HP_MNC  (a_hpm);\n\n'
            f'  ! b-path + direct c-path (career stage as control)\n'
            f'  CI_T3 ON DP_T2  (b_dp)\n'
            f'           JCP_T1 (c_jcp)\n'
            f'           HP_T1  (c_hp)\n'
            f'           PP_T1\n'
            f'           EXP_C  MAINT_C;\n\n'
            f'  JCP_T1 WITH HP_T1 PP_T1;\n'
            f'  HP_T1  WITH PP_T1;\n\n'
            f'MODEL CONSTRAINT:\n'
            f'  NEW(ie_e_j ie_m_j ie_r_j\n'
            f'      ie_e_h ie_m_h ie_r_h);\n'
            f'  ! Effects coding: EXP_C=+1 for EXP, MAINT_C=+1 for MAINT, both=-0.5 for ESTAB\n'
            f'  ! Conditional a-path (JCP) per stage = a_jcp + a_jce*EXP_C + a_jcm*MAINT_C\n'
            f'  ie_e_j = (a_jcp + a_jce*1    + a_jcm*(-0.5)) * b_dp;  ! EXP  vs other-two mean\n'
            f'  ie_m_j = (a_jcp + a_jce*(-0.5) + a_jcm*1   ) * b_dp;  ! MAINT vs other-two mean\n'
            f'  ie_r_j = (a_jcp + a_jce*(-0.5) + a_jcm*(-0.5)) * b_dp; ! ESTAB vs other-two mean\n'
            f'  ! Conditional a-path (HP) per stage = a_hp + a_hpe*EXP_C + a_hpm*MAINT_C\n'
            f'  ie_e_h = (a_hp  + a_hpe*1    + a_hpm*(-0.5)) * b_dp;  ! EXP\n'
            f'  ie_m_h = (a_hp  + a_hpe*(-0.5) + a_hpm*1   ) * b_dp;  ! MAINT\n'
            f'  ie_r_h = (a_hp  + a_hpe*(-0.5) + a_hpm*(-0.5)) * b_dp; ! ESTAB\n\n'
            f'OUTPUT:\n  STDYX;  CINTERVAL(HPD);\n'
        )

    # [PP 調節中介已停用] 生成調節中介路徑模型 .inp（保留函式供日後參考，不執行）
    # _path_fname = f'PATH_ModMed_{ts}'
    # _path_content = make_path_model(ts, mplus_dat_filename)
    # _path_utf8, _ = save_inp_dual_encoding(_path_content, run_dir, _path_fname)

    # 生成純中介模型（主模型：JCP/HP(T1)→DP(T2)→CI(T3)，不含 PP）
    _nopp_fname = f'PATH_NoPP_{ts}'
    _nopp_content = make_nopp_mediation_model(ts, mplus_dat_filename)
    _nopp_utf8, _ = save_inp_dual_encoding(_nopp_content, run_dir, _nopp_fname)

    # 生成基礎中介模型（無調節）.inp
    _base_fname = f'PATH_Baseline_{ts}'
    _base_content = make_baseline_mediation_model(ts, mplus_dat_filename)
    _base_utf8, _ = save_inp_dual_encoding(_base_content, run_dir, _base_fname)

    # 生成 JCP-only 模型（排除 HP，multicollinearity 確認）
    _jcp_fname = f'PATH_JCP_only_{ts}'
    _jcp_content = make_jcp_only_model(ts, mplus_dat_filename)
    _jcp_utf8, _ = save_inp_dual_encoding(_jcp_content, run_dir, _jcp_fname)

    # [PP 調節中介已停用] JCP-only + PP 調節模型
    # _jcp_pp_fname = f'PATH_JCP_PP_{ts}'
    # _jcp_pp_content = make_jcp_pp_model(ts, mplus_dat_filename)
    # _jcp_pp_utf8, _ = save_inp_dual_encoding(_jcp_pp_content, run_dir, _jcp_pp_fname)

    # 生成職涯階段 Bayesian 路徑模型 .inp（效果編碼）
    _cs_fname = f'PATH_CareerStage_{ts}'
    _cs_content = make_career_stage_model(ts, mplus_dat_filename)
    _cs_utf8, _ = save_inp_dual_encoding(_cs_content, run_dir, _cs_fname)

    path_inp_list = [
        ('PATH_NoPP',        _nopp_utf8),
        ('PATH_Baseline',    _base_utf8),
        ('PATH_JCP_only',    _jcp_utf8),
        ('PATH_CareerStage', _cs_utf8),
    ]

    # ---- 自動執行 Mplus ----
    if 'cfa' in phases:
        print("\n[Mplus] 自動執行 CFA 模型...")
        cfa_run_results = run_all_mplus(cfa_inp_list, mplus_exe)
    else:
        cfa_run_results = []

    print("[Mplus] 自動執行縱貫中介路徑模型（JCP/HP(T1)→DP(T2)→CI(T3)）...")
    path_run_results = run_all_mplus(path_inp_list, mplus_exe)

    # ---- 解析 CFA 結果 ----
    for label, ok, out_path, err in cfa_run_results:
        if ok:
            entry = {'fit': parse_mplus_fit(out_path), 'out': out_path}
            entry['loadings'] = parse_mplus_cfa_loadings(out_path)
            entry['ave_cr'] = calculate_ave_cr(out_path)
            all_results[label] = entry

    # ---- 解析路徑模型結果 ----
    _path_map = {
        'H1a: JCP(T1)→DP(T2) [at mean PP]': ('DP_T2', 'JCP_T1'),
        'H1b: HP(T1)→DP(T2) [at mean PP]':  ('DP_T2', 'HP_T1'),
        'H2a: PP×JCP→DP (moderation)':       ('DP_T2', 'JCP_PP'),
        'H2b: PP×HP→DP (moderation)':        ('DP_T2', 'HP_PP'),
        'H3:  DP(T2)→CI(T3) [at mean PP]':  ('CI_T3', 'DP_T2'),
        'H4:  PP×DP→CI (moderation)':        ('CI_T3', 'DP_PP'),
        'H5a: JCP(T1)→CI(T3) [at mean PP]': ('CI_T3', 'JCP_T1'),
        'H5b: HP(T1)→CI(T3) [at mean PP]':  ('CI_T3', 'HP_T1'),
        'H6a: PP×JCP→CI (moderation)':       ('CI_T3', 'JCP_PP'),
        'H6b: PP×HP→CI (moderation)':        ('CI_T3', 'HP_PP'),
        'PP→DP (main)':                      ('DP_T2', 'PP_T1'),
        'PP→CI (main)':                      ('CI_T3', 'PP_T1'),
    }
    _modconstraint_params = [
        'IND_HI_J', 'IND_LO_J', 'IND_HI_H', 'IND_LO_H',
        'DIR_HI_J', 'DIR_LO_J', 'DIR_HI_H', 'DIR_LO_H',
    ]

    _nopp_path_map = {
        'H1a: JCP(T1)→DP(T2)':      ('DP_T2', 'JCP_T1'),
        'H1b: HP(T1)→DP(T2)':       ('DP_T2', 'HP_T1'),
        'H2: DP(T2)→CI(T3)':        ('CI_T3', 'DP_T2'),
        'H3a: JCP(T1)→CI(T3) 直接': ('CI_T3', 'JCP_T1'),
        'H3b: HP(T1)→CI(T3) 直接':  ('CI_T3', 'HP_T1'),
    }
    _nopp_mc_params = ['IND_JCP', 'IND_HP']

    _base_path_map = {
        'JCP(T1)→DP(T2)':        ('DP_T2', 'JCP_T1'),
        'HP(T1)→DP(T2)':         ('DP_T2', 'HP_T1'),
        'PP→DP (控制)':           ('DP_T2', 'PP_T1'),
        'DP(T2)→CI(T3)':         ('CI_T3', 'DP_T2'),
        'JCP(T1)→CI(T3) 直接':   ('CI_T3', 'JCP_T1'),
        'HP(T1)→CI(T3) 直接':    ('CI_T3', 'HP_T1'),
        'PP→CI (控制)':           ('CI_T3', 'PP_T1'),
    }
    _base_mc_params = ['IND_JCP', 'IND_HP']

    _cs_path_map = {
        'JCP(T1)→DP(T2) [ESTAB均值]':   ('DP_T2', 'JCP_T1'),
        'HP(T1)→DP(T2) [ESTAB均值]':    ('DP_T2', 'HP_T1'),
        'JCP×EXP_C→DP (探索期效果)':     ('DP_T2', 'JCP_EXC'),
        'JCP×MAINT_C→DP (維持期效果)':   ('DP_T2', 'JCP_MNC'),
        'HP×EXP_C→DP (探索期效果)':      ('DP_T2', 'HP_EXC'),
        'HP×MAINT_C→DP (維持期效果)':    ('DP_T2', 'HP_MNC'),
        'DP(T2)→CI(T3)':               ('CI_T3', 'DP_T2'),
        'JCP(T1)→CI(T3) 直接':          ('CI_T3', 'JCP_T1'),
        'HP(T1)→CI(T3) 直接':           ('CI_T3', 'HP_T1'),
        'EXP_C→DP (控制)':              ('DP_T2', 'EXP_C'),
        'MAINT_C→DP (控制)':            ('DP_T2', 'MAINT_C'),
        'PP→DP (控制)':                 ('DP_T2', 'PP_T1'),
    }
    _cs_mc_params = ['IE_E_J', 'IE_M_J', 'IE_R_J', 'IE_E_H', 'IE_M_H', 'IE_R_H']

    _jcp_path_map = {
        'JCP(T1)→DP(T2) [JCP-only]': ('DP_T2', 'JCP_T1'),
        'PP→DP (控制)':               ('DP_T2', 'PP_T1'),
        'DP(T2)→CI(T3)':             ('CI_T3', 'DP_T2'),
        'JCP(T1)→CI(T3) 直接':       ('CI_T3', 'JCP_T1'),
        'PP→CI (控制)':               ('CI_T3', 'PP_T1'),
    }
    _jcp_mc_params = ['IND_JCP']

    _jcp_pp_path_map = {
        'JCP(T1)→DP(T2) [at mean PP]': ('DP_T2', 'JCP_T1'),
        'JCP×PP→DP (a-path 調節)':      ('DP_T2', 'JCP_PP'),
        'DP(T2)→CI(T3) [at mean PP]':  ('CI_T3', 'DP_T2'),
        'DP×PP→CI (b-path 調節)':       ('CI_T3', 'DP_PP'),
        'JCP(T1)→CI(T3) 直接':         ('CI_T3', 'JCP_T1'),
        'PP→DP (控制)':                 ('DP_T2', 'PP_T1'),
        'PP→CI (控制)':                 ('CI_T3', 'PP_T1'),
    }
    _jcp_pp_mc_params = ['IND_HI', 'IND_LO']

    for label, ok, out_path, err in path_run_results:
        if label == 'PATH_NoPP':
            if ok:
                all_results[label] = {
                    'fit':       parse_mplus_fit(out_path),
                    'paths':     parse_mplus_stdyx(out_path, _nopp_path_map),
                    'modconstr': parse_mplus_modconstraint(out_path, _nopp_mc_params),
                    'out':       out_path
                }
        elif label == 'PATH (T1→T2→T3)':
            if ok:
                all_results[label] = {
                    'fit':       parse_mplus_fit(out_path),
                    'paths':     parse_mplus_stdyx(out_path, _path_map),
                    'modconstr': parse_mplus_modconstraint(out_path, _modconstraint_params),
                    'out':       out_path
                }
        elif label == 'PATH_Baseline':
            if ok:
                all_results[label] = {
                    'fit':       parse_mplus_fit(out_path),
                    'paths':     parse_mplus_stdyx(out_path, _base_path_map),
                    'modconstr': parse_mplus_modconstraint(out_path, _base_mc_params),
                    'out':       out_path
                }
        elif label == 'PATH_JCP_only':
            if ok:
                all_results[label] = {
                    'fit':       parse_mplus_fit(out_path),
                    'paths':     parse_mplus_stdyx(out_path, _jcp_path_map),
                    'modconstr': parse_mplus_modconstraint(out_path, _jcp_mc_params),
                    'out':       out_path
                }
        elif label == 'PATH_JCP_PP':
            if ok:
                all_results[label] = {
                    'fit':       parse_mplus_fit(out_path),
                    'paths':     parse_mplus_stdyx(out_path, _jcp_pp_path_map),
                    'modconstr': parse_mplus_modconstraint(out_path, _jcp_pp_mc_params),
                    'out':       out_path
                }
        elif label == 'PATH_CareerStage':
            if ok:
                all_results[label] = {
                    'fit':       parse_mplus_fit(out_path),
                    'paths':     parse_mplus_bayes_paths(out_path, _cs_path_map),
                    'modconstr': parse_mplus_bayes_modconstraint(out_path, _cs_mc_params),
                    'out':       out_path
                }

    return all_results, cfa_inp_list + path_inp_list


# ==========================================
# MODULE F: Excel 綜合報告產生
# ==========================================
def generate_excel_report(run_dir, ts, g3_sample, alpha_dict, corr_dict, all_results,
                          variant_label=None, exclude=None,
                          clpm_results=None):
    """
    產生 Excel 綜合報告 (Thesis_Results_YYYYMMDD_HHMM.xlsx)，含：
      Sheet 1: 樣本背景變項描述統計
      Sheet 2: 績效考核分析（三波）
      Sheet 3: 各量表題目 Item-level 描述統計 + 刪題建議
      Sheet 4: 各量表各波次敘述統計 + 信度 (Cronbach's α)
      Sheet 5: 三波段追蹤相關矩陣（完整 15×15，對角線為 α，附 M/SD）
      Sheet 6: CFA 適配指數 + 因素負荷量 + AVE/CR + 測量不變性（MI）
      Sheet 7: SPSS 語法（.sps 內容）
      Sheet 8: Mplus 語法（CFA + MI .inp 內容）
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [警告] 未安裝 openpyxl，跳過 Excel 報告。請執行: pip install openpyxl")
        return None

    wb = openpyxl.Workbook()
    df = g3_sample.copy()
    n_total = len(df)
    variant_note = f'【敏感性分析：{variant_label}】' if variant_label else ''

    # ── 共用樣式 ──────────────────────────────────────────────────
    thin = Side(border_style="thin", color="000000")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", fgColor="4472C4")
    sub_fill  = PatternFill("solid", fgColor="D9E1F2")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hdr_fill
        c.alignment = ctr
        c.border = bdr
        return c

    def cell(ws, row, col, val, bold=False, align='center', color=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, size=10, color=color if color else "000000")
        c.alignment = ctr if align == 'center' else lft
        c.border = bdr
        return c

    def title(ws, text, end_col=8):
        ws.cell(row=1, column=1, value=text).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)

    def set_widths(ws, pairs):
        for col_letter, w in pairs:
            ws.column_dimensions[col_letter].width = w

    # ── Sheet 1: 樣本背景變項 ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "1_背景變項"
    title(ws1, f"樣本背景變項描述統計（N = {n_total}，三波配對樣本）{variant_note}", end_col=5)
    r = 3
    for ci, h in enumerate(["變項", "類別 / 統計量", "人數 / 數值", "%", "單位"], 1):
        hdr(ws1, r, ci, h)
    r += 1

    # 連續變項
    for col_name, label, unit in [
        ('Age',           '年齡',       '歲'),
        ('NowJobTenure',  '現職年資',   '月'),
        ('JobTenure',     '整體工作年資', '月'),
    ]:
        if col_name not in df.columns:
            continue
        s = pd.to_numeric(df[col_name], errors='coerce').dropna()
        cell(ws1, r, 1, label, bold=True, align='left')
        cell(ws1, r, 2, 'M (SD)')
        cell(ws1, r, 3, f"{s.mean():.1f} ({s.std():.1f})")
        cell(ws1, r, 4, '')
        cell(ws1, r, 5, unit)
        r += 1
        cell(ws1, r, 1, '')
        cell(ws1, r, 2, '最小值 ~ 最大值')
        cell(ws1, r, 3, f"{s.min():.0f} ~ {s.max():.0f}")
        cell(ws1, r, 4, '')
        cell(ws1, r, 5, '')
        r += 1
        # 年齡區間分布（僅 Age 欄）
        if col_name == 'Age':
            age_bins = [
                ('21–30 歲', (s >= 21) & (s <= 30)),
                ('31–40 歲', (s >= 31) & (s <= 40)),
                ('41 歲以上', s >= 41),
            ]
            for grp_label, mask in age_bins:
                cnt = int(mask.sum())
                pct = cnt / n_total * 100 if n_total > 0 else 0
                cell(ws1, r, 1, '')
                cell(ws1, r, 2, grp_label, align='left')
                cell(ws1, r, 3, cnt)
                cell(ws1, r, 4, f"{pct:.1f}%")
                cell(ws1, r, 5, '')
                r += 1

    # 類別變項
    cat_vars = [
        ('Gender',    '性別',
         {1: '男', 2: '女', 3: '其他'}),
        ('Education', '教育程度',
         {1: '高中(職)', 2: '專科', 3: '大學', 4: '碩士', 5: '博士'}),
        ('Marriage',  '婚姻狀況',
         {1: '未婚', 2: '已婚無子女', 3: '已婚有子女', 4: '其他'}),
        ('Position',  '職位層級',
         {1: '一般員工', 2: '基層主管', 3: '中階主管', 4: '高階主管'}),
        ('Industry',  '所屬產業',
         {1: '製造', 2: '科技', 3: '金融', 4: '服務',
          5: '醫療', 6: '教育', 7: '公部門', 8: '其他'}),
    ]
    for col_name, label, val_map in cat_vars:
        if col_name not in df.columns:
            continue
        s = pd.to_numeric(df[col_name], errors='coerce').dropna()
        cell(ws1, r, 1, label, bold=True, align='left')
        for ci in range(2, 6):
            cell(ws1, r, ci, '')
        r += 1
        for v, v_label in val_map.items():
            cnt = int((s == v).sum())
            pct = cnt / n_total * 100 if n_total > 0 else 0
            cell(ws1, r, 1, '')
            cell(ws1, r, 2, v_label, align='left')
            cell(ws1, r, 3, cnt)
            cell(ws1, r, 4, f"{pct:.1f}%")
            cell(ws1, r, 5, '')
            r += 1

    set_widths(ws1, [('A', 18), ('B', 20), ('C', 16), ('D', 10), ('E', 12)])

    # ── Sheet 3: 敘述統計 + 信度（各量表 × 各波次）─────────────────
    ws2 = wb.create_sheet("3_敘述統計與信度")
    title(ws2, f"各量表各波次敘述統計與信度（N = {n_total}，三波配對樣本）", end_col=12)
    r = 3
    hdr_cols = ["量表", "說明", "題數",
                "T1 M", "T1 SD", "T1 α",
                "T2 M", "T2 SD", "T2 α",
                "T3 M", "T3 SD", "T3 α"]
    for ci, h in enumerate(hdr_cols, 1):
        hdr(ws2, r, ci, h)
    r += 1

    scale_defs = [
        ('HP',  'HP 階層停滯（CP 次量表）',        [f'HP{i}'  for i in range(1, 7)]),
        ('JCP', 'JCP 工作內容停滯（CP 次量表）',   [f'JCP{i}' for i in range(1, 7)]),
        ('CP',  'CP 職涯高原（HP+JCP 合併）',       [f'HP{i}'  for i in range(1, 7)] + [f'JCP{i}' for i in range(1, 7)]),
        ('PP',  'PP 主動型人格',                    [f'PP{i}'  for i in range(1, 7)]),
        ('DP',  'DP 決策拖延',                      [f'DP{i}'  for i in range(1, 6)]),
        ('CI',  'CI 職涯無所作為',                  [f'CI{i}'  for i in range(1, 9)]),
    ]
    for s_name, s_label, s_items in scale_defs:
        row_data = [s_name, s_label, len(s_items)]
        for wave in ['T1', 'T2', 'T3']:
            wave_cols = [f'{item}_{wave}' for item in s_items]
            valid_cols = [c for c in wave_cols if c in df.columns]
            if valid_cols:
                mat = df[valid_cols].apply(pd.to_numeric, errors='coerce')
                means = mat.mean(axis=1)
                m  = means.mean()
                sd = means.std()
                alpha_v = calculate_cronbach_alpha(mat.dropna())
                row_data += [
                    f"{m:.2f}",
                    f"{sd:.2f}",
                    f"{alpha_v:.3f}" if not np.isnan(alpha_v) else 'N/A'
                ]
            else:
                row_data += ['N/A', 'N/A', 'N/A']
        for ci, val in enumerate(row_data, 1):
            cell(ws2, r, ci, val, bold=(ci == 1),
                 align='left' if ci <= 2 else 'center')
        r += 1

    set_widths(ws2, [('A', 12), ('B', 30), ('C', 8),
                     ('D', 11), ('E', 11), ('F', 11),
                     ('G', 11), ('H', 11), ('I', 11),
                     ('J', 11), ('K', 11), ('L', 11)])

    # ── 相關矩陣 helper（供 T1/T2/T3 各波次共用）─────────────────
    scale_n_items = {'HP': 6, 'JCP': 6, 'CP': 12, 'PP': 6, 'DP': 5, 'CI': 8}
    # 次量表版（HP/JCP 分開）與 CP 合併版各自的變數順序
    scale_order_sub = ['HP', 'JCP', 'PP', 'DP', 'CI']   # 用於分析用相關矩陣
    scale_order_cp  = ['CP', 'PP', 'DP', 'CI']           # 用於 CP 合併相關矩陣

    def _build_wave_means(sn_list, wave):
        """計算指定波次中各量表的受試者平均分數。"""
        wave_means = {}
        for sn in sn_list:
            if sn == 'CP':
                cp_cols = ([f'HP{i}_{wave}' for i in range(1, 7)] +
                           [f'JCP{i}_{wave}' for i in range(1, 7)])
                valid = [c for c in cp_cols if c in df.columns]
            else:
                n_items = scale_n_items[sn]
                cols = [f'{sn}{i}_{wave}' for i in range(1, n_items + 1)]
                valid = [c for c in cols if c in df.columns]
            if valid:
                wave_means[sn] = df[valid].apply(pd.to_numeric, errors='coerce').mean(axis=1)
        return wave_means

    def make_corr_block(ws, wave, sn_list, subtitle_text, start_row):
        """在指定起始列寫入一個相關矩陣區塊，回傳下一個可用列號。"""
        n_scales = len(sn_list)
        wave_means = _build_wave_means(sn_list, wave)
        r = start_row

        # 小標題
        sub = ws.cell(row=r, column=1, value=subtitle_text)
        sub.font = Font(bold=True, size=10)
        sub.fill = PatternFill("solid", fgColor="D9E1F2")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_scales + 3)
        r += 1

        # 欄標題
        hdr(ws, r, 1, '變數')
        for ci, sn in enumerate(sn_list, 2):
            hdr(ws, r, ci, sn)
        hdr(ws, r, n_scales + 2, 'M')
        hdr(ws, r, n_scales + 3, 'SD')
        r += 1

        for i, sn_r in enumerate(sn_list):
            cell(ws, r, 1, sn_r, bold=True)
            for j, sn_c in enumerate(sn_list):
                col = j + 2
                if i == j:
                    cell(ws, r, col, '—')
                elif i < j and sn_r in wave_means and sn_c in wave_means:
                    valid_dat = pd.concat(
                        [wave_means[sn_r], wave_means[sn_c]], axis=1).dropna()
                    if len(valid_dat) > 2:
                        rv, pv = stats.pearsonr(
                            valid_dat.iloc[:, 0], valid_dat.iloc[:, 1])
                        star, _ = fmt_p(pv)
                        cell(ws, r, col, f"{rv:.2f}{star}")
                    else:
                        cell(ws, r, col, '—')
                else:
                    cell(ws, r, col, '')
            if sn_r in wave_means:
                m  = wave_means[sn_r].mean()
                sd = wave_means[sn_r].std()
                cell(ws, r, n_scales + 2, f"{m:.2f}")
                cell(ws, r, n_scales + 3, f"{sd:.2f}")
            r += 1

        return r + 1  # 空一列再接下一個區塊

    def make_corr_sheet(ws, wave):
        """在一張 Sheet 內上下疊放次量表版與 CP 合併版兩張相關矩陣。"""
        title(ws, f"相關分析矩陣（{wave}，N = {n_total}）  *** p<.001  ** p<.01  * p<.05",
              end_col=8)
        ws.cell(row=2, column=1,
                value="上表：次量表版（HP/JCP 分開，與 Mplus parceling 一致）；"
                      "下表：CP 合併版（HP+JCP 合一，供整體描述統計參考）"
                ).font = Font(italic=True, size=9)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

        next_r = make_corr_block(ws, wave, scale_order_sub,
                                 '表 A：次量表相關矩陣（HP / JCP / PP / DP / CI）', 3)
        make_corr_block(ws, wave, scale_order_cp,
                        '表 B：CP 合併相關矩陣（CP / PP / DP / CI）', next_r)

        set_widths(ws, [('A', 8), ('B', 10), ('C', 10), ('D', 10),
                        ('E', 10), ('F', 10), ('G', 8), ('H', 8)])

    # ── Sheet 4: 相關矩陣（上：12×12 無PP；下：15×15 完整）──────────────
    ws_corr = wb.create_sheet("4_相關矩陣")

    # helper：依 _iv_defs 建立合成分數與 α
    def _build_iv(defs):
        composites, alphas = [], []
        for _sn, _wv, _lbl, _ni in defs:
            _cols = [f'{_sn}{_k}_{_wv}' for _k in range(1, _ni + 1)]
            _vcols = [c for c in _cols if c in df.columns]
            if _vcols:
                _mat = df[_vcols].apply(pd.to_numeric, errors='coerce')
                composites.append(_mat.mean(axis=1))
                alphas.append(calculate_cronbach_alpha(_mat.dropna()))
            else:
                composites.append(pd.Series([np.nan] * len(df), index=df.index))
                alphas.append(np.nan)
        return composites, alphas

    # helper：畫下三角相關矩陣區塊，從 start_row 開始，回傳下一個可用 row
    def _draw_corr_block(ws, iv_defs, composites, alphas, start_row, title_str, note_str):
        _n = len(iv_defs)
        _end_col = _n + 2
        _wf = {
            'T1': PatternFill("solid", fgColor="E2EFDA"),
            'T2': PatternFill("solid", fgColor="DDEBF7"),
            'T3': PatternFill("solid", fgColor="FCE4D6"),
        }
        _r = start_row
        _tc = ws.cell(row=_r, column=1, value=title_str)
        _tc.font = Font(bold=True, size=12, color="FFFFFF")
        _tc.fill = hdr_fill; _tc.alignment = ctr
        ws.merge_cells(start_row=_r, start_column=1, end_row=_r, end_column=_end_col)
        _r += 1
        _nc = ws.cell(row=_r, column=1, value=note_str)
        _nc.font = Font(italic=True, size=9)
        ws.merge_cells(start_row=_r, start_column=1, end_row=_r, end_column=_end_col)
        _r += 1
        _h0 = ws.cell(row=_r, column=1, value='變數')
        _h0.font = Font(bold=True, color="FFFFFF", size=10)
        _h0.fill = hdr_fill; _h0.alignment = ctr; _h0.border = bdr
        for _jj in range(_n):
            _hc = ws.cell(row=_r, column=_jj + 2, value=str(_jj + 1))
            _hc.font = Font(bold=True, color="FFFFFF", size=10)
            _hc.fill = hdr_fill; _hc.alignment = ctr; _hc.border = bdr
        _r += 1
        _data_start = _r
        for _ii, (_sn, _wv, _lbl, _ni) in enumerate(iv_defs):
            _row = _data_start + _ii
            _rc = ws.cell(row=_row, column=1, value=f"({_ii + 1}) {_lbl}")
            _rc.font = Font(bold=True, size=10)
            _rc.fill = _wf[_wv]; _rc.alignment = lft; _rc.border = bdr
            for _jj in range(_n):
                _col = _jj + 2
                if _ii == _jj:
                    _a = alphas[_ii]
                    _av = f"({_a:.3f})" if not np.isnan(_a) else "(—)"
                    _cc = ws.cell(row=_row, column=_col, value=_av)
                    _cc.font = Font(bold=True, size=10)
                    _cc.fill = _wf[_wv]; _cc.alignment = ctr; _cc.border = bdr
                elif _ii > _jj:
                    _vd = pd.concat([composites[_ii], composites[_jj]], axis=1).dropna()
                    if len(_vd) > 2:
                        _rv, _pv = stats.pearsonr(_vd.iloc[:, 0], _vd.iloc[:, 1])
                        _star, _ = fmt_p(_pv)
                        cell(ws, _row, _col, f"{_rv:.2f}{_star}")
                    else:
                        cell(ws, _row, _col, '—')
                else:
                    _uc = ws.cell(row=_row, column=_col, value='')
                    _uc.border = bdr; _uc.alignment = ctr
        _r = _data_start + _n
        for _lbl_r in ['M', 'SD']:
            _lc = ws.cell(row=_r, column=1, value=_lbl_r)
            _lc.font = Font(bold=True, size=10); _lc.alignment = ctr; _lc.border = bdr
            for _jj, _comp in enumerate(composites):
                _col = _jj + 2
                _v = (_comp.mean() if _lbl_r == 'M' else _comp.std()) if _comp.notna().sum() > 1 else np.nan
                _vc = ws.cell(row=_r, column=_col,
                              value=f"{_v:.2f}" if not np.isnan(_v) else '—')
                _vc.font = Font(size=10); _vc.alignment = ctr; _vc.border = bdr
            _r += 1
        _fn = ws.cell(row=_r, column=1,
            value="對角線括號內為 Cronbach's α；*** p<.001  ** p<.01  * p<.05"
                  "（下三角為 Pearson 相關係數；成對刪除法處理遺漏值）")
        _fn.font = Font(italic=True, size=9)
        ws.merge_cells(start_row=_r, start_column=1, end_row=_r, end_column=_end_col)
        return _r + 1

    # ── 上半部：表 4A  12×12（HP/JCP/DP/CI × T1~T3，不含 PP）──
    _iv_defs_12 = [
        ('HP',  'T1', 'HP（T1）階層停滯',     6),
        ('JCP', 'T1', 'JCP（T1）工作停滯',    6),
        ('DP',  'T1', 'DP（T1）決策拖延',     5),
        ('CI',  'T1', 'CI（T1）職涯無所作為', 8),
        ('HP',  'T2', 'HP（T2）階層停滯',     6),
        ('JCP', 'T2', 'JCP（T2）工作停滯',    6),
        ('DP',  'T2', 'DP（T2）決策拖延',     5),
        ('CI',  'T2', 'CI（T2）職涯無所作為', 8),
        ('HP',  'T3', 'HP（T3）階層停滯',     6),
        ('JCP', 'T3', 'JCP（T3）工作停滯',    6),
        ('DP',  'T3', 'DP（T3）決策拖延',     5),
        ('CI',  'T3', 'CI（T3）職涯無所作為', 8),
    ]
    _comp_12, _alpha_12 = _build_iv(_iv_defs_12)
    _next_r = _draw_corr_block(
        ws_corr, _iv_defs_12, _comp_12, _alpha_12,
        start_row=1,
        title_str=f"表 4A  三波段追蹤相關矩陣（12×12，不含 PP）  N={n_total}  *** p<.001  ** p<.01  * p<.05",
        note_str="對角線（括號內）為 Cronbach's α；HP=階層停滯、JCP=工作停滯、DP=決策拖延、CI=職涯無所作為"
    )

    _next_r += 1  # 分隔空行

    # ── 下半部：表 4B  15×15（含 PP，完整）──
    _iv_defs = [
        ('HP',  'T1', 'HP（T1）階層停滯',     6),
        ('JCP', 'T1', 'JCP（T1）工作停滯',    6),
        ('DP',  'T1', 'DP（T1）決策拖延',     5),
        ('CI',  'T1', 'CI（T1）職涯無所作為', 8),
        ('PP',  'T1', 'PP（T1）主動型人格',   6),
        ('HP',  'T2', 'HP（T2）階層停滯',     6),
        ('JCP', 'T2', 'JCP（T2）工作停滯',    6),
        ('DP',  'T2', 'DP（T2）決策拖延',     5),
        ('CI',  'T2', 'CI（T2）職涯無所作為', 8),
        ('PP',  'T2', 'PP（T2）主動型人格',   6),
        ('HP',  'T3', 'HP（T3）階層停滯',     6),
        ('JCP', 'T3', 'JCP（T3）工作停滯',    6),
        ('DP',  'T3', 'DP（T3）決策拖延',     5),
        ('CI',  'T3', 'CI（T3）職涯無所作為', 8),
        ('PP',  'T3', 'PP（T3）主動型人格',   6),
    ]
    _n_iv = len(_iv_defs)
    _iv_composites, _iv_alphas_list = _build_iv(_iv_defs)
    _next_r = _draw_corr_block(
        ws_corr, _iv_defs, _iv_composites, _iv_alphas_list,
        start_row=_next_r,
        title_str=f"表 4B  三波段追蹤相關矩陣（15×15，含 PP）  N={n_total}  *** p<.001  ** p<.01  * p<.05",
        note_str="對角線（括號內）為 Cronbach's α；HP=階層停滯、JCP=工作停滯、DP=決策拖延、CI=職涯無所作為、PP=主動型人格"
    )
    _corr_note_r = _next_r

    ws_corr.column_dimensions['A'].width = 25
    for _jj in range(_n_iv):
        _cl = get_column_letter(_jj + 2)
        ws_corr.column_dimensions[_cl].width = 7

    # ── 分波次相關矩陣（T1 / T2 / T3，附 M/SD，堆疊於同一 sheet）────────
    _wave_sep_r = _corr_note_r + 2
    _wave_sec_hdr = ws_corr.cell(row=_wave_sep_r, column=1,
        value="各波次同時間點相關矩陣（HP / JCP / DP / CI / PP）")
    _wave_sec_hdr.font = Font(bold=True, size=11, color="FFFFFF")
    _wave_sec_hdr.fill = PatternFill("solid", fgColor="2E4057")
    ws_corr.merge_cells(start_row=_wave_sep_r, start_column=1,
                        end_row=_wave_sep_r, end_column=9)
    _wave_sep_r += 1
    for _wave in ['T1', 'T2', 'T3']:
        _wblk = ws_corr.cell(row=_wave_sep_r, column=1,
                             value=f"━━━  {_wave} 相關矩陣  ━━━")
        _wblk.font = Font(bold=True, size=11, color="FFFFFF")
        _wblk.fill = PatternFill("solid", fgColor="4472C4")
        ws_corr.merge_cells(start_row=_wave_sep_r, start_column=1,
                            end_row=_wave_sep_r, end_column=9)
        _wave_sep_r += 1
        _wave_sep_r = make_corr_block(ws_corr, _wave, scale_order_sub,
                                       f'次量表相關矩陣（{_wave}：HP/JCP/DP/CI/PP）',
                                       _wave_sep_r)

    # ── Sheet 4 & 5: 適配指數（共用內部函式）─────────────────────
    fit_hdr_cols = ["模型", "結構說明", "χ²", "df", "p(χ²)",
                    "CFI", "TLI", "RMSEA", "90% CI", "SRMR", "判斷"]

    def write_fit_sheet(ws, sheet_title, model_info_dict, start_row=1):
        if start_row == 1:
            title(ws, sheet_title, end_col=11)
        else:
            _tc = ws.cell(row=start_row, column=1, value=sheet_title)
            _tc.font = Font(bold=True, size=12)
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=start_row, end_column=11)
        r2 = start_row + 2
        for ci, h in enumerate(fit_hdr_cols, 1):
            hdr(ws, r2, ci, h)
        r2 += 1
        for mkey, mval in model_info_dict.items():
            # mval can be (display_name, description) or just a description string
            if isinstance(mval, tuple):
                mdisplay, mdesc = mval
            else:
                mdisplay, mdesc = mkey, mval
            fit = all_results.get(mkey, {}).get('fit', {})
            chi2_v  = fit.get('chi2')
            df_v    = fit.get('df')
            p_chi2  = fit.get('p_chi2')
            cfi_v   = fit.get('cfi')
            tli_v   = fit.get('tli')
            rmsea_v = fit.get('rmsea')
            rlo_v   = fit.get('rmsea_lo')
            rhi_v   = fit.get('rmsea_hi')
            srmr_v  = fit.get('srmr')
            ok = (isinstance(cfi_v, float) and cfi_v >= .90 and
                  isinstance(rmsea_v, float) and rmsea_v <= .08 and
                  isinstance(srmr_v, float) and srmr_v <= .08)
            verdict = '✅' if ok else ('⚠️' if fit else '（尚未執行）')
            vals = [
                mdisplay, mdesc,
                f"{chi2_v:.2f}" if isinstance(chi2_v, float) else '—',
                df_v if df_v is not None else '—',
                f"{p_chi2:.3f}" if isinstance(p_chi2, float) else '—',
                f"{cfi_v:.3f}" if isinstance(cfi_v, float) else '—',
                f"{tli_v:.3f}" if isinstance(tli_v, float) else '—',
                f"{rmsea_v:.3f}" if isinstance(rmsea_v, float) else '—',
                (f"[{rlo_v:.3f}, {rhi_v:.3f}]"
                 if isinstance(rlo_v, float) and isinstance(rhi_v, float) else '—'),
                f"{srmr_v:.3f}" if isinstance(srmr_v, float) else '—',
                verdict,
            ]
            for ci, val in enumerate(vals, 1):
                cell(ws, r2, ci, val, align='left' if ci <= 2 else 'center')
            r2 += 1
        r2 += 1
        ws.cell(row=r2, column=1,
                value="判斷標準：CFI ≥ .90；RMSEA ≤ .08；SRMR ≤ .08"
                ).font = Font(italic=True, size=9)
        ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=11)
        if start_row == 1:
            set_widths(ws, [('A', 22), ('B', 24), ('C', 8), ('D', 5), ('E', 8),
                            ('F', 7), ('G', 7), ('H', 9), ('I', 14), ('J', 7), ('K', 6)])

    ws4 = wb.create_sheet("5_CFA分析")
    write_fit_sheet(ws4, f"CFA 驗證性因素分析適配指數（T1, N = {n_total}）", {
        'CFA-A (JCP+DP+CI)':               ('模型1（CFA-A）', 'JCP + DP + CI（3因子，主路徑 A）'),
        'CFA-B (HP+DP+CI)':                ('模型2（CFA-B）', 'HP + DP + CI（3因子，主路徑 B）'),
        'CFA-C (JCP+PP+DP+CI)':            ('模型3（CFA-C）', 'JCP + PP + DP + CI（4因子）'),
        'CFA-D (HP+PP+DP+CI)':             ('模型4（CFA-D）', 'HP + PP + DP + CI（4因子）'),
        'CFA-E (HP+JCP+PP+DP+CI, 5F)':   ('模型5（CFA-E）', 'HP/JCP/PP/DP/CI 五因子（區別效度基準，T1）'),
        'CFA-F (CP_merged+PP+DP+CI, 4F)': ('模型6（CFA-F）', 'CP合併/PP/DP/CI 四因子（對照 CFA-E）'),
        'CFA-G (CP_merged+DP+CI, 3F)':    ('模型7（CFA-G）', 'CP合併/DP/CI 三因子（最簡對照）'),
        'CFA-H (Cross-Wave 5F)':           ('模型8（CFA-H）', 'HP/JCP/PP(T1) + DP(T2) + CI(T3) 跨波次五因子'),
        'CFA-I (Cross-Wave 4F, no PP)':    ('模型9（CFA-I）', 'HP/JCP(T1) + DP(T2) + CI(T3) 跨波次四因子（不含 PP）'),
    })

    # ── AVE / CR 聚合效度指標（CFA-E 五因子，附在 CFA適配 表下方）──
    # 找到目前寫入的最後一行
    _ave_r = ws4.max_row + 2
    _ave_title = ws4.cell(row=_ave_r, column=1,
        value="聚合效度指標（各 CFA 模型；CFA-H 為跨波次）")
    _ave_title.font = Font(bold=True, size=11)
    ws4.merge_cells(start_row=_ave_r, start_column=1,
                    end_row=_ave_r, end_column=6)
    _ave_r += 1
    _ave_hdr = ["模型／因子", "題數", "因素負荷量範圍", "AVE", "CR", "判斷（AVE≥.50 & CR≥.70）"]
    for _ci, _h in enumerate(_ave_hdr, 1):
        hdr(ws4, _ave_r, _ci, _h)
    _ave_r += 1

    _fac_labels = {
        'HP': 'HP 階層停滯', 'JCP': 'JCP 工作內容停滯',
        'PP': 'PP 主動型人格', 'DP': 'DP 決策拖延',
        'CI': 'CI 職涯無所作為', 'CP': 'CP 職涯高原（合併）',
    }
    # 模型6/7 使用 CP parcel 合成分數，僅供區別效度比較，不計算 AVE/CR
    _cfa_ave_models = [
        ('CFA-A (JCP+DP+CI)',            '模型1（CFA-A）', ['JCP', 'DP', 'CI']),
        ('CFA-B (HP+DP+CI)',             '模型2（CFA-B）', ['HP',  'DP', 'CI']),
        ('CFA-C (JCP+PP+DP+CI)',         '模型3（CFA-C）', ['JCP', 'PP', 'DP', 'CI']),
        ('CFA-D (HP+PP+DP+CI)',          '模型4（CFA-D）', ['HP',  'PP', 'DP', 'CI']),
        ('CFA-E (HP+JCP+PP+DP+CI, 5F)','模型5（CFA-E）', ['HP', 'JCP', 'PP', 'DP', 'CI']),
        ('CFA-H (Cross-Wave 5F)',        '模型8（CFA-H）', ['HP', 'JCP', 'PP', 'DP', 'CI']),
        ('CFA-I (Cross-Wave 4F, no PP)', '模型9（CFA-I）', ['HP', 'JCP', 'DP', 'CI']),
    ]
    for _mkey, _mname, _facs in _cfa_ave_models:
        _ave_cr_data = all_results.get(_mkey, {}).get('ave_cr', {})
        # 模型標題列
        _mtitle = ws4.cell(row=_ave_r, column=1, value=_mname)
        _mtitle.font = Font(bold=True, size=10)
        ws4.merge_cells(start_row=_ave_r, start_column=1,
                        end_row=_ave_r, end_column=6)
        _ave_r += 1
        for _fac in _facs:
            _fd    = _ave_cr_data.get(_fac, {})
            _lams  = _fd.get('loadings', [])
            _n     = _fd.get('n_items', len(_lams))
            _ave_v = _fd.get('AVE', float('nan'))
            _cr_v  = _fd.get('CR',  float('nan'))
            _lam_range = f"{min(_lams):.3f} ~ {max(_lams):.3f}" if _lams else '—'
            _ave_str   = f"{_ave_v:.3f}" if not np.isnan(_ave_v) else '—'
            _cr_str    = f"{_cr_v:.3f}"  if not np.isnan(_cr_v)  else '—'
            _ok = (not np.isnan(_ave_v) and not np.isnan(_cr_v)
                   and _ave_v >= 0.50 and _cr_v >= 0.70)
            _verdict = '✅ 通過' if _ok else ('⚠️ 未達標' if _lams else '（尚未執行）')
            _row_vals = [f"  {_fac_labels.get(_fac, _fac)}", _n, _lam_range,
                         _ave_str, _cr_str, _verdict]
            for _ci, _v in enumerate(_row_vals, 1):
                cell(ws4, _ave_r, _ci, _v,
                     color=('006100' if _ok else ('C00000' if _lams else '808080')),
                     bold=(_ci == 6 and bool(_lams)))
            _ave_r += 1
        _ave_r += 1  # 模型間空一行

    _note_r = _ave_r
    ws4.cell(row=_note_r, column=1,
             value="判斷標準：AVE ≥ .50（聚合效度）；CR ≥ .70（組合信度）；資料來源：各模型 STDYX 標準化因素負荷量"
             ).font = Font(italic=True, size=9)
    ws4.merge_cells(start_row=_note_r, start_column=1,
                    end_row=_note_r, end_column=6)

    # ── Sheet 6: 假設檢驗 ─────────────────────────────────────────
    ws_hyp = wb.create_sheet("6_假設檢驗")

    def _hyp_ci_fmt(lo, hi):
        if np.isnan(lo) or np.isnan(hi): return '—'
        return f"[{lo:.3f}, {hi:.3f}]"

    def _write_fit_row(ws, row, model_label, res):
        fit = res.get('fit', {})
        def _f(k, fmt='.3f'):
            v = fit.get(k, np.nan)
            return format(v, fmt) if isinstance(v, float) and not np.isnan(v) else '—'
        rmsea_lo = fit.get('rmsea_lo', np.nan)
        rmsea_hi = fit.get('rmsea_hi', np.nan)
        rmsea_str = (f"{fit.get('rmsea',np.nan):.3f} [{rmsea_lo:.3f}, {rmsea_hi:.3f}]"
                     if not (np.isnan(rmsea_lo) or np.isnan(rmsea_hi)) else _f('rmsea'))
        vals = [model_label, _f('chi2','.2f'), _f('df','.0f'),
                _f('cfi'), _f('tli'), rmsea_str, _f('srmr'),
                _f('aic','.1f'), _f('bic','.1f'), '']
        for _fi, _fv in enumerate(vals, 1):
            _fc = ws.cell(row=row, column=_fi, value=_fv)
            _fc.font = Font(size=10)
            _fc.alignment = ctr if _fi > 1 else lft
            _fc.border = bdr

    green_fill = PatternFill("solid", fgColor="E2EFDA")
    red_fill   = PatternFill("solid", fgColor="FFE0E0")
    _hyp_r = 3

    if variant_label == 'nopp':
        # ── NoPP 版本：純中介模型 H1~H4 ────────────────────────────
        _nopp_res  = all_results.get('PATH_NoPP', {})
        title(ws_hyp,
              f"假設檢驗：純中介分析（不含 PP 調節）H1~H4  N={n_total}"
              f"  *** p<.001  ** p<.01  * p<.05",
              end_col=10)
        ws_hyp.cell(row=2, column=1,
            value="Mplus ML + Bootstrap 5000（BCBOOTSTRAP CI）；β = 標準化路徑係數；"
                  "HP=階層停滯、JCP=工作內容停滯、DP=決策拖延、CI=職涯無所作為"
        ).font = Font(italic=True, size=9)
        ws_hyp.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

        # 適配指數
        _fit_sec = ws_hyp.cell(row=_hyp_r, column=1, value="路徑模型適配指數（ML Estimator）")
        _fit_sec.font = Font(bold=True, size=11, color="FFFFFF")
        _fit_sec.fill = PatternFill("solid", fgColor="2E4057")
        ws_hyp.merge_cells(start_row=_hyp_r, start_column=1, end_row=_hyp_r, end_column=10)
        _hyp_r += 1
        for _fi, _fh in enumerate(["模型", "χ²", "df", "CFI", "TLI",
                                    "RMSEA [90% CI]", "SRMR", "AIC", "BIC", "備注"], 1):
            hdr(ws_hyp, _hyp_r, _fi, _fh)
        _hyp_r += 1
        _write_fit_row(ws_hyp, _hyp_r,
                       'PATH_NoPP: JCP/HP(T1)→DP(T2)→CI(T3)', _nopp_res)
        _hyp_r += 2

        # 直接路徑 H1~H3
        _main_sec = ws_hyp.cell(row=_hyp_r, column=1,
            value="假設檢驗：直接路徑係數（Mplus STDYX 標準化 β）")
        _main_sec.font = Font(bold=True, size=11, color="FFFFFF")
        _main_sec.fill = PatternFill("solid", fgColor="4472C4")
        ws_hyp.merge_cells(start_row=_hyp_r, start_column=1, end_row=_hyp_r, end_column=10)
        _hyp_r += 1
        for _ci, _h in enumerate(["假設", "路徑說明", "時間點", "預期方向",
                                   "IV", "DV", "β", "SE", "95% BC CI", "支持？"], 1):
            hdr(ws_hyp, _hyp_r, _ci, _h)
        _hyp_r += 1

        def _get_nopp_path(path_key):
            d = _nopp_res.get('paths', {}).get(path_key, {})
            return (d.get('est', np.nan), d.get('se', np.nan), d.get('p', np.nan),
                    d.get('ci_lo', np.nan), d.get('ci_hi', np.nan))

        _nopp_direct_hyps = [
            ('H1a', 'JCP（工作內容停滯）→ DP（決策拖延）', 'T1→T2',
             '+', 'JCP（T1）', 'DP（T2）', 'H1a: JCP(T1)→DP(T2)'),
            ('H1b', 'HP（階層停滯）→ DP（決策拖延）',      'T1→T2',
             '+', 'HP（T1）',  'DP（T2）', 'H1b: HP(T1)→DP(T2)'),
            ('H2',  'DP（決策拖延）→ CI（職涯無所作為）',  'T2→T3',
             '+', 'DP（T2）',  'CI（T3）', 'H2: DP(T2)→CI(T3)'),
            ('H3a', 'JCP（工作內容停滯）→ CI（職涯無所作為）直接', 'T1→T3',
             '+', 'JCP（T1）', 'CI（T3）', 'H3a: JCP(T1)→CI(T3) 直接'),
            ('H3b', 'HP（階層停滯）→ CI（職涯無所作為）直接',      'T1→T3',
             '+', 'HP（T1）',  'CI（T3）', 'H3b: HP(T1)→CI(T3) 直接'),
        ]
        for _hid, _hdesc, _tpt, _exp_dir, _iv_lbl, _dv_lbl, _pkey in _nopp_direct_hyps:
            _beta, _se, _pv, _ci_lo, _ci_hi = _get_nopp_path(_pkey)
            _star, _ = fmt_p(_pv) if not np.isnan(_pv) else ('', '—')
            _b_str  = f"{_beta:.3f}{_star}" if not np.isnan(_beta) else '待填入'
            _se_str = f"({_se:.3f})"        if not np.isnan(_se)   else '—'
            _ci_str = _hyp_ci_fmt(_ci_lo, _ci_hi)
            _sig    = not np.isnan(_pv) and _pv < .05
            _dir_ok = ((_exp_dir == '+' and not np.isnan(_beta) and _beta > 0) or
                       (_exp_dir == '-' and not np.isnan(_beta) and _beta < 0))
            _support = _sig and _dir_ok
            _verdict = '支持' if _support else ('不支持' if not np.isnan(_beta) else '待計算')
            _rfill = green_fill if _support else (red_fill if not np.isnan(_beta) else None)
            for _ci, _v in enumerate([_hid, _hdesc, _tpt, _exp_dir,
                                       _iv_lbl, _dv_lbl, _b_str, _se_str,
                                       _ci_str, _verdict], 1):
                _c = ws_hyp.cell(row=_hyp_r, column=_ci, value=_v)
                _c.font = Font(size=10, bold=(_ci == 10 and _support),
                               color=('006100' if _support
                                      else ('C00000' if not np.isnan(_beta) and not _support
                                            else '000000')))
                _c.alignment = ctr if _ci >= 3 else lft
                _c.border = bdr
                if _rfill: _c.fill = _rfill
            _hyp_r += 1

        # 間接效果 H4
        _hyp_r += 1
        _h4_sec = ws_hyp.cell(row=_hyp_r, column=1,
            value="H4 間接效果：JCP/HP(T1)→DP(T2)→CI(T3)"
                  "（Mplus MODEL CONSTRAINT，BC Bootstrap 95% CI）")
        _h4_sec.font = Font(bold=True, size=11, color="FFFFFF")
        _h4_sec.fill = PatternFill("solid", fgColor="2E4057")
        ws_hyp.merge_cells(start_row=_hyp_r, start_column=1, end_row=_hyp_r, end_column=10)
        _hyp_r += 1
        for _ci, _h in enumerate(["假設", "路徑說明", "—",
                                   "IV", "DV", "β (indirect)", "SE", "95% BC CI", "顯著？", "—"], 1):
            hdr(ws_hyp, _hyp_r, _ci, _h)
        _hyp_r += 1
        _nopp_mc_res = _nopp_res.get('modconstr', {})
        for _hid4, _path_lbl, _iv4, _dv4, _mc_key in [
            ('H4a', 'JCP(T1)→DP(T2)→CI(T3)', 'JCP（T1）', 'CI（T3）', 'IND_JCP'),
            ('H4b', 'HP(T1)→DP(T2)→CI(T3)',  'HP（T1）',  'CI（T3）', 'IND_HP'),
        ]:
            _mc4 = _nopp_mc_res.get(_mc_key, {})
            _ie    = _mc4.get('est', np.nan)
            _ie_se = _mc4.get('se',  np.nan)
            _ie_lo = _mc4.get('ci_lo', np.nan)
            _ie_hi = _mc4.get('ci_hi', np.nan)
            _ie_sig = _mc4.get('sig', False)
            for _ci, _v in enumerate([
                _hid4, _path_lbl, '—', _iv4, _dv4,
                f"{_ie:.3f}"      if not np.isnan(_ie)    else '待填入',
                f"({_ie_se:.3f})" if not np.isnan(_ie_se) else '—',
                _hyp_ci_fmt(_ie_lo, _ie_hi),
                '顯著' if _ie_sig else ('不顯著' if not np.isnan(_ie) else '—'),
                '',
            ], 1):
                _hc = ws_hyp.cell(row=_hyp_r, column=_ci, value=_v)
                _hc.font = Font(size=10, bold=(_ci == 9 and _ie_sig),
                                color='006100' if _ie_sig else '000000')
                _hc.alignment = ctr if _ci >= 2 else lft
                _hc.border = bdr
                if _ie_sig: _hc.fill = green_fill
            _hyp_r += 1

    else:
        # ── WithPP 版本：完整調節中介 H1~H7 ───────────────────────
        _PATH_KEY = 'PATH (T1→T2→T3)'
        _path_res  = all_results.get(_PATH_KEY, {})
        title(ws_hyp,
              f"假設檢驗：完整調節中介分析  PP(T1)調節 a/b/c' 三條路徑  H1~H7  N={n_total}"
              f"  *** p<.001  ** p<.01  * p<.05",
              end_col=10)
        ws_hyp.cell(row=2, column=1,
            value="Mplus ML + Bootstrap 5000（BCBOOTSTRAP CI）；β = 標準化路徑係數；"
                  "HP=階層停滯、JCP=工作內容停滯、DP=決策拖延、CI=職涯無所作為、PP=主動型人格（調節變項）"
        ).font = Font(italic=True, size=9)
        ws_hyp.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

        # 路徑模型適配指數
        _fit_sec = ws_hyp.cell(row=_hyp_r, column=1,
            value="路徑模型適配指數（ML Estimator）")
        _fit_sec.font = Font(bold=True, size=11, color="FFFFFF")
        _fit_sec.fill = PatternFill("solid", fgColor="2E4057")
        ws_hyp.merge_cells(start_row=_hyp_r, start_column=1, end_row=_hyp_r, end_column=10)
        _hyp_r += 1
        for _fi, _fh in enumerate(["模型", "χ²", "df", "CFI", "TLI",
                                    "RMSEA [90% CI]", "SRMR", "AIC", "BIC", "備注"], 1):
            hdr(ws_hyp, _hyp_r, _fi, _fh)
        _hyp_r += 1
        _write_fit_row(ws_hyp, _hyp_r,
                       'PATH: T1(JCP/HP/PP)→T2(DP)→T3(CI)', _path_res)
        _hyp_r += 2

        # 假設主表（直接效果 H1/H2/H3/H4/H5/H6）
        _main_sec = ws_hyp.cell(row=_hyp_r, column=1,
            value="假設檢驗：直接路徑係數（Mplus STDYX 標準化 β）")
        _main_sec.font = Font(bold=True, size=11, color="FFFFFF")
        _main_sec.fill = PatternFill("solid", fgColor="4472C4")
        ws_hyp.merge_cells(start_row=_hyp_r, start_column=1, end_row=_hyp_r, end_column=10)
        _hyp_r += 1
        for _ci, _h in enumerate(["假設", "路徑說明", "時間點", "預期方向",
                                   "IV", "DV", "β", "SE", "95% BC CI", "支持？"], 1):
            hdr(ws_hyp, _hyp_r, _ci, _h)
        _hyp_r += 1

        def _get_path(path_key):
            d = _path_res.get('paths', {}).get(path_key, {})
            return (d.get('est', np.nan), d.get('se', np.nan), d.get('p', np.nan),
                    d.get('ci_lo', np.nan), d.get('ci_hi', np.nan))

        def _var_label(sn, wv):
            return next((l for s, w, l, n in _iv_defs if s == sn and w == wv), f'{sn}_{wv}')

        _direct_hyps = [
            ('H1a', 'JCP（工作內容停滯）→ DP（決策拖延）[at mean PP]', 'T1→T2',
             '+', 'JCP（T1）',    'DP（T2）', 'H1a: JCP(T1)→DP(T2) [at mean PP]'),
            ('H1b', 'HP（階層停滯）→ DP（決策拖延）[at mean PP]',      'T1→T2',
             '+', 'HP（T1）',     'DP（T2）', 'H1b: HP(T1)→DP(T2) [at mean PP]'),
            ('H2a', 'PP 調節 JCP→DP（PP×JCP 交互）',  'T1×T1→T2',
             '-', 'PP×JCP（T1）', 'DP（T2）', 'H2a: PP×JCP→DP (moderation)'),
            ('H2b', 'PP 調節 HP→DP（PP×HP 交互）',    'T1×T1→T2',
             '-', 'PP×HP（T1）',  'DP（T2）', 'H2b: PP×HP→DP (moderation)'),
            ('H3',  'DP（決策拖延）→ CI（職涯無所作為）[at mean PP]', 'T2→T3',
             '+', 'DP（T2）',     'CI（T3）', 'H3:  DP(T2)→CI(T3) [at mean PP]'),
            ('H4',  'PP 調節 DP→CI（PP×DP 交互）',    'T2×T1→T3',
             '-', 'PP×DP（T1/T2）', 'CI（T3）', 'H4:  PP×DP→CI (moderation)'),
            ('H5a', 'JCP（工作內容停滯）→ CI（職涯無所作為）直接效果 [at mean PP]', 'T1→T3',
             '+', 'JCP（T1）',    'CI（T3）', 'H5a: JCP(T1)→CI(T3) [at mean PP]'),
            ('H5b', 'HP（階層停滯）→ CI（職涯無所作為）直接效果 [at mean PP]',      'T1→T3',
             '+', 'HP（T1）',     'CI（T3）', 'H5b: HP(T1)→CI(T3) [at mean PP]'),
            ('H6a', 'PP 調節 JCP→CI 直接效果（PP×JCP 交互）',  'T1×T1→T3',
             '-', 'PP×JCP（T1）', 'CI（T3）', 'H6a: PP×JCP→CI (moderation)'),
            ('H6b', 'PP 調節 HP→CI 直接效果（PP×HP 交互）',    'T1×T1→T3',
             '-', 'PP×HP（T1）',  'CI（T3）', 'H6b: PP×HP→CI (moderation)'),
        ]
        for _hid, _hdesc, _tpt, _exp_dir, _iv_lbl, _dv_lbl, _pkey in _direct_hyps:
            _beta, _se, _pv, _ci_lo, _ci_hi = _get_path(_pkey)
            _star, _p_str = fmt_p(_pv) if not np.isnan(_pv) else ('', '—')
            _b_str    = f"{_beta:.3f}{_star}" if not np.isnan(_beta) else '待填入'
            _se_str   = f"({_se:.3f})"        if not np.isnan(_se)   else '—'
            _ci_str   = _hyp_ci_fmt(_ci_lo, _ci_hi)
            _sig    = not np.isnan(_pv) and _pv < .05
            _dir_ok = ((_exp_dir == '+' and not np.isnan(_beta) and _beta > 0) or
                       (_exp_dir == '-' and not np.isnan(_beta) and _beta < 0))
            _support = _sig and _dir_ok
            _verdict = '支持' if _support else ('不支持' if not np.isnan(_beta) else '待計算')
            _rfill = green_fill if _support else (red_fill if not np.isnan(_beta) else None)
            for _ci, _v in enumerate([_hid, _hdesc, _tpt, _exp_dir,
                                       _iv_lbl, _dv_lbl, _b_str, _se_str,
                                       _ci_str, _verdict], 1):
                _c = ws_hyp.cell(row=_hyp_r, column=_ci, value=_v)
                _c.font = Font(size=10,
                               bold=(_ci == 10 and _support),
                               color=('006100' if _support
                                      else ('C00000' if not np.isnan(_beta) and not _support
                                            else '000000')))
                _c.alignment = ctr if _ci >= 3 else lft
                _c.border = bdr
                if _rfill: _c.fill = _rfill
            _hyp_r += 1

        # H6 條件直接效果
        _hyp_r += 1
        _h6_sec = ws_hyp.cell(row=_hyp_r, column=1,
            value="H6 條件直接效果：JCP/HP(T1)→CI(T3)，PP = ±1SD（c'-path 隨 PP 水準變化）"
                  "（Mplus MODEL CONSTRAINT，BC Bootstrap 95% CI）")
        _h6_sec.font = Font(bold=True, size=11, color="FFFFFF")
        _h6_sec.fill = PatternFill("solid", fgColor="375623")
        ws_hyp.merge_cells(start_row=_hyp_r, start_column=1, end_row=_hyp_r, end_column=10)
        _hyp_r += 1
        for _ci, _h in enumerate(["假設", "路徑說明", "PP 水準",
                                   "IV", "DV", "β (direct)", "SE", "95% BC CI", "顯著？", "—"], 1):
            hdr(ws_hyp, _hyp_r, _ci, _h)
        _hyp_r += 1
        _mc_res = _path_res.get('modconstr', {})
        for _hid6, _path_lbl, _pp_lbl, _iv6, _dv6, _mc_key in [
            ('H6a', 'JCP(T1)→CI(T3) 直接', 'PP +1SD（主動型人格高）', 'JCP（T1）', 'CI（T3）', 'DIR_HI_J'),
            ('H6a', 'JCP(T1)→CI(T3) 直接', 'PP -1SD（主動型人格低）', 'JCP（T1）', 'CI（T3）', 'DIR_LO_J'),
            ('H6b', 'HP(T1)→CI(T3) 直接',  'PP +1SD（主動型人格高）', 'HP（T1）',  'CI（T3）', 'DIR_HI_H'),
            ('H6b', 'HP(T1)→CI(T3) 直接',  'PP -1SD（主動型人格低）', 'HP（T1）',  'CI（T3）', 'DIR_LO_H'),
        ]:
            _mc6 = _mc_res.get(_mc_key, {})
            _de    = _mc6.get('est', np.nan)
            _de_se = _mc6.get('se',  np.nan)
            _de_lo = _mc6.get('ci_lo', np.nan)
            _de_hi = _mc6.get('ci_hi', np.nan)
            _de_sig = _mc6.get('sig', False)
            for _ci, _v in enumerate([
                _hid6, _path_lbl, _pp_lbl, _iv6, _dv6,
                f"{_de:.3f}"      if not np.isnan(_de)    else '待填入',
                f"({_de_se:.3f})" if not np.isnan(_de_se) else '—',
                _hyp_ci_fmt(_de_lo, _de_hi),
                '顯著' if _de_sig else ('不顯著' if not np.isnan(_de) else '—'),
                '',
            ], 1):
                _hc = ws_hyp.cell(row=_hyp_r, column=_ci, value=_v)
                _hc.font = Font(size=10, bold=(_ci == 9 and _de_sig),
                                color='006100' if _de_sig else '000000')
                _hc.alignment = ctr if _ci >= 2 else lft
                _hc.border = bdr
                if _de_sig: _hc.fill = green_fill
            _hyp_r += 1

        # H7 條件間接效果
        _hyp_r += 1
        _h7_sec = ws_hyp.cell(row=_hyp_r, column=1,
            value="H7 條件間接效果：CP(T1)→DP(T2)→CI(T3)，PP = ±1SD（a-path＋b-path 均條件化）"
                  "（Mplus MODEL CONSTRAINT，BC Bootstrap 95% CI）")
        _h7_sec.font = Font(bold=True, size=11, color="FFFFFF")
        _h7_sec.fill = PatternFill("solid", fgColor="2E4057")
        ws_hyp.merge_cells(start_row=_hyp_r, start_column=1, end_row=_hyp_r, end_column=10)
        _hyp_r += 1
        for _ci, _h in enumerate(["假設", "路徑說明", "PP 水準",
                                   "IV", "DV", "β (indirect)", "SE", "95% BC CI", "顯著？", "—"], 1):
            hdr(ws_hyp, _hyp_r, _ci, _h)
        _hyp_r += 1
        for _hid7, _path_lbl, _pp_lbl, _iv7, _dv7, _mc_key in [
            ('H7a', 'JCP(T1)→DP(T2)→CI(T3)', 'PP +1SD（主動型人格高）', 'JCP（T1）', 'CI（T3）', 'IND_HI_J'),
            ('H7a', 'JCP(T1)→DP(T2)→CI(T3)', 'PP -1SD（主動型人格低）', 'JCP（T1）', 'CI（T3）', 'IND_LO_J'),
            ('H7b', 'HP(T1)→DP(T2)→CI(T3)',  'PP +1SD（主動型人格高）', 'HP（T1）',  'CI（T3）', 'IND_HI_H'),
            ('H7b', 'HP(T1)→DP(T2)→CI(T3)',  'PP -1SD（主動型人格低）', 'HP（T1）',  'CI（T3）', 'IND_LO_H'),
        ]:
            _mc7 = _mc_res.get(_mc_key, {})
            _ie    = _mc7.get('est', np.nan)
            _ie_se = _mc7.get('se',  np.nan)
            _ie_lo = _mc7.get('ci_lo', np.nan)
            _ie_hi = _mc7.get('ci_hi', np.nan)
            _ie_sig = _mc7.get('sig', False)
            for _ci, _v in enumerate([
                _hid7, _path_lbl, _pp_lbl, _iv7, _dv7,
                f"{_ie:.3f}"      if not np.isnan(_ie)    else '待填入',
                f"({_ie_se:.3f})" if not np.isnan(_ie_se) else '—',
                _hyp_ci_fmt(_ie_lo, _ie_hi),
                '顯著' if _ie_sig else ('不顯著' if not np.isnan(_ie) else '—'),
                '',
            ], 1):
                _hc = ws_hyp.cell(row=_hyp_r, column=_ci, value=_v)
                _hc.font = Font(size=10, bold=(_ci == 9 and _ie_sig),
                                color='006100' if _ie_sig else '000000')
                _hc.alignment = ctr if _ci >= 2 else lft
                _hc.border = bdr
                if _ie_sig: _hc.fill = green_fill
            _hyp_r += 1

    set_widths(ws_hyp, [('A', 10), ('B', 40), ('C', 30), ('D', 16),
                        ('E', 18), ('F', 18), ('G', 10), ('H', 18),
                        ('I', 16), ('J', 10)])

    # ── Sheet 9: Mplus 語法（.inp 內容）── 語法附錄，最後 ────────
    ws8 = wb.create_sheet("9_Mplus語法")
    ws8.cell(row=1, column=1,
             value="Mplus 分析語法（.inp）— 可直接複製至 Mplus 執行"
             ).font = Font(bold=True, size=12)
    ws8.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    hdr(ws8, 2, 1, "模型")
    hdr(ws8, 2, 2, "檔案名稱")
    hdr(ws8, 2, 3, "語法內容")
    ws8.column_dimensions['A'].width = 28
    ws8.column_dimensions['B'].width = 40
    ws8.column_dimensions['C'].width = 80

    # 模型名稱對照表（檔名 → 中文標題）
    MPLUS_MODEL_NAME_MAP = {
        'CFA A JCP DP CI':               '模型1（CFA-A）：JCP + DP + CI',
        'CFA B HP DP CI':                '模型2（CFA-B）：HP + DP + CI',
        'CFA C JCP PP DP CI':            '模型3（CFA-C）：JCP + PP + DP + CI',
        'CFA D HP PP DP CI':             '模型4（CFA-D）：HP + PP + DP + CI',
        'CFA E FiveFactor':              '模型5（CFA-E）：HP/JCP/PP/DP/CI 五因子',
        'CFA F FourFactor CP merged':    '模型6（CFA-F）：CP合併/PP/DP/CI 四因子',
        'CFA G ThreeFactor CP DP CI':    '模型7（CFA-G）：CP合併/DP/CI 三因子',
        'PATH ModMed':                      '調節中介路徑模型：PP(T1)調節 JCP/HP(T1)→DP(T2)→CI(T3)',
    }

    # 掃 run_dir 下所有 .inp（順序：CFA → MI → PATH → 其他）
    def _mplus_sort_key(fn):
        if fn.startswith('CFA_'):   return (0, fn)
        elif fn.startswith('MI_'):  return (1, fn)
        elif fn.startswith('PATH_'): return (2, fn)
        else:                       return (3, fn)

    all_inp_paths = []
    for fname in sorted(os.listdir(run_dir), key=_mplus_sort_key):
        if fname.endswith('.inp'):
            raw_lbl = fname.replace(f'_{ts}.inp', '').replace('_', ' ').replace('-', ' ')
            lbl = MPLUS_MODEL_NAME_MAP.get(raw_lbl, raw_lbl)
            all_inp_paths.append((lbl, os.path.join(run_dir, fname)))

    r8 = 3
    for label, inp_path in all_inp_paths:
        if not os.path.isfile(inp_path):
            continue
        try:
            with open(inp_path, 'r', encoding='utf-8', errors='replace') as fh:
                syntax = fh.read()
        except Exception:
            syntax = '（讀取失敗）'
        fname = os.path.basename(inp_path)
        c1 = ws8.cell(row=r8, column=1, value=label)
        c1.font = Font(bold=True, size=10)
        c1.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c1.border = bdr
        c2 = ws8.cell(row=r8, column=2, value=fname)
        c2.font = Font(size=9)
        c2.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c2.border = bdr
        c3 = ws8.cell(row=r8, column=3, value=syntax)
        c3.font = Font(name='Courier New', size=9)
        c3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c3.border = bdr
        ws8.row_dimensions[r8].height = max(80, syntax.count('\n') * 12)
        r8 += 1

    # ── 輔助：計算 CITC（校正後題項-總分相關）與刪題後 α ──────────
    def _citc_and_alpha_if_deleted(df_sub, all_cols, target_col):
        """
        CITC = 題目與「其餘題目總分」的 Pearson r
        Alpha_if_deleted = 刪掉此題後剩餘題目的 Cronbach's α
        """
        rest_cols = [c for c in all_cols if c != target_col]
        valid     = [c for c in rest_cols if c in df_sub.columns]
        if len(valid) < 2:
            return np.nan, np.nan
        data = df_sub[[target_col] + valid].dropna()
        if len(data) < 5:
            return np.nan, np.nan
        rest_sum = data[valid].sum(axis=1)
        citc = data[target_col].corr(rest_sum)
        # Alpha if deleted
        k   = len(valid)
        var_items = data[valid].var(axis=0, ddof=1).sum()
        var_total = data[valid].sum(axis=1).var(ddof=1)
        alpha_del = (k / (k - 1)) * (1 - var_items / var_total) if (k > 1 and var_total > 0) else np.nan
        return round(citc, 3), round(alpha_del, 3)

    def _cronbach(df_sub, cols):
        valid = [c for c in cols if c in df_sub.columns]
        data  = df_sub[valid].dropna()
        if len(data) < 2 or len(valid) < 2:
            return np.nan
        k = len(valid)
        var_sum   = data.var(axis=0, ddof=1).sum()
        var_total = data.sum(axis=1).var(ddof=1)
        return round((k / (k - 1)) * (1 - var_sum / var_total), 3) if var_total > 0 else np.nan

    # ── Sheet 7: 職涯階段補充分析（Bayesian Path）────────────────────
    ws_cs = wb.create_sheet("7_職涯階段分析")
    _BASE_KEY = 'PATH_Baseline'
    _CS_KEY   = 'PATH_CareerStage'
    _base_res = all_results.get(_BASE_KEY, {})
    _cs_res   = all_results.get(_CS_KEY, {})

    _CS_NCOL = 10  # 假設, 路徑說明, 時間點, 預期方向, IV, DV, β, SE, CI[lo,hi], 支持？
    # 計算各職涯階段 n（效果編碼用）
    _cs_age = df['Age'].replace(-999, np.nan)
    _n_exp   = int(((df['Age'] != -999) & (_cs_age >= 21) & (_cs_age <= 30)).sum())
    _n_maint = int(((df['Age'] != -999) & (_cs_age >= 41)).sum())
    _n_estab = int(((df['Age'] != -999) & (_cs_age >= 31) & (_cs_age <= 40)).sum())
    title(ws_cs,
          f"職涯階段路徑分析  N={n_total}  效果編碼：各階段 vs. 其他兩階段平均"
          f"  EXP=探索期(21-30,n={_n_exp})  ESTAB=建立期(31-40,n={_n_estab})  MAINT=維持期(41+,n={_n_maint})",
          end_col=_CS_NCOL)
    ws_cs.cell(row=2, column=1,
        value="上半部：基礎中介模型（ML + Bootstrap 5000，無調節）；"
              "下半部：職涯階段調節模型（Bayesian，10000次迭代）；"
              "β = STDYX；CI = BC Bootstrap 95%（基礎）/ HPD 95%（Bayesian）；顯著 = CI 不含 0"
    ).font = Font(italic=True, size=9)
    ws_cs.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_CS_NCOL)

    _csr = 3
    green_fill_cs = PatternFill("solid", fgColor="E2EFDA")
    red_fill_cs   = PatternFill("solid", fgColor="FFE0E0")

    def _cs_hdr(label, color="2E4057"):
        nonlocal _csr
        _sec = ws_cs.cell(row=_csr, column=1, value=label)
        _sec.font = Font(bold=True, size=11, color="FFFFFF")
        _sec.fill = PatternFill("solid", fgColor=color)
        ws_cs.merge_cells(start_row=_csr, start_column=1, end_row=_csr, end_column=_CS_NCOL)
        _csr += 1

    def _cs_col_hdr(cols):
        nonlocal _csr
        for ci, h in enumerate(cols, 1):
            if h: hdr(ws_cs, _csr, ci, h)
        _csr += 1

    def _base_path(key):
        d = _base_res.get('paths', {}).get(key, {})
        return (d.get('est', np.nan), d.get('se', np.nan),
                d.get('ci_lo', np.nan), d.get('ci_hi', np.nan),
                d.get('p', np.nan))

    def _cs_path(key):
        d = _cs_res.get('paths', {}).get(key, {})
        return (d.get('est', np.nan), d.get('sd', np.nan),
                d.get('ci_lo', np.nan), d.get('ci_hi', np.nan),
                d.get('sig', False))

    def _cs_row(cells, sig=False, fill=None):
        nonlocal _csr
        for ci, v in enumerate(cells, 1):
            c = ws_cs.cell(row=_csr, column=ci, value=v)
            c.font = Font(size=10, bold=(ci == len(cells) and sig),
                          color='006100' if sig else '000000')
            c.alignment = ctr if ci >= 3 else lft
            c.border = bdr
            if fill: c.fill = fill
        _csr += 1

    def _ci_fmt(lo, hi):
        if np.isnan(lo) or np.isnan(hi): return '—'
        return f"[{lo:.3f}, {hi:.3f}]"

    _STD_HDR = ["假設", "路徑說明", "時間點", "預期方向", "IV", "DV", "β (STDYX)", "SE", "95% BC CI", "支持？"]
    _IND_HDR = ["假設", "路徑說明", "條件／時間點", "預期方向", "IV", "DV", "β (indirect)", "SE", "95% BC CI", "支持？"]
    _BAY_HDR = ["假設", "路徑說明", "時間點", "預期方向", "IV", "DV", "β (post.mean)", "SD", "95% HPD CI", "支持？"]
    _BAY_IND = ["假設", "路徑說明", "條件／時間點", "預期方向", "IV", "DV", "β (indirect)", "SD", "95% HPD CI", "支持？"]

    def _verdict(sig, dir_ok, has_data):
        if not has_data: return '待計算'
        return '支持' if (sig and dir_ok) else '不支持'

    # ── 基礎中介模型結果 ──────────────────────────────────────────
    _cs_hdr("第一步：基礎中介模型（ML + Bootstrap 5000，無調節，全樣本 N=357）", color="4472C4")
    _cs_col_hdr(_STD_HDR)

    # (_hid, 路徑說明, 時間點, 預期方向, IV, DV, path_key)
    _base_rows = [
        ('H1a', 'JCP(T1)→DP(T2)（無調節）',  'T1→T2', '+', 'JCP（T1）', 'DP（T2）', 'JCP(T1)→DP(T2)'),
        ('H1b', 'HP(T1)→DP(T2)（無調節）',   'T1→T2', '+', 'HP（T1）',  'DP（T2）', 'HP(T1)→DP(T2)'),
        ('H3',  'DP(T2)→CI(T3)（無調節）',   'T2→T3', '+', 'DP（T2）',  'CI（T3）', 'DP(T2)→CI(T3)'),
        ('H5a', 'JCP(T1)→CI(T3) 直接效果',  'T1→T3', '+', 'JCP（T1）', 'CI（T3）', 'JCP(T1)→CI(T3) 直接'),
        ('H5b', 'HP(T1)→CI(T3) 直接效果',   'T1→T3', '+', 'HP（T1）',  'CI（T3）', 'HP(T1)→CI(T3) 直接'),
        ('',    'PP→DP（控制）',              'T1→T2', '', 'PP（T1）',  'DP（T2）', 'PP→DP (控制)'),
    ]
    for _hid, _desc, _tpt, _exp, _iv, _dv, _key in _base_rows:
        _est, _se, _lo, _hi, _pv = _base_path(_key)
        _star, _ = fmt_p(_pv) if not np.isnan(_pv) else ('', '')
        _sig = not np.isnan(_lo) and not (_lo <= 0 <= _hi)
        _dir_ok = (_exp == '' or
                   (_exp == '+' and not np.isnan(_est) and _est > 0) or
                   (_exp == '-' and not np.isnan(_est) and _est < 0))
        _support = _sig and _dir_ok and (_exp != '')
        _fill = green_fill_cs if _support else None
        _cs_row([
            _hid, _desc, _tpt, _exp, _iv, _dv,
            f"{_est:.3f}{_star}" if not np.isnan(_est) else '—',
            f"({_se:.3f})" if not np.isnan(_se) else '—',
            _ci_fmt(_lo, _hi),
            _verdict(_sig, _dir_ok, not np.isnan(_est)) if _exp != '' else '—',
        ], sig=_support, fill=_fill)

    # 間接效果
    _csr += 1
    _cs_hdr("基礎間接效果（BC Bootstrap 95% CI）", color="375623")
    _cs_col_hdr(_IND_HDR)
    for _hid_i, _pname, _desc_i, _iv_i, _dv_i in [
        ('', 'IND_JCP', 'JCP→DP→CI 間接效果（基礎中介）', 'JCP（T1）', 'CI（T3）'),
        ('', 'IND_HP',  'HP→DP→CI 間接效果（基礎中介）',  'HP（T1）',  'CI（T3）'),
    ]:
        _mc = _base_res.get('modconstr', {}).get(_pname, {})
        _est = _mc.get('est', np.nan)
        _se  = _mc.get('se',  np.nan)
        _lo  = _mc.get('ci_lo', np.nan)
        _hi  = _mc.get('ci_hi', np.nan)
        _sig = _mc.get('sig', False)
        _fill = green_fill_cs if _sig else None
        _cs_row([
            _hid_i, _desc_i, 'T1→T2→T3', '+', _iv_i, _dv_i,
            f"{_est:.3f}" if not np.isnan(_est) else '待填入',
            f"({_se:.3f})" if not np.isnan(_se) else '—',
            _ci_fmt(_lo, _hi),
            _verdict(_sig, True, not np.isnan(_est)),
        ], sig=_sig, fill=_fill)

    _csr += 1  # 空行隔開兩部分

    # ── JCP-only 模型（排除 HP，共線性確認）──────────────────────────
    _JCP_KEY = 'PATH_JCP_only'
    _jcp_res = all_results.get(_JCP_KEY, {})

    def _jcp_path(key):
        d = _jcp_res.get('paths', {}).get(key, {})
        return (d.get('est', np.nan), d.get('se', np.nan),
                d.get('ci_lo', np.nan), d.get('ci_hi', np.nan),
                d.get('p', np.nan))

    _cs_hdr("補充：JCP-only 模型（排除 HP，確認 JCP 獨立效果；ML + Bootstrap 5000）", color="7030A0")
    _cs_col_hdr(_STD_HDR)

    _jcp_rows = [
        ('H1a', 'JCP(T1)→DP(T2)（無HP）',   'T1→T2', '+', 'JCP（T1）', 'DP（T2）', 'JCP(T1)→DP(T2) [JCP-only]'),
        ('H3',  'DP(T2)→CI(T3)',             'T2→T3', '+', 'DP（T2）',  'CI（T3）', 'DP(T2)→CI(T3)'),
        ('H5a', 'JCP(T1)→CI(T3) 直接效果',  'T1→T3', '+', 'JCP（T1）', 'CI（T3）', 'JCP(T1)→CI(T3) 直接'),
        ('',    'PP→DP（控制）',              'T1→T2', '', 'PP（T1）',  'DP（T2）', 'PP→DP (控制)'),
    ]
    for _hid, _desc, _tpt, _exp, _iv, _dv, _key in _jcp_rows:
        _est, _se, _lo, _hi, _pv = _jcp_path(_key)
        _star, _ = fmt_p(_pv) if not np.isnan(_pv) else ('', '')
        _sig = not np.isnan(_lo) and not (_lo <= 0 <= _hi)
        _dir_ok = (_exp == '' or
                   (_exp == '+' and not np.isnan(_est) and _est > 0) or
                   (_exp == '-' and not np.isnan(_est) and _est < 0))
        _support = _sig and _dir_ok and (_exp != '')
        _fill = green_fill_cs if _support else None
        _cs_row([
            _hid, _desc, _tpt, _exp, _iv, _dv,
            f"{_est:.3f}{_star}" if not np.isnan(_est) else '—',
            f"({_se:.3f})" if not np.isnan(_se) else '—',
            _ci_fmt(_lo, _hi),
            _verdict(_sig, _dir_ok, not np.isnan(_est)) if _exp != '' else '—',
        ], sig=_support, fill=_fill)

    # JCP-only 間接效果
    _csr += 1
    _cs_hdr("JCP-only 間接效果（BC Bootstrap 95% CI）", color="6B3080")
    _cs_col_hdr(_IND_HDR)
    _jcp_mc = _jcp_res.get('modconstr', {}).get('IND_JCP', {})
    _est = _jcp_mc.get('est', np.nan)
    _se  = _jcp_mc.get('se',  np.nan)
    _lo  = _jcp_mc.get('ci_lo', np.nan)
    _hi  = _jcp_mc.get('ci_hi', np.nan)
    _sig = _jcp_mc.get('sig', False)
    _fill = green_fill_cs if _sig else None
    _cs_row([
        '', 'JCP→DP→CI 間接效果（無HP）', 'T1→T2→T3', '+', 'JCP（T1）', 'CI（T3）',
        f"{_est:.3f}" if not np.isnan(_est) else '待填入',
        f"({_se:.3f})" if not np.isnan(_se) else '—',
        _ci_fmt(_lo, _hi),
        _verdict(_sig, True, not np.isnan(_est)),
    ], sig=_sig, fill=_fill)

    _csr += 1  # 空行

    # ── JCP-only + PP 調節模型（確認低 PP 條件間接效果）────────────────
    _JCP_PP_KEY = 'PATH_JCP_PP'
    _jcp_pp_res = all_results.get(_JCP_PP_KEY, {})

    def _jcp_pp_path(key):
        d = _jcp_pp_res.get('paths', {}).get(key, {})
        return (d.get('est', np.nan), d.get('se', np.nan),
                d.get('ci_lo', np.nan), d.get('ci_hi', np.nan),
                d.get('p', np.nan))

    _cs_hdr("補充：JCP-only + PP 調節模型（確認低PP條件間接效果；無HP；ML + Bootstrap 5000）", color="9B2335")
    _cs_col_hdr(_STD_HDR)

    _jcp_pp_rows = [
        ('', 'JCP(T1)→DP(T2)（PP取均值）',  'T1→T2',    '+', 'JCP（T1）',  'DP（T2）', 'JCP(T1)→DP(T2) [at mean PP]'),
        ('', 'JCP×PP→DP（PP調節a-path）',    'T1×T1→T2', '-', 'JCP×PP（T1）','DP（T2）', 'JCP×PP→DP (a-path 調節)'),
        ('', 'DP(T2)→CI(T3)（PP取均值）',   'T2→T3',    '+', 'DP（T2）',   'CI（T3）', 'DP(T2)→CI(T3) [at mean PP]'),
        ('', 'DP×PP→CI（PP調節b-path）',     'T2×T1→T3', '-', 'DP×PP（T2）','CI（T3）', 'DP×PP→CI (b-path 調節)'),
        ('', 'JCP(T1)→CI(T3) 直接效果',     'T1→T3',    '+', 'JCP（T1）',  'CI（T3）', 'JCP(T1)→CI(T3) 直接'),
    ]
    for _hid, _desc, _tpt, _exp, _iv, _dv, _key in _jcp_pp_rows:
        _est, _se, _lo, _hi, _pv = _jcp_pp_path(_key)
        _star, _ = fmt_p(_pv) if not np.isnan(_pv) else ('', '')
        _sig = not np.isnan(_lo) and not (_lo <= 0 <= _hi)
        _dir_ok = (_exp == '' or
                   (_exp == '+' and not np.isnan(_est) and _est > 0) or
                   (_exp == '-' and not np.isnan(_est) and _est < 0))
        _support = _sig and _dir_ok and (_exp != '')
        _fill = green_fill_cs if _support else None
        _cs_row([
            _hid, _desc, _tpt, _exp, _iv, _dv,
            f"{_est:.3f}{_star}" if not np.isnan(_est) else '—',
            f"({_se:.3f})" if not np.isnan(_se) else '—',
            _ci_fmt(_lo, _hi),
            _verdict(_sig, _dir_ok, not np.isnan(_est)) if _exp != '' else '—',
        ], sig=_support, fill=_fill)

    # JCP-only + PP 條件間接效果
    _csr += 1
    _cs_hdr("JCP-only + PP 條件間接效果（BC Bootstrap 95% CI，PP ±1SD）", color="7A1C2A")
    _cs_col_hdr(_IND_HDR)
    for _mc_key, _pp_lbl in [('IND_HI', 'PP +1SD（主動型人格高）'), ('IND_LO', 'PP -1SD（主動型人格低）')]:
        _mc = _jcp_pp_res.get('modconstr', {}).get(_mc_key, {})
        _est = _mc.get('est', np.nan)
        _se  = _mc.get('se',  np.nan)
        _lo  = _mc.get('ci_lo', np.nan)
        _hi  = _mc.get('ci_hi', np.nan)
        _sig = _mc.get('sig', False)
        _fill = green_fill_cs if _sig else None
        _cs_row([
            '', 'JCP→DP→CI 條件間接效果', _pp_lbl, '+', 'JCP（T1）', 'CI（T3）',
            f"{_est:.3f}" if not np.isnan(_est) else '待填入',
            f"({_se:.3f})" if not np.isnan(_se) else '—',
            _ci_fmt(_lo, _hi),
            _verdict(_sig, True, not np.isnan(_est)),
        ], sig=_sig, fill=_fill)

    _csr += 1  # 空行

    # ── a-path 主效果 + 職涯階段調節 ──
    _cs_hdr("第二步：職涯階段調節 a-path（Bayesian，效果編碼：各階段 vs. 其他兩階段平均）")
    _cs_col_hdr(_BAY_HDR)

    # (_hid, 路徑說明, 時間點, 預期方向, IV, DV, path_key)
    _cs_a_rows = [
        ('H1a', 'JCP(T1)→DP(T2)（所有階段平均）',          'T1→T2',    '+', 'JCP（T1）',    'DP（T2）', 'JCP(T1)→DP(T2) [ESTAB均值]'),
        ('H1b', 'HP(T1)→DP(T2)（所有階段平均）',           'T1→T2',    '+', 'HP（T1）',     'DP（T2）', 'HP(T1)→DP(T2) [ESTAB均值]'),
        ('H2a', 'JCP×EXP_C→DP（探索期 vs 其他兩期平均）',  'T1×T1→T2', '+', 'JCP×EXP_C',   'DP（T2）', 'JCP×EXP_C→DP (探索期效果)'),
        ('H2b', 'JCP×MAINT_C→DP（維持期 vs 其他兩期平均）','T1×T1→T2', '-', 'JCP×MAINT_C', 'DP（T2）', 'JCP×MAINT_C→DP (維持期效果)'),
        ('H3a', 'HP×EXP_C→DP（探索期 vs 其他兩期平均）',   'T1×T1→T2', '?', 'HP×EXP_C',    'DP（T2）', 'HP×EXP_C→DP (探索期效果)'),
        ('H3b', 'HP×MAINT_C→DP（維持期 vs 其他兩期平均）', 'T1×T1→T2', '?', 'HP×MAINT_C',  'DP（T2）', 'HP×MAINT_C→DP (維持期效果)'),
    ]
    for _hid, _desc, _tpt, _exp, _iv, _dv, _key in _cs_a_rows:
        _est, _sd, _lo, _hi, _sig = _cs_path(_key)
        _dir_ok = (_exp == '?' or
                   (_exp == '+' and not np.isnan(_est) and _est > 0) or
                   (_exp == '-' and not np.isnan(_est) and _est < 0))
        _support = _sig and _dir_ok and (_exp != '?')
        _fill = green_fill_cs if _support else (red_fill_cs if not np.isnan(_est) and _sig and not _dir_ok else None)
        _cs_row([
            _hid, _desc, _tpt, _exp, _iv, _dv,
            f"{_est:.3f}" if not np.isnan(_est) else '—',
            f"({_sd:.3f})" if not np.isnan(_sd) else '—',
            _ci_fmt(_lo, _hi),
            _verdict(_sig, _dir_ok, not np.isnan(_est)) if _exp != '?' else ('顯著' if _sig else '不顯著'),
        ], sig=_support, fill=_fill)

    # ── b-path + 直接效果 ──
    _csr += 1
    _cs_hdr("b-path / 直接效果：DP(T2)→CI(T3)，JCP/HP→CI 直接")
    _cs_col_hdr(_BAY_HDR)

    _cs_b_rows = [
        ('H3',   'DP(T2)→CI(T3)',         'T2→T3', '+', 'DP（T2）',  'CI（T3）', 'DP(T2)→CI(T3)'),
        ('H5a',  'JCP(T1)→CI(T3) 直接效果','T1→T3', '+', 'JCP（T1）', 'CI（T3）', 'JCP(T1)→CI(T3) 直接'),
        ('H5b',  'HP(T1)→CI(T3) 直接效果', 'T1→T3', '+', 'HP（T1）',  'CI（T3）', 'HP(T1)→CI(T3) 直接'),
    ]
    for _hid, _desc, _tpt, _exp, _iv, _dv, _key in _cs_b_rows:
        _est, _sd, _lo, _hi, _sig = _cs_path(_key)
        _dir_ok = (_exp == '+' and not np.isnan(_est) and _est > 0) or \
                  (_exp == '-' and not np.isnan(_est) and _est < 0)
        _support = _sig and _dir_ok
        _fill = green_fill_cs if _support else None
        _cs_row([
            _hid, _desc, _tpt, _exp, _iv, _dv,
            f"{_est:.3f}" if not np.isnan(_est) else '—',
            f"({_sd:.3f})" if not np.isnan(_sd) else '—',
            _ci_fmt(_lo, _hi),
            _verdict(_sig, _dir_ok, not np.isnan(_est)),
        ], sig=_support, fill=_fill)

    # ── 條件間接效果 ──
    _csr += 1
    _cs_hdr("條件間接效果：JCP/HP→DP→CI 依職涯階段（Bayesian MODEL CONSTRAINT，效果編碼，95% HPD CI）")
    _cs_col_hdr(_BAY_IND)

    _cs_mc = _cs_res.get('modconstr', {})
    _cs_ind_specs = [
        ('H4a', 'JCP→DP→CI 條件間接', f'探索期（EXP，n={_n_exp}）',    'JCP（T1）', 'CI（T3）', 'IE_E_J'),
        ('H4a', 'JCP→DP→CI 條件間接', f'建立期（ESTAB，n={_n_estab}）', 'JCP（T1）', 'CI（T3）', 'IE_R_J'),
        ('H4a', 'JCP→DP→CI 條件間接', f'維持期（MAINT，n={_n_maint}）', 'JCP（T1）', 'CI（T3）', 'IE_M_J'),
        ('H4b', 'HP→DP→CI 條件間接',  f'探索期（EXP，n={_n_exp}）',    'HP（T1）',  'CI（T3）', 'IE_E_H'),
        ('H4b', 'HP→DP→CI 條件間接',  f'建立期（ESTAB，n={_n_estab}）', 'HP（T1）',  'CI（T3）', 'IE_R_H'),
        ('H4b', 'HP→DP→CI 條件間接',  f'維持期（MAINT，n={_n_maint}）', 'HP（T1）',  'CI（T3）', 'IE_M_H'),
    ]
    for _hid, _path_lbl, _stage, _iv_i, _dv_i, _key in _cs_ind_specs:
        _mc = _cs_mc.get(_key, {})
        _est = _mc.get('est', np.nan)
        _sd  = _mc.get('sd',  np.nan)
        _lo  = _mc.get('ci_lo', np.nan)
        _hi  = _mc.get('ci_hi', np.nan)
        _sig = _mc.get('sig', False)
        _fill = green_fill_cs if _sig else None
        _cs_row([
            _hid, _path_lbl, _stage, '+', _iv_i, _dv_i,
            f"{_est:.3f}" if not np.isnan(_est) else '待填入',
            f"({_sd:.3f})" if not np.isnan(_sd) else '—',
            _ci_fmt(_lo, _hi),
            _verdict(_sig, True, not np.isnan(_est)),
        ], sig=_sig, fill=_fill)

    set_widths(ws_cs, [('A', 10), ('B', 40), ('C', 30), ('D', 16),
                       ('E', 18), ('F', 18), ('G', 10), ('H', 18), ('I', 16), ('J', 10)])

    # ── Sheet 8: SPSS 語法（.sps 內容）── 語法附錄 ───────────────────
    ws_spss = wb.create_sheet("8_SPSS語法")
    ws_spss.cell(row=1, column=1,
                 value="SPSS 分析語法（.sps）— 可直接複製至 SPSS 語法視窗執行"
                 ).font = Font(bold=True, size=12)
    ws_spss.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    hdr(ws_spss, 2, 1, "用途說明")
    hdr(ws_spss, 2, 2, "檔案名稱")
    hdr(ws_spss, 2, 3, "語法內容")
    ws_spss.column_dimensions['A'].width = 30
    ws_spss.column_dimensions['B'].width = 44
    ws_spss.column_dimensions['C'].width = 80

    # 依執行順序列出所有 .sps 檔
    spss_file_specs = [
        (f"SPSS_Syntax_{ts}.sps",   "步驟1：匯入資料 + 變數標籤（先執行）"),
        (f"SPSS_Analysis_{ts}.sps", "步驟2：完整分析（描述統計/相關/CMV/信度/CITC/題目間相關/t檢定）"),
    ]
    r_spss = 3
    for sps_fname, sps_label in spss_file_specs:
        sps_path = os.path.join(run_dir, sps_fname)
        if not os.path.isfile(sps_path):
            continue
        try:
            with open(sps_path, 'r', encoding='utf-8-sig', errors='replace') as fh:
                sps_syntax = fh.read()
        except Exception:
            sps_syntax = '（讀取失敗）'
        c1 = ws_spss.cell(row=r_spss, column=1, value=sps_label)
        c1.font = Font(bold=True, size=10)
        c1.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c1.border = bdr
        c2 = ws_spss.cell(row=r_spss, column=2, value=sps_fname)
        c2.font = Font(size=9)
        c2.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c2.border = bdr
        c3 = ws_spss.cell(row=r_spss, column=3, value=sps_syntax)
        c3.font = Font(name='Courier New', size=9)
        c3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c3.border = bdr
        ws_spss.row_dimensions[r_spss].height = max(80, sps_syntax.count('\n') * 12)
        r_spss += 1

    # ── CFA 因素負荷量（附加在 5_CFA分析 下方，模型1/2/5）───────────
    _ldg_start = ws4.max_row + 2
    _ldg_sec = ws4.cell(row=_ldg_start, column=1,
        value=f"CFA 標準化因素負荷量（STDYX，T1，N = {n_total}）— 模型1/2/5")
    _ldg_sec.font = Font(bold=True, size=12)
    ws4.merge_cells(start_row=_ldg_start, start_column=1,
                    end_row=_ldg_start, end_column=7)
    _ldg_start += 1
    _ldg_note = ws4.cell(row=_ldg_start, column=1,
        value="判斷標準：λ ≥ 0.50 為可接受；< 0.40 建議刪題（標示橘色）；< 0.50 邊緣（標示黃色）；p < .05 顯著")
    _ldg_note.font = Font(italic=True, size=9, color='C00000')
    ws4.merge_cells(start_row=_ldg_start, start_column=1,
                    end_row=_ldg_start, end_column=7)
    _ldg_start += 1
    for ci, h in enumerate(["因子", "題目", "β (STDYX)", "SE", "z 值", "p 值", "建議"], 1):
        hdr(ws4, _ldg_start, ci, h)
    _ldg_start += 1
    ws12 = ws4  # alias so _write_loading_section code below is unchanged

    orange_fill = PatternFill("solid", fgColor="FFB347")
    yellow_fill = PatternFill("solid", fgColor="FFFF99")
    blue_fill   = PatternFill("solid", fgColor="BDD7EE")

    factor_label = {
        'HP': 'HP 階層停滯', 'JCP': 'JCP 工作內容停滯',
        'PP': 'PP 主動型人格', 'DP': 'DP 決策拖延', 'CI': 'CI 職涯無所作為',
    }

    def _write_loading_section(ws, r_start, model_key, section_title):
        """Write one CFA model's loadings block; return next row."""
        r = r_start
        loadings = all_results.get(model_key, {}).get('loadings', [])
        # Section header
        sec = ws.cell(row=r, column=1, value=section_title)
        sec.font = Font(bold=True, size=11)
        sec.fill = blue_fill
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 1
        if not loadings:
            ws.cell(row=r, column=1,
                    value=f"（{model_key} 尚未執行或無法解析，請確認 .out 檔）"
                    ).font = Font(italic=True, color='808080')
            r += 2
            return r
        prev_factor = None
        for row_d in loadings:
            fac  = row_d['factor']
            beta = row_d['beta']
            p    = row_d['p']
            if fac != prev_factor:
                sep = ws.cell(row=r, column=1, value=factor_label.get(fac, fac))
                sep.font = Font(bold=True, size=10)
                sep.fill = PatternFill("solid", fgColor="D9E1F2")
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
                r += 1
                prev_factor = fac
            sig_star = '***' if p < .001 else ('**' if p < .01 else ('*' if p < .05 else ''))
            beta_str = f"{beta:.3f}{sig_star}"
            if abs(beta) < 0.40:
                verdict = '⚠ 建議刪題'
                row_fill = orange_fill
            elif abs(beta) < 0.50:
                verdict = '△ 邊緣'
                row_fill = yellow_fill
            else:
                verdict = '✅ 保留'
                row_fill = None
            vals = [fac, row_d['item'], beta_str,
                    f"{row_d['se']:.3f}", f"{row_d['z']:.3f}",
                    f"{p:.3f}", verdict]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.alignment = Alignment(horizontal='center')
                c.border = bdr
                if row_fill:
                    c.fill = row_fill
            r += 1
        r += 1  # blank separator
        return r

    r12 = _ldg_start
    r12 = _write_loading_section(ws12, r12,
        'CFA-A (JCP+DP+CI)',
        '模型1（CFA-A）：JCP + DP + CI 三因子標準化因素負荷量')
    r12 = _write_loading_section(ws12, r12,
        'CFA-B (HP+DP+CI)',
        '模型2（CFA-B）：HP + DP + CI 三因子標準化因素負荷量')
    r12 = _write_loading_section(ws12, r12,
        'CFA-E (HP+JCP+PP+DP+CI, 5F)',
        '模型5（CFA-E）：HP/JCP/PP/DP/CI 五因子標準化因素負荷量（區別效度基準）')
    r12 = _write_loading_section(ws12, r12,
        'CFA-H (Cross-Wave 5F)',
        '模型8（CFA-H）：跨波次五因子因素負荷量（HP/JCP/PP=T1，DP=T2，CI=T3）')
    r12 = _write_loading_section(ws12, r12,
        'CFA-I (Cross-Wave 4F, no PP)',
        '模型9（CFA-I）：跨波次四因子因素負荷量（HP/JCP=T1，DP=T2，CI=T3，不含 PP）')

    set_widths(ws4, [('A', 22), ('B', 10), ('C', 16), ('D', 8),
                     ('E', 9),  ('F', 28), ('G', 14)])

    # ── 績效考核分析（T1/T2/T3 + 三波相關，同一張表）────────────────
    # 同類表格集中在同一 sheet，依論文章節邏輯分區呈現
    _ws_pm = wb.create_sheet("2_績效考核分析")
    title(_ws_pm, f"績效考核背景變項分析（三波次，N = {n_total}）", end_col=13)
    _pr = 3
    _sec_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')

    _pm_waves = [
        ('T1', '第一波（T1）', {
            'has': 'PM_Has_T1', 'supervisor': 'PM_Supervisor_T1',
            'self': 'PM_Self_T1', 'interview': 'PM_Interview_T1',
            'other': 'PM_Other_T1', 'result': 'PM_Result_T1', 'help': 'PM_Help_T1',
        }),
        ('T2', '第二波（T2）', {
            'has': 'PM_Has_T2', 'supervisor': 'PM_Supervisor_T2',
            'self': 'PM_Self_T2', 'interview': 'PM_Interview_T2',
            'other': 'PM_Other_T2', 'result': 'PM_Result_T2', 'help': 'PM_Help_T2',
        }),
        ('T3', '第三波（T3）', {
            'has': 'PM_Has_T3', 'supervisor': 'PM_Supervisor_T3',
            'self': 'PM_Self_T3', 'interview': 'PM_Interview_T3',
            'other': 'PM_Other_T3', 'result': 'PM_Result_T3', 'help': 'PM_Help_T3',
        }),
    ]
    for _wave, _wave_lbl, _cols in _pm_waves:
        # section header (shaded)
        _sec_cell = _ws_pm.cell(row=_pr, column=1, value=_wave_lbl)
        _sec_cell.font = Font(bold=True, size=11)
        _sec_cell.fill = _sec_fill
        _ws_pm.merge_cells(start_row=_pr, start_column=1, end_row=_pr, end_column=4)
        for _ci in range(1, 5):
            _ws_pm.cell(row=_pr, column=_ci).fill = _sec_fill
        _pr += 1
        for _ci, _h in enumerate(["變項", "類別／統計量", "人數／數值", "%"], 1):
            hdr(_ws_pm, _pr, _ci, _h)
        _pr += 1

        _has_s = pd.to_numeric(df.get(_cols['has'], pd.Series(dtype=float)), errors='coerce')
        _n_has = int((_has_s == 1).sum())
        _n_not = int((_has_s == 0).sum())

        cell(_ws_pm, _pr, 1, '是否有績效考核', bold=True, align='left')
        for _ci in range(2, 5): cell(_ws_pm, _pr, _ci, '')
        _pr += 1
        for _lbl, _cnt in [('有', _n_has), ('無', _n_not)]:
            _pct = _cnt / n_total * 100 if n_total > 0 else 0
            cell(_ws_pm, _pr, 1, ''); cell(_ws_pm, _pr, 2, _lbl, align='left')
            cell(_ws_pm, _pr, 3, _cnt); cell(_ws_pm, _pr, 4, f"{_pct:.1f}%")
            _pr += 1

        cell(_ws_pm, _pr, 1, '考核形式（有考核者，可複選）', bold=True, align='left')
        for _ci in range(2, 5): cell(_ws_pm, _pr, _ci, '')
        _pr += 1
        _base = _n_has if _n_has > 0 else 1
        for _fkey, _flbl in [('supervisor','主管評核'),('self','自我評核'),
                               ('interview','績效面談'),('other','其他')]:
            _fs = pd.to_numeric(df.get(_cols[_fkey], pd.Series(dtype=float)), errors='coerce')
            _fn = int((_fs == 1).sum())
            cell(_ws_pm, _pr, 1, ''); cell(_ws_pm, _pr, 2, _flbl, align='left')
            cell(_ws_pm, _pr, 3, _fn); cell(_ws_pm, _pr, 4, f"{_fn/_base*100:.1f}%")
            _pr += 1

        _res_s = pd.to_numeric(df.get(_cols['result'], pd.Series(dtype=float)), errors='coerce')
        cell(_ws_pm, _pr, 1, '考核結果性質', bold=True, align='left')
        for _ci in range(2, 5): cell(_ws_pm, _pr, _ci, '')
        _pr += 1
        for _val, _lbl in [(1,'負向'),(2,'中立／持平'),(3,'正向')]:
            _cnt = int((_res_s == _val).sum())
            _pct = _cnt / n_total * 100 if n_total > 0 else 0
            cell(_ws_pm, _pr, 1, ''); cell(_ws_pm, _pr, 2, _lbl, align='left')
            cell(_ws_pm, _pr, 3, _cnt); cell(_ws_pm, _pr, 4, f"{_pct:.1f}%")
            _pr += 1

        _hlp_s = pd.to_numeric(df.get(_cols['help'], pd.Series(dtype=float)), errors='coerce').dropna()
        cell(_ws_pm, _pr, 1, '考核對職涯幫助程度（1–5）', bold=True, align='left')
        for _ci in range(2, 5): cell(_ws_pm, _pr, _ci, '')
        _pr += 1
        if len(_hlp_s) > 0:
            cell(_ws_pm, _pr, 1, ''); cell(_ws_pm, _pr, 2, 'M (SD)', align='left')
            cell(_ws_pm, _pr, 3, f"{_hlp_s.mean():.2f} ({_hlp_s.std():.2f})"); cell(_ws_pm, _pr, 4, '')
            _pr += 1
            cell(_ws_pm, _pr, 1, ''); cell(_ws_pm, _pr, 2, '最小值 ~ 最大值', align='left')
            cell(_ws_pm, _pr, 3, f"{_hlp_s.min():.0f} ~ {_hlp_s.max():.0f}"); cell(_ws_pm, _pr, 4, '')
            _pr += 1
        _pr += 1  # blank separator between waves

    # 注腳
    _ws_pm.cell(row=_pr, column=1,
        value="考核形式百分比以「有績效考核」者為分母（可複選）；結果性質及幫助程度以全樣本為分母"
    ).font = Font(italic=True, size=9)
    _ws_pm.merge_cells(start_row=_pr, start_column=1, end_row=_pr, end_column=4)
    _pr += 2

    # ── 三波次相關矩陣（接在同一 sheet 下方）────────────────────────
    _sec2 = _ws_pm.cell(row=_pr, column=1, value="三波次績效考核主要變項相關矩陣")
    _sec2.font = Font(bold=True, size=11)
    _sec2.fill = _sec_fill
    _ws_pm.merge_cells(start_row=_pr, start_column=1, end_row=_pr, end_column=13)
    for _ci in range(1, 14):
        _ws_pm.cell(row=_pr, column=_ci).fill = _sec_fill
    _pr += 1

    _pmcorr_vars = [
        ('PM_Has_T1',    'T1 有績效考核'),
        ('PM_Result_T1', 'T1 考核結果（1–3）'),
        ('PM_Help_T1',   'T1 考核幫助程度'),
        ('PM_Has_T2',    'T2 有績效考核'),
        ('PM_Result_T2', 'T2 考核結果（1–3）'),
        ('PM_Help_T2',   'T2 考核幫助程度'),
        ('PM_Has_T3',    'T3 有績效考核'),
        ('PM_Result_T3', 'T3 考核結果（1–3）'),
        ('PM_Help_T3',   'T3 考核幫助程度'),
    ]
    _nv = len(_pmcorr_vars)
    _pmcorr_hdrs = ['#', '變項', 'M', 'SD'] + [str(i+1) for i in range(_nv)]
    for _ci, _h in enumerate(_pmcorr_hdrs, 1):
        hdr(_ws_pm, _pr, _ci, _h)
    _pr += 1

    _pm_series = [pd.to_numeric(df.get(_col, pd.Series(dtype=float)), errors='coerce')
                  for _col, _ in _pmcorr_vars]
    for _i, ((_col, _lbl), _si) in enumerate(zip(_pmcorr_vars, _pm_series)):
        cell(_ws_pm, _pr, 1, _i + 1)
        cell(_ws_pm, _pr, 2, _lbl, align='left')
        cell(_ws_pm, _pr, 3, f"{_si.mean():.2f}" if _si.notna().sum() > 0 else '—')
        cell(_ws_pm, _pr, 4, f"{_si.std():.2f}"  if _si.notna().sum() > 1 else '—')
        for _j, _sj in enumerate(_pm_series):
            if _j < _i:
                _vld = pd.concat([_si, _sj], axis=1).dropna()
                if len(_vld) > 2:
                    _rv = _vld.iloc[:, 0].corr(_vld.iloc[:, 1])
                    _clr = ('C00000' if abs(_rv) >= .5 else ('996633' if abs(_rv) >= .3 else '000000'))
                    cell(_ws_pm, _pr, 5 + _j, f"{_rv:.2f}", color=_clr)
                else:
                    cell(_ws_pm, _pr, 5 + _j, '—')
            elif _j == _i:
                cell(_ws_pm, _pr, 5 + _j, '—')
        _pr += 1

    _ws_pm.cell(row=_pr, column=1,
        value="下三角相關係數；|r| ≥ .50 紅色，|r| ≥ .30 棕色；成對刪除法處理遺漏值"
    ).font = Font(italic=True, size=9)
    _ws_pm.merge_cells(start_row=_pr, start_column=1, end_row=_pr, end_column=_nv + 4)
    set_widths(_ws_pm, [('A', 30), ('B', 26), ('C', 12), ('D', 10)] +
               [(chr(ord('E') + i), 7) for i in range(_nv)])

    # ── NoPP 版本：移除職涯階段分析 Sheet ──────────────────────────
    if variant_label == 'nopp' and "7_職涯階段分析" in [s.title for s in wb._sheets]:
        del wb["7_職涯階段分析"]

    # ── 依論文研究方法章節順序排列索引標籤 ──────────────────────────
    # SPSS分析在前，Mplus分析在後，語法附錄最後
    if variant_label == 'nopp':
        _desired_sheet_order = [
            "1_背景變項",
            "2_績效考核分析",
            "3_敘述統計與信度",
            "4_相關矩陣",
            "5_CFA分析",
            "6_假設檢驗",
            "8_SPSS語法",
            "9_Mplus語法",
        ]
    else:
        _desired_sheet_order = [
            "1_背景變項",          # SPSS: 人口統計
            "2_績效考核分析",      # SPSS: 績效考核（三波 + 相關）
            "3_敘述統計與信度",    # SPSS: M/SD/α
            "4_相關矩陣",          # Python: 三波段追蹤相關矩陣（15×15）
            "5_CFA分析",           # Mplus: CFA適配 + 負荷量 + AVE/CR
            "6_假設檢驗",          # Mplus: 完整調節中介 H1~H7
            "7_職涯階段分析",      # Mplus Bayes: 職涯階段補充分析
            "8_SPSS語法",          # 附錄：SPSS語法
            "9_Mplus語法",         # 附錄：Mplus語法（CFA + PATH）
        ]
    _existing_names = [s.title for s in wb._sheets]
    _ordered = []
    for _sname in _desired_sheet_order:
        if _sname in _existing_names:
            _ordered.append(wb[_sname])
    # Append any sheets not in the desired list (preserve them at end)
    for _s in wb._sheets:
        if _s not in _ordered:
            _ordered.append(_s)
    wb._sheets = _ordered

    # ── 儲存 ─────────────────────────────────────────────────────
    if variant_label == 'nopp':
        excel_path = os.path.join(run_dir, f"Thesis_NoPP_{ts}.xlsx")
    else:
        vtag = ('_' + variant_label.replace(' ', '').replace('+', '_')) if variant_label else ''
        excel_path = os.path.join(run_dir, f"Thesis_Results{vtag}_{ts}.xlsx")
    try:
        wb.save(excel_path)
        print(f"  [OK] Excel 報告已儲存：{excel_path}")
    except Exception as e:
        print(f"  [錯誤] Excel 儲存失敗：{e}")
        return None
    return excel_path


# ==========================================
# MODULE G: Analysis Summary .md（根目錄）
# ==========================================
def generate_analysis_summary_md(run_dir, ts, g3_sample, alpha_dict, corr_dict,
                                  all_results, attrition_md, desc_md,
                                  variant=None):
    """
    產生完整 Pipeline_Master_Report_{ts}.md 至 run_dir。
    variant='nopp': 輸出 Report_NoPP_{ts}.md，使用 PATH_NoPP 結果，跳過 7/8/9 補充分析。
    """
    is_nopp = (variant == 'nopp')
    date_tag = ts[:8]
    n = len(g3_sample) if g3_sample is not None else 357

    # ── 輔助：格式化單一路徑結果 ─────────────────────────────────────────────
    def _v(paths, label, key='est', na='—'):
        d = paths.get(label, {})
        return f"{d[key]:.3f}" if (d and key in d and d[key] is not None) else na

    def _ci(paths, label, na='—'):
        d = paths.get(label, {})
        if d and 'ci_lo' in d and 'ci_hi' in d:
            return f"[{d['ci_lo']:.3f}, {d['ci_hi']:.3f}]"
        return na

    def _pval(paths, label, na='—'):
        d = paths.get(label, {})
        if not d or 'p' not in d: return na
        p = d['p']
        if p < 0.001: return '<.001'
        return f'.{round(p*1000):03d}'[:4].lstrip('0') or '.000'

    def _sup(paths, label, expected_dir='+'):
        d = paths.get(label, {})
        if not d: return '（待執行）'
        ci_lo = d.get('ci_lo'); ci_hi = d.get('ci_hi')
        p = d.get('p', 1.0); est = d.get('est', 0)
        if ci_lo is not None and ci_hi is not None:
            sig = not (ci_lo <= 0 <= ci_hi)
        else:
            sig = p < 0.05
        if not sig: return '✗ 不支持'
        if expected_dir == '+' and est > 0: return '✓ 支持'
        if expected_dir == '-' and est < 0: return '✓ 支持'
        return '✗（方向不符）'

    def _mc_v(mc, key, val_key='est', na='—'):
        d = mc.get(key, {})
        return f"{d[val_key]:.3f}" if (d and val_key in d) else na

    def _mc_ci(mc, key, na='—'):
        d = mc.get(key, {})
        if d and 'ci_lo' in d and 'ci_hi' in d:
            return f"[{d['ci_lo']:.3f}, {d['ci_hi']:.3f}]"
        return na

    def _mc_sup(mc, key):
        d = mc.get(key, {})
        if not d: return '（待執行）'
        ci_lo = d.get('ci_lo'); ci_hi = d.get('ci_hi')
        if ci_lo is not None and ci_hi is not None:
            return '✗ 不顯著' if (ci_lo <= 0 <= ci_hi) else '✓ 顯著'
        return '—'

    # ── PATH ModMed 結果 ─────────────────────────────────────────────────────
    _pm = all_results.get('PATH (T1→T2→T3)', {})
    pm_paths = _pm.get('paths', {})
    pm_mc    = _pm.get('modconstr', {})

    # ── PATH Baseline 結果 ───────────────────────────────────────────────────
    _bl = all_results.get('PATH_Baseline', {})
    bl_paths = _bl.get('paths', {})
    bl_mc    = _bl.get('modconstr', {})

    # ── PATH JCP-only ────────────────────────────────────────────────────────
    _jcp = all_results.get('PATH_JCP_only', {})
    jcp_paths = _jcp.get('paths', {})
    jcp_mc    = _jcp.get('modconstr', {})

    # ── PATH JCP+PP ──────────────────────────────────────────────────────────
    _jpp = all_results.get('PATH_JCP_PP', {})
    jpp_paths = _jpp.get('paths', {})
    jpp_mc    = _jpp.get('modconstr', {})

    # ── PATH CareerStage（Bayesian）──────────────────────────────────────────
    _cs = all_results.get('PATH_CareerStage', {})
    cs_paths = _cs.get('paths', {})
    cs_mc    = _cs.get('modconstr', {})

    # ── 績效考核統計（從 g3_sample 計算）────────────────────────────────────
    def _pm_stats(wave):
        """回傳 (has_pct, neg_n, neu_n, pos_n, help_m, help_sd)"""
        has_col    = f'PM_Has_T{wave}'
        res_col    = f'PM_Result_T{wave}'
        help_col   = f'PM_Help_T{wave}'
        if g3_sample is None or has_col not in g3_sample.columns:
            return ('—', '—', '—', '—', '—', '—')
        df = g3_sample
        has_pct = f"{df[has_col].mean()*100:.1f}%" if has_col in df.columns else '—'
        if res_col in df.columns:
            neg = int((df[res_col] < 3).sum())
            neu = int((df[res_col] == 3).sum())
            pos = int((df[res_col] > 3).sum())
        else:
            neg = neu = pos = '—'
        if help_col in df.columns:
            hm = f"{df[help_col].mean():.2f}"
            hs = f"{df[help_col].std():.2f}"
        else:
            hm = hs = '—'
        return (has_pct, neg, neu, pos, f"{hm} ({hs})")

    # ════════════════════════════════════════════════════════════════════════
    lines = []
    A = lines.append

    if is_nopp:
        A(f"# 分析報告【不含 PP 版本】純中介模型（產生時間: {ts}）")
    else:
        A(f"# 全階段資料分析自動化整合報告（產生時間: {ts}）")
    A(f"> 資料：三波配對樣本 N = {n}（Group = 3）")
    A("")
    A("---")
    A("")

    # ── 一、研究設計概述 ──────────────────────────────────────────────────────
    A("## 一、研究設計概述")
    A("")
    A("- **縱貫三波**：T1（預測變項）→ T2（中介）→ T3（結果）")
    A("- **主要路徑**：HP / JCP (T1) → 決策拖延 DP (T2) → 職涯無所作為 CI (T3)")
    if is_nopp:
        A("- **版本說明**：本版本移除 PP（主動型人格）調節，檢驗純中介主路徑（H1~H4）")
    else:
        A("- **調節變項**：主動型人格 PP (T1)，調節 a-path 與 b-path")
        A("- **補充分析**：職涯階段（探索／建立／維護）Bayesian 調節")
    A("")
    A("---")
    A("")

    # ── 二、樣本流失分析 ─────────────────────────────────────────────────────
    # attrition_md 已由 analyze_attrition() 格式化，直接嵌入並調整標題層級
    attrition_section = attrition_md.replace(
        "## 1. 樣本流失分析 (Attrition Analysis)",
        "## 二、樣本流失分析（Attrition Analysis）"
    )
    A(attrition_section.strip())
    A("")
    A("---")
    A("")

    # ── 三、敘述統計與信度分析 ＋ 四、相關矩陣 ──────────────────────────────
    desc_section = desc_md
    desc_section = desc_section.replace(
        "## 2. 敘述性統計與信度分析 (Descriptives & Reliability)",
        "## 三、敘述統計與信度分析（N = {n}）".format(n=n)
    )
    # 移除 desc_md 中的 T1 相關矩陣區塊（改由下方完整版取代）
    _corr_sep = "\n## 3. 相關矩陣 (Correlation Matrix)"
    if _corr_sep in desc_section:
        desc_section = desc_section[:desc_section.index(_corr_sep)]
    A(desc_section.strip())
    A("")

    # ── 四、三波段追蹤相關矩陣（完整，12×12 + 15×15）──────────────────────────
    A(f"## 四、三波段追蹤相關矩陣（N = {n}）")
    A("")
    A("> 對角線括號內為 Cronbach's α；*** p<.001  ** p<.01  * p<.05；成對刪除法處理遺漏值")
    A("")

    _df_md = g3_sample if g3_sample is not None else pd.DataFrame()

    def _md_compute_composites(iv_defs, df_data):
        composites, alphas = [], []
        for sn, wv, lbl, ni in iv_defs:
            cols = [f'{sn}{k}_{wv}' for k in range(1, ni+1)]
            vcols = [c for c in cols if c in df_data.columns]
            if vcols:
                mat = df_data[vcols].apply(pd.to_numeric, errors='coerce')
                composites.append(mat.mean(axis=1))
                alphas.append(calculate_cronbach_alpha(mat.dropna()))
            else:
                composites.append(pd.Series([np.nan]*len(df_data)))
                alphas.append(np.nan)
        return composites, alphas

    def _md_corr_table(iv_defs, composites, alphas):
        _n = len(iv_defs)
        hdr = "| 變數 | " + " | ".join(str(i+1) for i in range(_n)) + " | M | SD |"
        sep = "|---|" + "---|"*_n + "---|---|"
        rows = [hdr, sep]
        for i, (sn, wv, lbl, ni) in enumerate(iv_defs):
            cells = []
            for j in range(_n):
                if i == j:
                    a = alphas[i]
                    cells.append(f"({a:.3f})" if not np.isnan(a) else "(—)")
                elif i > j:
                    vd = pd.concat([composites[i], composites[j]], axis=1).dropna()
                    if len(vd) > 2:
                        rv, pv = stats.pearsonr(vd.iloc[:,0], vd.iloc[:,1])
                        star, _ = fmt_p(pv)
                        cells.append(f"{rv:.2f}{star}")
                    else:
                        cells.append("—")
                else:
                    cells.append(" ")
            mv = composites[i].mean() if composites[i].notna().sum() > 1 else np.nan
            sv = composites[i].std()  if composites[i].notna().sum() > 1 else np.nan
            rows.append(f"| {lbl} | " + " | ".join(cells) +
                        f" | {mv:.2f} | {sv:.2f} |" if not np.isnan(mv) else
                        f"| {lbl} | " + " | ".join(cells) + " | — | — |")
        return "\n".join(rows)

    # 12×12（不含 PP）
    _defs_12 = [
        ('HP',  'T1', 'HP（T1）',  6), ('JCP', 'T1', 'JCP（T1）', 6),
        ('DP',  'T1', 'DP（T1）',  5), ('CI',  'T1', 'CI（T1）',  8),
        ('HP',  'T2', 'HP（T2）',  6), ('JCP', 'T2', 'JCP（T2）', 6),
        ('DP',  'T2', 'DP（T2）',  5), ('CI',  'T2', 'CI（T2）',  8),
        ('HP',  'T3', 'HP（T3）',  6), ('JCP', 'T3', 'JCP（T3）', 6),
        ('DP',  'T3', 'DP（T3）',  5), ('CI',  'T3', 'CI（T3）',  8),
    ]
    _c12, _a12 = _md_compute_composites(_defs_12, _df_md)
    A("### 4A  12×12 追蹤相關矩陣（HP / JCP / DP / CI × T1~T3，不含 PP）")
    A("")
    A(_md_corr_table(_defs_12, _c12, _a12))
    A("")

    # 15×15（含 PP）
    _defs_15 = [
        ('HP',  'T1', 'HP（T1）',  6), ('JCP', 'T1', 'JCP（T1）', 6),
        ('DP',  'T1', 'DP（T1）',  5), ('CI',  'T1', 'CI（T1）',  8),
        ('PP',  'T1', 'PP（T1）',  6),
        ('HP',  'T2', 'HP（T2）',  6), ('JCP', 'T2', 'JCP（T2）', 6),
        ('DP',  'T2', 'DP（T2）',  5), ('CI',  'T2', 'CI（T2）',  8),
        ('PP',  'T2', 'PP（T2）',  6),
        ('HP',  'T3', 'HP（T3）',  6), ('JCP', 'T3', 'JCP（T3）', 6),
        ('DP',  'T3', 'DP（T3）',  5), ('CI',  'T3', 'CI（T3）',  8),
        ('PP',  'T3', 'PP（T3）',  6),
    ]
    _c15, _a15 = _md_compute_composites(_defs_15, _df_md)
    A("### 4B  15×15 追蹤相關矩陣（含 PP，N = {n}）".format(n=n))
    A("")
    A(_md_corr_table(_defs_15, _c15, _a15))
    A("")

    A("---")
    A("")

    # ── 四之一、CFA 測量模型適配指數 ─────────────────────────────────────────
    A("## 四之一、CFA 測量模型適配指數")
    A("")
    A("| 模型 | 結構 | CFI | RMSEA | SRMR | 判斷 |")
    A("|------|------|-----|-------|------|------|")
    def _cfa_row(key, label, desc):
        fit = all_results.get(key, {}).get('fit', {})
        cfi  = fit.get('cfi');   cfi_s  = f"{cfi:.3f}"  if isinstance(cfi,  float) else '—'
        rmse = fit.get('rmsea'); rmse_s = f"{rmse:.3f}" if isinstance(rmse, float) else '—'
        srmr = fit.get('srmr');  srmr_s = f"{srmr:.3f}" if isinstance(srmr, float) else '—'
        ok = (isinstance(cfi, float) and cfi >= .90 and
              isinstance(rmse, float) and rmse <= .08 and
              isinstance(srmr, float) and srmr <= .08)
        verdict = '✅' if ok else ('⚠️' if fit else '（未執行）')
        return f"| {label} | {desc} | {cfi_s} | {rmse_s} | {srmr_s} | {verdict} |"
    A(_cfa_row('CFA-E (HP+JCP+PP+DP+CI, 5F)', '模型5 CFA-E', 'HP/JCP/PP/DP/CI 五因子（T1）'))
    A(_cfa_row('CFA-H (Cross-Wave 5F)',         '模型8 CFA-H', 'HP/JCP/PP(T1) + DP(T2) + CI(T3) 跨波次'))
    A(_cfa_row('CFA-I (Cross-Wave 4F, no PP)', '模型9 CFA-I', 'HP/JCP(T1) + DP(T2) + CI(T3) 跨波次，不含 PP'))
    A("")
    A("> 判斷標準：CFI ≥ .90；RMSEA ≤ .08；SRMR ≤ .08。CFA-H 使用跨波次測量：HP/JCP/PP 取 T1，DP 取 T2，CI 取 T3。")
    A("")
    A("---")
    A("")

    # ── 五、假設檢驗 ──────────────────────────────────────────────────────────
    if is_nopp:
        # NoPP 版本：純中介模型 H1~H4
        _np = all_results.get('PATH_NoPP', {})
        np_paths = _np.get('paths', {})
        np_mc    = _np.get('modconstr', {})
        A("## 五、假設檢驗結果：純中介模型（PATH_NoPP，不含 PP）")
        A("")
        A("> 估計方法：ML + Bootstrap 5000；信賴區間：BC Bootstrap 95% CI")
        A("> JCP_T1、HP_T1、DP_T2 均以 grand-mean 中心化")
        A("")
        A("### 5-1 a-path：HP / JCP → DP")
        A("")
        A("| 假設 | 路徑 | β (STDYX) | SE | 95% BC CI | p | 支持？ |")
        A("|------|------|-----------|-----|-----------|---|--------|")
        A(f"| H1a | JCP(T1)→DP(T2) | {_v(np_paths,'H1a: JCP(T1)→DP(T2)')} | {_v(np_paths,'H1a: JCP(T1)→DP(T2)','se')} | {_ci(np_paths,'H1a: JCP(T1)→DP(T2)')} | {_pval(np_paths,'H1a: JCP(T1)→DP(T2)')} | {_sup(np_paths,'H1a: JCP(T1)→DP(T2)','+')} |")
        A(f"| H1b | HP(T1)→DP(T2) | {_v(np_paths,'H1b: HP(T1)→DP(T2)')} | {_v(np_paths,'H1b: HP(T1)→DP(T2)','se')} | {_ci(np_paths,'H1b: HP(T1)→DP(T2)')} | {_pval(np_paths,'H1b: HP(T1)→DP(T2)')} | {_sup(np_paths,'H1b: HP(T1)→DP(T2)','+')} |")
        A("")
        A("### 5-2 b-path：DP → CI")
        A("")
        A("| 假設 | 路徑 | β (STDYX) | SE | 95% BC CI | p | 支持？ |")
        A("|------|------|-----------|-----|-----------|---|--------|")
        A(f"| H2 | DP(T2)→CI(T3) | {_v(np_paths,'H2: DP(T2)→CI(T3)')} | {_v(np_paths,'H2: DP(T2)→CI(T3)','se')} | {_ci(np_paths,'H2: DP(T2)→CI(T3)')} | {_pval(np_paths,'H2: DP(T2)→CI(T3)')} | {_sup(np_paths,'H2: DP(T2)→CI(T3)','+')} |")
        A("")
        A("### 5-3 c'-path：HP / JCP 直接 → CI")
        A("")
        A("| 假設 | 路徑 | β (STDYX) | SE | 95% BC CI | p | 支持？ |")
        A("|------|------|-----------|-----|-----------|---|--------|")
        A(f"| H3a | JCP(T1)→CI(T3) 直接 | {_v(np_paths,'H3a: JCP(T1)→CI(T3) 直接')} | {_v(np_paths,'H3a: JCP(T1)→CI(T3) 直接','se')} | {_ci(np_paths,'H3a: JCP(T1)→CI(T3) 直接')} | {_pval(np_paths,'H3a: JCP(T1)→CI(T3) 直接')} | {_sup(np_paths,'H3a: JCP(T1)→CI(T3) 直接','+')} |")
        A(f"| H3b | HP(T1)→CI(T3) 直接 | {_v(np_paths,'H3b: HP(T1)→CI(T3) 直接')} | {_v(np_paths,'H3b: HP(T1)→CI(T3) 直接','se')} | {_ci(np_paths,'H3b: HP(T1)→CI(T3) 直接')} | {_pval(np_paths,'H3b: HP(T1)→CI(T3) 直接')} | {_sup(np_paths,'H3b: HP(T1)→CI(T3) 直接','+')} |")
        A("")
        A("### 5-4 間接效果 H4（Bootstrap 5000）")
        A("")
        A("| 假設 | 路徑 | β (indirect) | 95% BC CI | 顯著？ |")
        A("|------|------|--------------|-----------|--------|")
        A(f"| H4a | JCP→DP→CI | {_mc_v(np_mc,'IND_JCP')} | {_mc_ci(np_mc,'IND_JCP')} | {_mc_sup(np_mc,'IND_JCP')} |")
        A(f"| H4b | HP→DP→CI | {_mc_v(np_mc,'IND_HP')} | {_mc_ci(np_mc,'IND_HP')} | {_mc_sup(np_mc,'IND_HP')} |")
        A("")
        A("---")
        A("")
    else:
        A("## 五、假設檢驗結果：完整調節中介模型（PATH_ModMed）")
        A("")
        A("> 估計方法：ML + Bootstrap 5000；信賴區間：BC Bootstrap 95% CI")
        A("> 所有 T1 變項及 DP_T2 均以 grand-mean 中心化")
        A("")
        # a-path
        A("### 5-1 a-path：HP / JCP → DP（PP 調節）")
        A("")
        A("| 假設 | 路徑 | β (STDYX) | SE | 95% BC CI | p | 支持？ |")
        A("|------|------|-----------|-----|-----------|---|--------|")
        _H1a_k = 'H1a: JCP(T1)→DP(T2) [at mean PP]'
        _H1b_k = 'H1b: HP(T1)→DP(T2) [at mean PP]'
        _H2a_k = 'H2a: PP×JCP→DP (moderation)'
        _H2b_k = 'H2b: PP×HP→DP (moderation)'
        A(f"| H1a | JCP(T1)→DP(T2)（at mean PP） | {_v(pm_paths,_H1a_k)} | {_v(pm_paths,_H1a_k,'se')} | {_ci(pm_paths,_H1a_k)} | {_pval(pm_paths,_H1a_k)} | {_sup(pm_paths,_H1a_k,'+')} |")
        A(f"| H1b | HP(T1)→DP(T2)（at mean PP） | {_v(pm_paths,_H1b_k)} | {_v(pm_paths,_H1b_k,'se')} | {_ci(pm_paths,_H1b_k)} | {_pval(pm_paths,_H1b_k)} | {_sup(pm_paths,_H1b_k,'+')} |")
        A(f"| H2a | JCP×PP→DP | {_v(pm_paths,_H2a_k)} | {_v(pm_paths,_H2a_k,'se')} | {_ci(pm_paths,_H2a_k)} | {_pval(pm_paths,_H2a_k)} | {_sup(pm_paths,_H2a_k,'-')} |")
        A(f"| H2b | HP×PP→DP | {_v(pm_paths,_H2b_k)} | {_v(pm_paths,_H2b_k,'se')} | {_ci(pm_paths,_H2b_k)} | {_pval(pm_paths,_H2b_k)} | {_sup(pm_paths,_H2b_k,'-')} |")
        A("")
        # b-path
        A("### 5-2 b-path：DP → CI（PP 調節）")
        A("")
        A("| 假設 | 路徑 | β (STDYX) | SE | 95% BC CI | p | 支持？ |")
        A("|------|------|-----------|-----|-----------|---|--------|")
        _H3_k = 'H3:  DP(T2)→CI(T3) [at mean PP]'
        _H4_k = 'H4:  PP×DP→CI (moderation)'
        A(f"| H3 | DP(T2)→CI(T3)（at mean PP） | {_v(pm_paths,_H3_k)} | {_v(pm_paths,_H3_k,'se')} | {_ci(pm_paths,_H3_k)} | {_pval(pm_paths,_H3_k)} | {_sup(pm_paths,_H3_k,'+')} |")
        A(f"| H4 | DP×PP→CI | {_v(pm_paths,_H4_k)} | {_v(pm_paths,_H4_k,'se')} | {_ci(pm_paths,_H4_k)} | {_pval(pm_paths,_H4_k)} | {_sup(pm_paths,_H4_k,'-')} |")
        A("")
        # c'-path
        A("### 5-3 c'-path：HP / JCP 直接 → CI（PP 調節）")
        A("")
        A("| 假設 | 路徑 | β (STDYX) | SE | 95% BC CI | p | 支持？ |")
        A("|------|------|-----------|-----|-----------|---|--------|")
        _H5a_k = 'H5a: JCP(T1)→CI(T3) [at mean PP]'
        _H5b_k = 'H5b: HP(T1)→CI(T3) [at mean PP]'
        _H6a_k = 'H6a: PP×JCP→CI (moderation)'
        _H6b_k = 'H6b: PP×HP→CI (moderation)'
        A(f"| H5a | JCP(T1)→CI(T3) 直接 | {_v(pm_paths,_H5a_k)} | {_v(pm_paths,_H5a_k,'se')} | {_ci(pm_paths,_H5a_k)} | {_pval(pm_paths,_H5a_k)} | {_sup(pm_paths,_H5a_k,'+')} |")
        A(f"| H5b | HP(T1)→CI(T3) 直接 | {_v(pm_paths,_H5b_k)} | {_v(pm_paths,_H5b_k,'se')} | {_ci(pm_paths,_H5b_k)} | {_pval(pm_paths,_H5b_k)} | {_sup(pm_paths,_H5b_k,'+')} |")
        A(f"| H6a | JCP×PP→CI | {_v(pm_paths,_H6a_k)} | {_v(pm_paths,_H6a_k,'se')} | {_ci(pm_paths,_H6a_k)} | {_pval(pm_paths,_H6a_k)} | {_sup(pm_paths,_H6a_k,'-')} |")
        A(f"| H6b | HP×PP→CI | {_v(pm_paths,_H6b_k)} | {_v(pm_paths,_H6b_k,'se')} | {_ci(pm_paths,_H6b_k)} | {_pval(pm_paths,_H6b_k)} | {_sup(pm_paths,_H6b_k,'-')} |")
        A("")
        # H7 條件間接
        A("### 5-4 條件間接效果 H7（Bootstrap 5000，PP ±1SD）")
        A("")
        A("| 假設 | 路徑 | PP 水準 | β (indirect) | 95% BC CI | 支持？ |")
        A("|------|------|---------|--------------|-----------|--------|")
        A(f"| H7a | JCP→DP→CI | 高 PP (+1SD) | {_mc_v(pm_mc,'IND_HI_J')} | {_mc_ci(pm_mc,'IND_HI_J')} | {_mc_sup(pm_mc,'IND_HI_J')} |")
        A(f"| H7a | JCP→DP→CI | 低 PP (−1SD) | {_mc_v(pm_mc,'IND_LO_J')} | {_mc_ci(pm_mc,'IND_LO_J')} | {_mc_sup(pm_mc,'IND_LO_J')} |")
        A(f"| H7b | HP→DP→CI | 高 PP (+1SD) | {_mc_v(pm_mc,'IND_HI_H')} | {_mc_ci(pm_mc,'IND_HI_H')} | {_mc_sup(pm_mc,'IND_HI_H')} |")
        A(f"| H7b | HP→DP→CI | 低 PP (−1SD) | {_mc_v(pm_mc,'IND_LO_H')} | {_mc_ci(pm_mc,'IND_LO_H')} | {_mc_sup(pm_mc,'IND_LO_H')} |")
        A("")
        A("---")
        A("")

        # ── 六、基礎中介模型 ──────────────────────────────────────────────────
        A("## 六、基礎中介模型（PATH_Baseline）—「不含調節」主路徑")
        A("")
        A("> 用途：確認主路徑方向與顯著性，排除調節效果干擾")
        A("")
        A("| 假設 | 路徑 | β (STDYX) | SE | 95% CI | p | 支持？ |")
        A("|------|------|-----------|-----|--------|---|--------|")
        for hyp, lbl, exp in [
            ("H1a", "JCP(T1)→DP(T2)",       '+'),
            ("H1b", "HP(T1)→DP(T2)",        '+'),
            ("H3",  "DP(T2)→CI(T3)",        '+'),
            ("H5a", "JCP(T1)→CI(T3) 直接",  '+'),
            ("H5b", "HP(T1)→CI(T3) 直接",   '+'),
        ]:
            A(f"| {hyp} | {lbl} | {_v(bl_paths,lbl)} | {_v(bl_paths,lbl,'se')} | {_ci(bl_paths,lbl)} | {_pval(bl_paths,lbl)} | {_sup(bl_paths,lbl,exp)} |")
        A(f"| — | 間接效果 JCP→DP→CI | {_mc_v(bl_mc,'IND_JCP')} | — | {_mc_ci(bl_mc,'IND_JCP')} | — | {_mc_sup(bl_mc,'IND_JCP')} |")
        A(f"| — | 間接效果 HP→DP→CI | {_mc_v(bl_mc,'IND_HP')} | — | {_mc_ci(bl_mc,'IND_HP')} | — | {_mc_sup(bl_mc,'IND_HP')} |")
        A("")
        A("---")
        A("")

    # ── NoPP 版本：早期返回，跳過補充分析 ────────────────────────────────────
    if is_nopp:
        A(f"*此摘要依據 {ts} pipeline 輸出（純中介版本，不含 PP 調節）*")
        content = "\n".join(lines)
        out_path = os.path.join(run_dir, f"Report_NoPP_{ts}.md")
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"  [OK] NoPP Report: {out_path}")
        return out_path

    # ── 七、JCP-only ─────────────────────────────────────────────────────────
    A("## 七、JCP-only（PATH_JCP_only，移除 HP）")
    A("")
    A("> 目的：確認完整模型中 JCP 不顯著是否由 HP 共線性抑制所致")
    A("")
    A("| 路徑 | β (STDYX) | SE | 95% CI | p | 說明 |")
    A("|------|-----------|-----|--------|---|------|")
    for lbl, note in [
        ("JCP(T1)→DP(T2)",      ""),
        ("DP(T2)→CI(T3)",       ""),
        ("JCP(T1)→CI(T3) 直接", "若顯著→HP 抑制 JCP"),
    ]:
        A(f"| {lbl} | {_v(jcp_paths,lbl)} | {_v(jcp_paths,lbl,'se')} | {_ci(jcp_paths,lbl)} | {_pval(jcp_paths,lbl)} | {note} |")
    A(f"| 間接效果 JCP→DP→CI | {_mc_v(jcp_mc,'IND_JCP')} | — | {_mc_ci(jcp_mc,'IND_JCP')} | — | |")
    A("")
    A("---")
    A("")

    # ── 八、JCP-only + PP ────────────────────────────────────────────────────
    A("## 八、JCP-only + PP 調節模型（PATH_JCP_PP）")
    A("")
    A("> 確認 PP 是否調節 JCP 路徑（移除 HP，保留 PP）")
    A("")
    A("| 路徑 | β (STDYX) | SE | 95% CI | p | 說明 |")
    A("|------|-----------|-----|--------|---|------|")
    for lbl, note in [
        ("JCP(T1)→DP(T2) [at mean PP]",  ""),
        ("JCP×PP→DP (a-path 調節)",       ""),
        ("DP(T2)→CI(T3) [at mean PP]",   ""),
        ("DP×PP→CI (b-path 調節)",        ""),
        ("JCP(T1)→CI(T3) 直接",           ""),
    ]:
        A(f"| {lbl} | {_v(jpp_paths,lbl)} | {_v(jpp_paths,lbl,'se')} | {_ci(jpp_paths,lbl)} | {_pval(jpp_paths,lbl)} | {note} |")
    A("")
    A("**條件間接效果（PP ±1SD）：**")
    A("")
    A("| PP 水準 | β (indirect) | 95% BC CI | 說明 |")
    A("|---------|--------------|-----------|------|")
    A(f"| 高 PP (+1SD) | {_mc_v(jpp_mc,'IND_HI')} | {_mc_ci(jpp_mc,'IND_HI')} | {_mc_sup(jpp_mc,'IND_HI')} |")
    A(f"| 低 PP (−1SD) | {_mc_v(jpp_mc,'IND_LO')} | {_mc_ci(jpp_mc,'IND_LO')} | {_mc_sup(jpp_mc,'IND_LO')} |")
    A("")
    A("---")
    A("")

    # ── 九、職涯階段補充分析（Bayesian）──────────────────────────────────────
    A("## 九、職涯階段補充分析（PATH_CareerStage，Bayesian）")
    A("")
    A("> 估計方法：Bayesian MCMC；信賴區間：95% HPD CI（可信區間）")
    A("> 效果編碼：探索期 EXP_C=+1；維護期 MAINT_C=+1；建立期兩者均 −0.5")
    A("")
    A("### 9-1 樣本職涯階段分布")
    A("")
    if g3_sample is not None and 'Age_Group' in g3_sample.columns:
        exp_n  = int((g3_sample['Age_Group'] == 'EXP').sum())
        est_n  = int((g3_sample['Age_Group'] == 'ESTAB').sum())
        mnt_n  = int((g3_sample['Age_Group'] == 'MAINT').sum())
    else:
        exp_n, est_n, mnt_n = 64, 235, 58
    total = exp_n + est_n + mnt_n
    A("| 階段 | 年齡 | N | 佔比 |")
    A("|------|------|---|------|")
    A(f"| 探索期（EXP）  | 21–30 歲 | {exp_n}  | {exp_n/total*100:.1f}% |")
    A(f"| 建立期（ESTAB）| 31–40 歲 | {est_n} | {est_n/total*100:.1f}% |")
    A(f"| 維護期（MAINT）| 41 歲以上 | {mnt_n}  | {mnt_n/total*100:.1f}% |")
    A("")
    A("### 9-2 主效果（averaged across all stages）")
    A("")
    A("| 路徑 | β (post.mean) | SD | 95% HPD | p（one-tail） | 說明 |")
    A("|------|--------------|-----|---------|---------------|------|")
    for lbl, hyp in [
        ("JCP(T1)→DP(T2) [ESTAB均值]",  "H1a"),
        ("HP(T1)→DP(T2) [ESTAB均值]",   "H1b"),
        ("DP(T2)→CI(T3)",               "H3"),
        ("JCP(T1)→CI(T3) 直接",          "H5a"),
        ("HP(T1)→CI(T3) 直接",           "H5b"),
    ]:
        A(f"| {lbl} | {_v(cs_paths,lbl)} | {_v(cs_paths,lbl,'se')} | {_ci(cs_paths,lbl)} | {_pval(cs_paths,lbl)} | {hyp} |")
    A("")
    A("### 9-3 職涯階段交互作用項（a-path 調節）")
    A("")
    A("| 交互作用項 | β | SD | 95% HPD | 說明 |")
    A("|-----------|----|----|---------|------|")
    for lbl in ['JCP×EXP_C→DP (探索期效果)', 'JCP×MAINT_C→DP (維護期效果)',
                'HP×EXP_C→DP (探索期效果)',  'HP×MAINT_C→DP (維護期效果)']:
        A(f"| {lbl} | {_v(cs_paths,lbl)} | {_v(cs_paths,lbl,'se')} | {_ci(cs_paths,lbl)} | {_sup(cs_paths,lbl,'-')} |")
    A("")
    A("### 9-4 條件間接效果（依職涯階段）")
    A("")
    A("| 條件間接效果 | 階段 | β (indirect) | 95% HPD | 顯著？ |")
    A("|-------------|------|--------------|---------|--------|")
    for mc_key, label, stage in [
        ('ie_e_j', 'JCP→DP→CI', '探索期'), ('ie_m_j', 'JCP→DP→CI', '維護期'), ('ie_r_j', 'JCP→DP→CI', '建立期'),
        ('ie_e_h', 'HP→DP→CI',  '探索期'), ('ie_m_h', 'HP→DP→CI',  '維護期'), ('ie_r_h', 'HP→DP→CI',  '建立期'),
    ]:
        A(f"| {label} | {stage} | {_mc_v(cs_mc,mc_key)} | {_mc_ci(cs_mc,mc_key)} | {_mc_sup(cs_mc,mc_key)} |")
    A("")
    A("---")
    A("")

    # ── 十、績效考核補充資料 ─────────────────────────────────────────────────
    A("## 十、績效考核補充資料（待老師確認方向）")
    A("")
    A("> 資料來源：Analysis_Ready_Data 中 PM_Has / PM_Result / PM_Help 欄位")
    A("")
    A("| 波次 | 有考核比例 | 負向 n | 中立 n | 正向 n | 幫助程度 M (SD) |")
    A("|------|-----------|--------|--------|--------|----------------|")
    for wave, label in [(1,'T1'), (2,'T2'), (3,'T3')]:
        stats_ = _pm_stats(wave)
        A(f"| {label} | {stats_[0]} | {stats_[1]} | {stats_[2]} | {stats_[3]} | {stats_[4]} |")
    A("")
    A("---")
    A("")

    # ── 十一、分析語法索引 ───────────────────────────────────────────────────
    A("## 十一、分析語法索引")
    A("")
    A("| 工具 | 分析目的 | 檔案 |")
    A("|------|---------|------|")
    A(f"| SPSS | 匯入資料 + 變數標籤 | `SPSS_Syntax_{ts}.sps` |")
    A(f"| SPSS | 完整分析（描述統計／相關／CMV／CITC／信度）| `SPSS_Analysis_{ts}.sps` |")
    A(f"| Mplus | CFA-E 五因子 (HP/JCP/PP/DP/CI, T1) | `CFA_E_FiveFactor_{ts}.inp` |")
    A(f"| Mplus | CFA-H 跨波次五因子 (HP/JCP/PP@T1, DP@T2, CI@T3) | `CFA_H_CrossWave_{ts}.inp` |")
    A(f"| Mplus | CFA-I 跨波次四因子，不含 PP (HP/JCP@T1, DP@T2, CI@T3) | `CFA_I_CrossWave_4F_{ts}.inp` |")
    A(f"| Mplus | CFA-F 四因子 (CP合併/PP/DP/CI) | `CFA_F_FourFactor_CP_merged_{ts}.inp` |")
    A(f"| Mplus | CFA-G 三因子 (CP/DP/CI) | `CFA_G_ThreeFactor_CP_DP_CI_{ts}.inp` |")
    A(f"| Mplus | 測量不變性（MI） | `MI_A/B_*_{ts}.inp` |")
    A(f"| Mplus | 完整調節中介路徑模型 | `PATH_ModMed_{ts}.inp` |")
    A(f"| Mplus | 基礎中介（無調節）| `PATH_Baseline_{ts}.inp` |")
    A(f"| Mplus | JCP-only 模型 | `PATH_JCP_only_{ts}.inp` |")
    A(f"| Mplus | JCP+PP 調節模型 | `PATH_JCP_PP_{ts}.inp` |")
    A(f"| Mplus | 職涯階段 Bayesian 模型 | `PATH_CareerStage_{ts}.inp` |")
    A("")
    A("---")
    A("")

    # ── 十二、整體結論摘要 ───────────────────────────────────────────────────
    A("## 十二、整體結論摘要")
    A("")
    A("### 支持的假設")
    A("")
    A("| 假設 | 路徑 | 模型 |")
    A("|------|------|------|")
    # 動態判斷哪些假設支持
    _supported = {
        'H1b': (_sup(pm_paths,_H1b_k,'+') == '✓ 支持', "HP(T1)→DP(T2)", "完整模型"),
        'H3':  (_sup(pm_paths,_H3_k, '+') == '✓ 支持', "DP(T2)→CI(T3)", "完整模型"),
        'H5b': (_sup(pm_paths,_H5b_k,'+') == '✓ 支持', "HP(T1)→CI(T3) 直接", "完整模型"),
    }
    has_supported = False
    for hyp, (ok, path, model) in _supported.items():
        if ok:
            A(f"| {hyp} | {path} | {model} |")
            has_supported = True
    # H7b 間接
    h7b_hi_ok = _mc_sup(pm_mc,'IND_HI_H') == '✓ 顯著'
    h7b_lo_ok = _mc_sup(pm_mc,'IND_LO_H') == '✓ 顯著'
    if h7b_hi_ok:
        A("| H7b（高PP）| HP→DP→CI 條件間接 | 完整模型，高 PP 時顯著 |")
        has_supported = True
    if h7b_lo_ok:
        A("| H7b（低PP）| HP→DP→CI 條件間接 | 完整模型，低 PP 時顯著 |")
        has_supported = True
    if not has_supported:
        A("| （待 Mplus 執行後填入）| — | — |")
    A("")
    A("### 不支持的假設")
    A("")
    A("| 假設 | 路徑 | 備註 |")
    A("|------|------|------|")
    _not_supported = [
        ('H1a', _sup(pm_paths,_H1a_k,'+'), "JCP(T1)→DP(T2)"),
        ('H2a', _sup(pm_paths,_H2a_k,'-'), "JCP×PP→DP"),
        ('H2b', _sup(pm_paths,_H2b_k,'-'), "HP×PP→DP"),
        ('H4',  _sup(pm_paths,_H4_k, '-'), "DP×PP→CI"),
        ('H5a', _sup(pm_paths,_H5a_k,'+'), "JCP(T1)→CI(T3) 直接"),
        ('H6a', _sup(pm_paths,_H6a_k,'-'), "JCP×PP→CI"),
        ('H6b', _sup(pm_paths,_H6b_k,'-'), "HP×PP→CI"),
    ]
    has_not = False
    for hyp, result, path in _not_supported:
        if '不支持' in result or '待執行' in result:
            A(f"| {hyp} | {path} | {result} |")
            has_not = True
    if not has_not:
        A("| （待 Mplus 執行後填入）| — | — |")
    A("")
    A("### JCP 的特殊發現（與 HP 共線性）")
    A("")
    _jcp_ci_full = _ci(pm_paths, _H1a_k)
    A(f"- 完整模型（含 HP）：JCP→DP 路徑 β={_v(pm_paths,_H1a_k)}，CI={_jcp_ci_full}（不顯著）")
    _jcp_direct_jcponly = _v(jcp_paths, 'JCP(T1)→CI(T3) 直接')
    A(f"- JCP-only 模型（移除 HP）：JCP→CI 直接路徑 β={_jcp_direct_jcponly}")
    A("- **可能解釋**：HP 與 JCP 相關，HP 的效果在完整模型中壓制了 JCP（競爭預測）")
    A("")
    A("---")
    A("")
    A(f"*此摘要依據 {ts} pipeline 輸出，β 值均為 STDYX 標準化係數（職涯階段模型除外）*")

    # ════════════════════════════════════════════════════════════════════════
    content = "\n".join(lines)
    fname = f"Report_NoPP_{ts}.md" if is_nopp else f"Pipeline_Master_Report_{ts}.md"
    out_path = os.path.join(run_dir, fname)
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"  [OK] Pipeline Master Report: {out_path}")
    return out_path


# ==========================================
# TODO-2: MODULE G — Word 論文草稿（python-docx）
# ==========================================
# 尚未實作。建議在此處新增 generate_word_report(run_dir, ts, g3_sample, all_mplus_results, ...) 函式。
#
# 功能規劃：
#   - 使用 python-docx（pip install python-docx）
#   - APA 格式表格，內容對應 Excel Sheet 2~7
#   - 表格標題、標註、顯著性說明（*** p<.001 等）
#   - 章節：研究方法（樣本背景）→ 信效度 → 相關矩陣 → CFA 適配 → RI-CLPM 結果
#   - 在 main() 中呼叫：word_path = generate_word_report(run_dir, ts, g3_sample_full, alpha_dict, corr_dict, all_mplus_results)
#
# ==========================================
# TODO-3: MODULE H — PPT 簡報（python-pptx）
# ==========================================
# 尚未實作。建議在此處新增 generate_pptx_report(run_dir, ts, all_mplus_results, ...) 函式。
#
# 功能規劃：
#   - 使用 python-pptx（pip install python-pptx）
#   - 投影片規劃：研究架構圖 → 樣本說明 → CFA 適配表 → RI-CLPM 路徑圖（含係數）→ RI 相關 → 結論
#   - 在 main() 中呼叫：pptx_path = generate_pptx_report(run_dir, ts, all_mplus_results)
#
# ==========================================
# TODO-4: 刪題目設定（等老師確認後執行）
# ==========================================
# 問題題目：
#   - JCP6：標準化因素負荷量為負值（約 -.092 ~ -.097），且未達顯著
#   - DP1：因素負荷量極低（約 .276），與其他題目差距大
#
# 確認要刪後的修改方式：
#   1. 在程式最上方新增設定區塊（建議放在 OUTPUT_DIR 附近）：
#        ITEMS_TO_EXCLUDE = ['JCP6', 'DP1']   # 老師確認後填入
#   2. 在 generate_cfa_dat()、run_and_parse_all_models() 的 CFA model_lines 中
#      將對應題目從 USEVARIABLES 和 BY 語法中移除
#   3. 重新執行 pipeline 並比較新舊 CFA 適配指數
#
# ==========================================
# TODO-5: 執行 pipeline 驗證（第一次跑完整流程）
# ==========================================
# 建議執行步驟：
#   1. python pipeline_master.py
#   2. 確認 Master_Pipeline_Output/<ts>/ 資料夾下有：
#        - Thesis_Results_<ts>.xlsx（9 個 Sheet 資料正確）
#        - CFA_A~D_<ts>.inp、CFA_E/F/G_<ts>.inp（純 UTF-8，英文注記）
#        - RI_CLPM_A~D_<ts>.inp
#        - 若已安裝 Mplus：.out 檔會自動產生，Excel 內數字會填入（含 MI）
#        - 若未安裝 Mplus：Excel 欄位顯示「（尚未執行）」，手動執行 .inp 後可重跑 pipeline
#   3. 確認 SPSS .sps 語法在 SPSS 中可正常執行（所有 * 行結尾有句點）
#
# ==========================================

# ==========================================
# MODULE H: 兩版本比較報告
# ==========================================
def generate_comparison_excel(run_dir, ts, all_results):
    """
    產生 Comparison_{ts}.xlsx：WithPP vs NoPP 路徑係數並排比較 + 模型適配。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("  [警告] 未安裝 openpyxl，跳過比較 Excel。")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "兩版本比較"

    thin = Side(border_style="thin", color="000000")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    nopp_fill = PatternFill("solid", fgColor="FCE4D6")
    wpp_fill  = PatternFill("solid", fgColor="E2EFDA")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _hdr(row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hdr_fill; c.alignment = ctr; c.border = bdr

    def _cell(row, col, val, bold=False, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, size=10)
        c.alignment = ctr; c.border = bdr
        if fill: c.fill = fill
        return c

    def _fmt(d, key='est'):
        v = d.get(key, float('nan')) if d else float('nan')
        try:
            return f"{v:.3f}" if v == v else '—'
        except Exception:
            return '—'

    def _ci_str(d):
        if not d: return '—'
        lo, hi = d.get('ci_lo', float('nan')), d.get('ci_hi', float('nan'))
        try:
            return f"[{lo:.3f}, {hi:.3f}]" if (lo == lo and hi == hi) else '—'
        except Exception:
            return '—'

    r = 1
    ws.cell(row=r, column=1,
            value=f"WithPP vs NoPP 兩版本路徑係數比較（{ts}）"
            ).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    # ── 適配指數比較 ──────────────────────────────────────────────
    sec = ws.cell(row=r, column=1, value="模型適配指數比較")
    sec.font = Font(bold=True, size=11, color="FFFFFF")
    sec.fill = PatternFill("solid", fgColor="2E4057")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for ci, h in enumerate(["版本", "模型", "CFI", "RMSEA", "SRMR", "AIC", "BIC", "備注"], 1):
        _hdr(r, ci, h)
    r += 1

    def _fit_row(label, version, res, fill):
        fit = res.get('fit', {}) if res else {}
        def _f(k):
            v = fit.get(k, float('nan'))
            try: return f"{v:.3f}" if v == v else '—'
            except: return '—'
        for ci, v in enumerate([version, label,
                                 _f('cfi'), _f('rmsea'), _f('srmr'),
                                 f"{fit.get('aic',float('nan')):.1f}" if fit.get('aic') else '—',
                                 f"{fit.get('bic',float('nan')):.1f}" if fit.get('bic') else '—',
                                 ''], 1):
            _cell(r, ci, v, fill=fill)

    _fit_row('PATH: JCP/HP/PP(T1)→DP(T2)→CI(T3)', 'WithPP',
             all_results.get('PATH (T1→T2→T3)', {}), wpp_fill)
    r += 1
    _fit_row('PATH_NoPP: JCP/HP(T1)→DP(T2)→CI(T3)', 'NoPP',
             all_results.get('PATH_NoPP', {}), nopp_fill)
    r += 2

    # ── 路徑係數並排比較 ──────────────────────────────────────────
    sec2 = ws.cell(row=r, column=1, value="主路徑係數並排比較（STDYX 標準化 β）")
    sec2.font = Font(bold=True, size=11, color="FFFFFF")
    sec2.fill = PatternFill("solid", fgColor="4472C4")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for ci, h in enumerate(["路徑", "對應假設",
                             "WithPP β", "WithPP CI",
                             "NoPP β",  "NoPP CI",
                             "方向一致？", "備注"], 1):
        _hdr(r, ci, h)
    r += 1

    pm   = all_results.get('PATH (T1→T2→T3)', {})
    npm  = all_results.get('PATH_NoPP', {})
    pm_p = pm.get('paths', {})
    np_p = npm.get('paths', {})

    comparisons = [
        ("JCP(T1)→DP(T2)", "H1a",
         pm_p.get('H1a: JCP(T1)→DP(T2) [at mean PP]', {}),
         np_p.get('H1a: JCP(T1)→DP(T2)', {})),
        ("HP(T1)→DP(T2)", "H1b",
         pm_p.get('H1b: HP(T1)→DP(T2) [at mean PP]', {}),
         np_p.get('H1b: HP(T1)→DP(T2)', {})),
        ("DP(T2)→CI(T3)", "H3/H2",
         pm_p.get('H3:  DP(T2)→CI(T3) [at mean PP]', {}),
         np_p.get('H2: DP(T2)→CI(T3)', {})),
        ("JCP(T1)→CI(T3) 直接", "H5a/H3a",
         pm_p.get('H5a: JCP(T1)→CI(T3) [at mean PP]', {}),
         np_p.get('H3a: JCP(T1)→CI(T3) 直接', {})),
        ("HP(T1)→CI(T3) 直接", "H5b/H3b",
         pm_p.get('H5b: HP(T1)→CI(T3) [at mean PP]', {}),
         np_p.get('H3b: HP(T1)→CI(T3) 直接', {})),
    ]

    pm_mc = pm.get('modconstr', {})
    np_mc = npm.get('modconstr', {})
    indirect_comparisons = [
        ("JCP→DP→CI 間接", "H7a(mean)/H4a",
         pm_mc.get('IND_HI_J', {}), np_mc.get('IND_JCP', {})),
        ("HP→DP→CI 間接", "H7b(mean)/H4b",
         pm_mc.get('IND_HI_H', {}), np_mc.get('IND_HP', {})),
    ]

    def _sign(d):
        e = d.get('est', float('nan')) if d else float('nan')
        try: return '+' if e > 0 else ('-' if e < 0 else '0')
        except: return '?'

    for path, hyp, wpp_d, npp_d in comparisons:
        w_sign = _sign(wpp_d); n_sign = _sign(npp_d)
        consistent = '✓' if w_sign == n_sign and w_sign != '?' else '△' if not wpp_d or not npp_d else '✗'
        fill = wpp_fill if consistent == '✓' else (PatternFill("solid", fgColor="FFE0E0") if consistent == '✗' else None)
        for ci, v in enumerate([path, hyp, _fmt(wpp_d), _ci_str(wpp_d),
                                 _fmt(npp_d), _ci_str(npp_d), consistent, ''], 1):
            _cell(r, ci, v, fill=fill)
        r += 1

    r += 1
    note = ws.cell(row=r, column=1, value="間接效果（間接路徑 Bootstrap）")
    note.font = Font(bold=True, size=10); note.border = bdr
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for path, hyp, wpp_d, npp_d in indirect_comparisons:
        w_sign = _sign(wpp_d); n_sign = _sign(npp_d)
        consistent = '✓' if w_sign == n_sign and w_sign != '?' else '△'
        fill = wpp_fill if consistent == '✓' else None
        for ci, v in enumerate([path, hyp, _fmt(wpp_d), _ci_str(wpp_d),
                                 _fmt(npp_d), _ci_str(npp_d), consistent, ''], 1):
            _cell(r, ci, v, fill=fill)
        r += 1

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14

    out_path = os.path.join(run_dir, f"Comparison_{ts}.xlsx")
    try:
        wb.save(out_path)
        print(f"  [OK] Comparison Excel 已儲存：{out_path}")
    except Exception as e:
        print(f"  [錯誤] Comparison Excel 儲存失敗：{e}")
        return None
    return out_path


def generate_comparison_md(run_dir, ts, all_results):
    """
    產生 Comparison_{ts}.md：兩版本差異敘述。
    """
    lines = []
    A = lines.append

    def _v(d, k='est'):
        v = d.get(k, float('nan')) if d else float('nan')
        try: return f"{v:.3f}" if v == v else '—'
        except: return '—'

    def _ci(d):
        if not d: return '—'
        lo, hi = d.get('ci_lo', float('nan')), d.get('ci_hi', float('nan'))
        try: return f"[{lo:.3f}, {hi:.3f}]" if (lo == lo and hi == hi) else '—'
        except: return '—'

    pm   = all_results.get('PATH (T1→T2→T3)', {})
    npm  = all_results.get('PATH_NoPP', {})
    pm_p = pm.get('paths', {})
    np_p = npm.get('paths', {})
    pm_mc = pm.get('modconstr', {})
    np_mc = npm.get('modconstr', {})

    A(f"# WithPP vs NoPP 兩版本結果比較（{ts}）")
    A("")
    A("> WithPP = 完整調節中介模型（PP 調節 a/b/c' 三條路徑）")
    A("> NoPP = 純中介模型（移除 PP 所有相關分析）")
    A("")
    A("---")
    A("")
    A("## 一、模型適配比較")
    A("")
    A("| 版本 | 模型 | CFI | RMSEA | SRMR | AIC | BIC |")
    A("|------|------|-----|-------|------|-----|-----|")

    def _fit_row_md(label, version, res):
        fit = res.get('fit', {}) if res else {}
        def _f(k):
            v = fit.get(k, float('nan'))
            try: return f"{v:.3f}" if v == v else '—'
            except: return '—'
        aic = fit.get('aic', float('nan'))
        bic = fit.get('bic', float('nan'))
        try: aic_s = f"{aic:.1f}" if aic == aic else '—'
        except: aic_s = '—'
        try: bic_s = f"{bic:.1f}" if bic == bic else '—'
        except: bic_s = '—'
        A(f"| {version} | {label} | {_f('cfi')} | {_f('rmsea')} | {_f('srmr')} | {aic_s} | {bic_s} |")

    _fit_row_md('PATH_ModMed', 'WithPP', pm)
    _fit_row_md('PATH_NoPP',   'NoPP',   npm)
    A("")
    A("---")
    A("")
    A("## 二、主路徑係數並排比較")
    A("")
    A("| 路徑 | 對應假設 | WithPP β | WithPP CI | NoPP β | NoPP CI | 方向一致？ |")
    A("|------|----------|----------|-----------|--------|---------|-----------|")

    rows = [
        ("JCP(T1)→DP(T2)", "H1a",
         pm_p.get('H1a: JCP(T1)→DP(T2) [at mean PP]', {}),
         np_p.get('H1a: JCP(T1)→DP(T2)', {})),
        ("HP(T1)→DP(T2)", "H1b",
         pm_p.get('H1b: HP(T1)→DP(T2) [at mean PP]', {}),
         np_p.get('H1b: HP(T1)→DP(T2)', {})),
        ("DP(T2)→CI(T3)", "H3/H2",
         pm_p.get('H3:  DP(T2)→CI(T3) [at mean PP]', {}),
         np_p.get('H2: DP(T2)→CI(T3)', {})),
        ("JCP(T1)→CI(T3) 直接", "H5a/H3a",
         pm_p.get('H5a: JCP(T1)→CI(T3) [at mean PP]', {}),
         np_p.get('H3a: JCP(T1)→CI(T3) 直接', {})),
        ("HP(T1)→CI(T3) 直接", "H5b/H3b",
         pm_p.get('H5b: HP(T1)→CI(T3) [at mean PP]', {}),
         np_p.get('H3b: HP(T1)→CI(T3) 直接', {})),
    ]
    for path, hyp, wd, nd in rows:
        we, ne = wd.get('est', float('nan')) if wd else float('nan'), nd.get('est', float('nan')) if nd else float('nan')
        try: consistent = '✓' if (we > 0) == (ne > 0) else '✗'
        except: consistent = '?'
        A(f"| {path} | {hyp} | {_v(wd)} | {_ci(wd)} | {_v(nd)} | {_ci(nd)} | {consistent} |")

    A("")
    A("## 三、間接效果比較")
    A("")
    A("| 路徑 | WithPP（at mean PP） β | WithPP CI | NoPP β | NoPP CI |")
    A("|------|----------------------|-----------|--------|---------|")
    A(f"| JCP→DP→CI | {_v(pm_mc.get('IND_HI_J',{}))} | {_ci(pm_mc.get('IND_HI_J',{}))} | {_v(np_mc.get('IND_JCP',{}))} | {_ci(np_mc.get('IND_JCP',{}))} |")
    A(f"| HP→DP→CI | {_v(pm_mc.get('IND_HI_H',{}))} | {_ci(pm_mc.get('IND_HI_H',{}))} | {_v(np_mc.get('IND_HP',{}))} | {_ci(np_mc.get('IND_HP',{}))} |")
    A("")
    A("---")
    A("")
    A("## 四、WithPP 版本獨有結果（PP 調節效果）")
    A("")
    A("| 假設 | 路徑 | β | CI | 說明 |")
    A("|------|------|---|-----|------|")
    for hyp, key in [
        ("H2a", "H2a: PP×JCP→DP (moderation)"),
        ("H2b", "H2b: PP×HP→DP (moderation)"),
        ("H4",  "H4:  PP×DP→CI (moderation)"),
        ("H6a", "H6a: PP×JCP→CI (moderation)"),
        ("H6b", "H6b: PP×HP→CI (moderation)"),
    ]:
        d = pm_p.get(key, {})
        A(f"| {hyp} | {key} | {_v(d)} | {_ci(d)} | 僅 WithPP 版本 |")
    A("")
    A(f"*此比較摘要依據 {ts} pipeline 輸出*")

    content = "\n".join(lines)
    out_path = os.path.join(run_dir, f"Comparison_{ts}.md")
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"  [OK] Comparison MD: {out_path}")
    return out_path


# ==========================================
# 3. MAIN PIPELINE
# ==========================================
def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    run_dir = os.path.join(OUTPUT_DIR, ts)   # e.g. Master_Pipeline_Output/20260325_1730
    os.makedirs(run_dir, exist_ok=True)
    
    print("--- Starting Master Pipeline ---")
    merged_df, escales, tracking = perform_matching()
    
    print("Running Analyses...")
    attrition_md, merged_df, anova_stats, chi_stats = analyze_attrition(merged_df, tracking)

    g3_sample = merged_df[merged_df['Group'] == 3].copy()
    desc_md, alpha_dict, corr_dict = run_descriptives_and_correlations(
        g3_sample if not g3_sample.empty else merged_df, escales)
    
    # === 產出分析用 CSV（移除身份識別欄位與暫時計算欄位）===
    drop_cols = ([c for c in merged_df.columns if c.startswith('_') or c == 'Edu'] +
                 ['Custom_UID', 'Timestamp', 'Matched_T1_ID', 'Matched_T1_ID_x',
                  'Matched_T1_ID_y', 'System_ID', 'key1', 'key2', 'key3',
                  'dedup_id', 'Email'])
    analysis_df = merged_df.drop(columns=[c for c in drop_cols if c in merged_df.columns], errors='ignore')
    analysis_filename = f"Analysis_Ready_Data_{ts}.csv"
    analysis_path = os.path.join(run_dir, analysis_filename)
    analysis_df.to_csv(analysis_path, index=False, encoding='utf-8-sig')

    # === 產出 SPSS 匯入 + 變數標籤語法 ===
    spss_syntax = generate_spss_syntax(analysis_path, ts)
    spss_sps_path = os.path.join(run_dir, f"SPSS_Syntax_{ts}.sps")
    with open(spss_sps_path, 'w', encoding='utf-8-sig') as f:
        f.write(spss_syntax)

    # SPSS_Reliability_*.sps 已停止輸出：詳細信度（CORR/MEANS/VARIANCE）已整合至 SPSS_Analysis_*.sps 步驟8

    # === 產出 SPSS 完整分析語法（新版，含 CMV/相關/描述統計/CITC）===
    _pp_cols = [f'PP{i+1}_T1' for i in range(6)]
    _pp_series = g3_sample[[c for c in _pp_cols if c in g3_sample.columns]].mean(axis=1)
    _pp_median = round(float(_pp_series.median()), 3)
    spss_analysis_syntax = generate_spss_analysis_syntax(
        analysis_path, ts, pp_median=_pp_median, n_total=len(g3_sample))
    spss_analysis_path = os.path.join(run_dir, f"SPSS_Analysis_{ts}.sps")
    with open(spss_analysis_path, 'w', encoding='utf-8-sig') as f:
        f.write(spss_analysis_syntax)

    # === 產出 Mplus .dat 資料檔（含控制變數）===
    g3_sample_full = merged_df[merged_df['Group'] == 3].copy() if 'Group' in merged_df.columns else merged_df.dropna(subset=['HP_T3','DP_T3','CI_T3']).copy()
    mplus_dat_path, mplus_dat_filename = generate_mplus_dat(g3_sample_full, run_dir, ts)

    # === 產出 CFA 用 dat（T1 原始題目）===
    cfa_dat_path, cfa_dat_filename = generate_cfa_dat(g3_sample_full, run_dir, ts)
    # === 產出 CFA-H 用 dat（跨波次：HP/JCP/PP@T1，DP@T2，CI@T3）===
    cfa_h_dat_path, cfa_h_dat_filename = generate_cfa_h_dat(g3_sample_full, run_dir, ts)

    # === Phase 1: 生成所有 .inp 並執行 CFA（論文順序：CFA → MI）===
    print("[Mplus] 生成所有 Mplus .inp 並執行 CFA 模型...")
    all_mplus_results, all_inp_list = run_and_parse_all_models(
        run_dir, mplus_dat_filename, cfa_dat_filename, ts, phases=['cfa'],
        cfa_h_dat_filename=cfa_h_dat_filename)
    cfa_paths = []  # 保留空串列供相容舊參照

    # === 產出 Excel 綜合報告（NoPP 純中介版本，主要輸出）===
    print("[Excel] 產生縱貫中介 Excel 綜合報告...")
    excel_nopp_path = generate_excel_report(
        run_dir, ts, g3_sample_full, alpha_dict, corr_dict, all_mplus_results,
        variant_label='nopp')

    # === 產出完整分析報告 .md（NoPP）===
    print("[MD] 產生縱貫中介完整分析報告...")
    report_nopp_path = generate_analysis_summary_md(
        run_dir, ts, g3_sample_full, alpha_dict, corr_dict,
        all_mplus_results, attrition_md, desc_md, variant='nopp')

    # [PP 調節中介已停用] WithPP 版本 Excel 與比較報告不再產出
    excel_path = None
    report_path = None
    comparison_excel_path = None
    comparison_md_path = None

    print(f"[OK] Pipeline Completed!")
    print(f"   - Report           : {report_nopp_path}")
    if excel_nopp_path:
        print(f"   - Excel            : {excel_nopp_path}")
    for label, inp_p in all_inp_list:
        print(f"   - Mplus (新) {label}: {os.path.basename(inp_p)}")
    print(f"   - Analysis Data    : {analysis_path}")
    print(f"   - SPSS 匯入語法    : {spss_sps_path}")
    print(f"   - SPSS 完整分析    : {spss_analysis_path}")
    # SPSS_Reliability_*.sps 已停止輸出，詳細信度已整合至 SPSS_Analysis_*.sps
    print(f"   - Mplus CFA dat    : {cfa_dat_path}")
    print(f"   - Mplus dat        : {mplus_dat_path}")

    # 刪題敏感性分析已停用：實際驗證後，刪除 JCP6、DP1 或兩者後
    # CFA-E (五因子) CFI 最高僅達 0.898，未達 .90 門檻，
    # 故刪題對模型適配無實質改善，不納入正式分析流程。
    print("\n--- Pipeline 完成 ---")


if __name__ == "__main__":
    main()
