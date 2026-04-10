"""
generate_spss_tables.py
自動讀取最新 Pipeline 輸出，產生第三章 / 第四章所需 SPSS 學術格式 Excel 表格。
每次跑完 pipeline_master.py 後執行此腳本，數字即自動更新。
"""
import pandas as pd
import numpy as np
import os
import re
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from scipy import stats as scipy_stats

# ============================================================
# 1. 自動找最新 Pipeline 輸出資料夾
# ============================================================
BASE_DIR   = r"g:\其他電腦\我的 PC\NSYSU_HRM\Thesis_LM\Research_Questionaire"
OUTPUT_BASE = os.path.join(BASE_DIR, "Master_Pipeline_Output")
EXCEL_OUT   = r"g:\其他電腦\我的 PC\NSYSU_HRM\Thesis_LM\05_會議報告與關聯圖\報告簡報\Canva_Tables_SPSS_Style_v2.xlsx"

subfolders = sorted(
    [d for d in os.listdir(OUTPUT_BASE)
     if os.path.isdir(os.path.join(OUTPUT_BASE, d)) and d[:8].isdigit()],
    reverse=True
)
if not subfolders:
    raise FileNotFoundError("找不到任何 Pipeline 輸出資料夾")
ts         = subfolders[0]
latest_dir = os.path.join(OUTPUT_BASE, ts)
print(f"[INFO] 使用最新 Pipeline 輸出: {ts}")

csv_path    = os.path.join(latest_dir, f"Analysis_Ready_Data_{ts}.csv")
report_path = os.path.join(latest_dir, f"Pipeline_Master_Report_{ts}.md")

# ============================================================
# 2. 讀取資料（只取三波完整樣本 Group=3）
# ============================================================
df_all = pd.read_csv(csv_path)
df     = df_all[df_all['Group'] == 3].copy().reset_index(drop=True)
N      = len(df)
print(f"[INFO] 三波完整樣本 N = {N}")

# ============================================================
# 3. 輔助函式
# ============================================================
def scale_mean(df, prefix, n, wave):
    cols = [f"{prefix}{i}_{wave}" for i in range(1, n+1)]
    valid = [c for c in cols if c in df.columns]
    return df[valid].mean(axis=1)

def cronbach_alpha(df, cols):
    valid = [c for c in cols if c in df.columns]
    data  = df[valid].dropna()
    if len(data) < 2 or len(valid) < 2:
        return np.nan
    k         = len(valid)
    var_sum   = data.var(axis=0, ddof=1).sum()
    var_total = data.sum(axis=1).var(ddof=1)
    return round((k / (k - 1)) * (1 - var_sum / var_total), 3)

def sig_stars(r, n):
    if pd.isna(r) or n < 3:
        return ''
    t_val = r * np.sqrt((n - 2) / max(1 - r**2, 1e-10))
    p     = 2 * scipy_stats.t.sf(abs(t_val), df=n - 2)
    if p < 0.001: return '***'
    if p < 0.01:  return '**'
    if p < 0.05:  return '*'
    return ''

def wave_stats(df, wave):
    """計算一個波次的六個量表 M / SD / α 及 Series"""
    hp_cols  = [f"HP{i}_{wave}"  for i in range(1, 7)]
    jcp_cols = [f"JCP{i}_{wave}" for i in range(1, 7)]
    cp_cols  = hp_cols + jcp_cols
    pp_cols  = [f"PP{i}_{wave}"  for i in range(1, 7)]
    dp_cols  = [f"DP{i}_{wave}"  for i in range(1, 6)]
    ci_cols  = [f"CI{i}_{wave}"  for i in range(1, 9)]

    hp_s  = scale_mean(df, 'HP',  6, wave)
    jcp_s = scale_mean(df, 'JCP', 6, wave)
    cp_s  = pd.concat([hp_s, jcp_s], axis=1).mean(axis=1)

    out = {}
    for code, s, cols in [
        ('HP',  hp_s,                hp_cols),
        ('JCP', jcp_s,               jcp_cols),
        ('CP',  cp_s,                cp_cols),
        ('PP',  scale_mean(df,'PP',6,wave),  pp_cols),
        ('DP',  scale_mean(df,'DP',5,wave),  dp_cols),
        ('CI',  scale_mean(df,'CI',8,wave),  ci_cols),
    ]:
        out[code] = {
            'M':     round(s.mean(), 2),
            'SD':    round(s.std(),  2),
            'alpha': cronbach_alpha(df, cols),
            's':     s
        }
    return out

def corr_matrix_df(stats, n):
    """下三角相關矩陣，對角線填 α，上三角空白，含 M/SD"""
    order  = ['HP', 'JCP', 'CP', 'PP', 'DP', 'CI']
    labels = ['1. HP', '2. JCP', '3. CP', '4. PP', '5. DP', '6. CI']
    series = pd.DataFrame({c: stats[c]['s'] for c in order}).dropna()
    corr   = series.corr()
    rows   = []
    for i, (code, lbl) in enumerate(zip(order, labels)):
        m, sd = stats[code]['M'], stats[code]['SD']
        row   = [f"{lbl}  M={m:.2f}, SD={sd:.2f}"]
        for j, c2 in enumerate(order):
            if j < i:
                r     = corr.loc[code, c2]
                stars = sig_stars(r, n)
                row.append(f"{r:.2f}{stars}")
            elif j == i:
                row.append(f"({stats[code]['alpha']:.2f})")
            else:
                row.append("")
        rows.append(row)
    return pd.DataFrame(rows, columns=["變數（M, SD）"] + labels)

# ============================================================
# 4. Mplus .out 解析函式
# ============================================================
def parse_mplus_fit(out_path):
    """從 Mplus .out 擷取適配指數 (CFI/TLI/RMSEA/SRMR/χ²/df/p)
    格式範例（Mplus 7.4）：
      CFI/TLI\n  CFI   0.908\n  TLI   0.894
      RMSEA (Root...)\n  Estimate   0.068\n  90% CI  0.060  0.076
      SRMR (...)\n  Value   0.066
      Chi-Square Test of Model Fit\n  Value  382.810*\n  Degrees of Freedom  149
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
        text = f.read()
    res = {}

    # CFI / TLI  (after "CFI/TLI" header)
    m = re.search(r'CFI\s+([\d.]+)', text)
    if m:
        try: res['CFI'] = float(m.group(1))
        except Exception: pass
    m = re.search(r'TLI\s+([\d.]+)', text)
    if m:
        try: res['TLI'] = float(m.group(1))
        except Exception: pass

    # RMSEA Estimate (line after RMSEA header)
    m = re.search(r'RMSEA.*?Estimate\s+([\d.]+)', text, re.DOTALL | re.IGNORECASE)
    if m:
        try: res['RMSEA'] = float(m.group(1))
        except Exception: pass

    # RMSEA 90% CI (two numbers on same line after "90 Percent")
    m = re.search(r'90 Percent C\.I\.\s+([\d.]+)\s+([\d.]+)', text, re.IGNORECASE)
    if m:
        try:
            res['RMSEA_lo'] = float(m.group(1))
            res['RMSEA_hi'] = float(m.group(2))
        except Exception: pass

    # SRMR Value
    m = re.search(r'SRMR.*?Value\s+([\d.]+)', text, re.DOTALL | re.IGNORECASE)
    if m:
        try: res['SRMR'] = float(m.group(1))
        except Exception: pass

    # chi2 Value, df, p
    m = re.search(r'Chi-Square Test.*?Value\s+([\d.]+)', text, re.DOTALL | re.IGNORECASE)
    if m:
        try: res['chi2'] = float(m.group(1))
        except Exception: pass
    m = re.search(r'Degrees of Freedom\s+([\d]+)', text, re.IGNORECASE)
    if m:
        try: res['df'] = int(m.group(1))
        except Exception: pass
    m = re.search(r'P-Value\s+([\d.]+)', text, re.IGNORECASE)
    if m:
        try: res['p_chi2'] = float(m.group(1))
        except Exception: pass

    return res


def parse_mplus_stdyx(out_path, path_map):
    """
    從 Mplus .out 的 STDYX 區段擷取路徑係數。
    path_map: { '顯示標籤': ('outcome_var', 'predictor_var') }
    回傳 { '標籤': {'est':, 'se':, 'z':, 'p':} }
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
        text = f.read()
    m = re.search(r'STDYX Standardization\s+(.*?)(?=\nR-SQUARE|\nSTD |\Z)', text, re.DOTALL)
    if not m:
        return {}
    stdyx = m.group(1)
    results = {}
    for label, (outcome, predictor) in path_map.items():
        on_m = re.search(
            rf'{re.escape(outcome)}\s+ON\s+(.*?)(?=\n\s*\n\s*\w|\Z)',
            stdyx, re.DOTALL | re.IGNORECASE)
        if not on_m:
            continue
        row = re.search(
            rf'\b{re.escape(predictor)}\s+([-\d.]+)\s+([\d.]+)\s+([-\d.]+)\s+([\d.]+)',
            on_m.group(1), re.IGNORECASE)
        if row:
            results[label] = {
                'est': float(row.group(1)),
                'se':  float(row.group(2)),
                'z':   float(row.group(3)),
                'p':   float(row.group(4))
            }
    return results


def parse_mplus_unstd(out_path, path_map):
    """
    從 Mplus .out 擷取非標準化估計值 b/SE/p 及 95% CI。
    MODEL RESULTS 格式: predictor  Estimate  S.E.  Est./S.E.  P-Value
    CI RESULTS 格式:    predictor  Lo.5%  Lo2.5%  Lo5%  Estimate  Hi5%  Hi2.5%  Hi.5%
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
        text = f.read()

    # ── 取 MODEL RESULTS 區段（非標準化）──────────────────────
    mr_start = text.find('MODEL RESULTS')
    mr_end   = text.find('STANDARDIZED MODEL RESULTS', mr_start) if mr_start >= 0 else -1
    mr_text  = text[mr_start:mr_end] if (mr_start >= 0 and mr_end > mr_start) else (text[mr_start:] if mr_start >= 0 else '')

    # ── 取 CONFIDENCE INTERVALS OF MODEL RESULTS 區段 ──────────
    ci_start = text.find('CONFIDENCE INTERVALS OF MODEL RESULTS')
    ci_end   = text.find('CONFIDENCE INTERVALS OF STANDARDIZED', ci_start) if ci_start >= 0 else -1
    ci_text  = text[ci_start:ci_end] if (ci_start >= 0 and ci_end > ci_start) else (text[ci_start:] if ci_start >= 0 else '')

    def _get_on_section(full_text, outcome):
        """Return text within 'outcome ON ...' block."""
        m = re.search(
            rf'{re.escape(outcome)}\s+ON\s+(.*?)(?=\n\s*\n\s*[A-Z\s]|\Z)',
            full_text, re.DOTALL | re.IGNORECASE)
        return m.group(1) if m else ''

    results = {}
    for label, (outcome, predictor) in path_map.items():
        row_data = {}

        # Non-standardized b, SE, p from MODEL RESULTS
        on_mr = _get_on_section(mr_text, outcome)
        if on_mr:
            row = re.search(
                rf'^\s*{re.escape(predictor)}\s+([-\d.]+)\s+([\d.]+)\s+[-\d.]+\s+([\d.]+)',
                on_mr, re.IGNORECASE | re.MULTILINE)
            if row:
                row_data['b']  = float(row.group(1))
                row_data['se'] = float(row.group(2))
                row_data['p']  = float(row.group(3))

        # 95% CI: 7 values per row → positions [1]=lo2.5%, [5]=hi2.5% (0-indexed)
        on_ci = _get_on_section(ci_text, outcome)
        if on_ci:
            row2 = re.search(
                rf'^\s*{re.escape(predictor)}\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)\s+([-\d.]+)',
                on_ci, re.IGNORECASE | re.MULTILINE)
            if row2:
                row_data['ci_lo'] = float(row2.group(2))   # Lower 2.5%
                row_data['ci_hi'] = float(row2.group(6))   # Upper 2.5%
                if 'b' not in row_data:
                    row_data['b'] = float(row2.group(4))   # Estimate (fallback)

        if row_data:
            results[label] = row_data
    return results


def parse_multigroup_paths(out_path, path_map, group_name):
    """
    從多群組 Mplus .out 擷取特定群組的路徑係數。
    先嘗試 STDYX；若未收斂，fallback 到 MODEL RESULTS（unstandardized Estimate only）。
    回傳 { label: {'est': float, 'converged': bool} }
    """
    if not out_path or not os.path.isfile(out_path):
        return {}
    with open(out_path, 'r', encoding='utf-8', errors='replace') as f:
        text = f.read()

    converged = 'NO CONVERGENCE' not in text and 'TERMINATED NORMALLY' in text

    # Try STDYX first
    grp_start = text.upper().find(group_name.upper())
    stdyx = ''
    if grp_start >= 0:
        text_after = text[grp_start:]
        stdyx_m = re.search(r'STDYX Standardization\s+(.*?)(?=\nR-SQUARE|\nGroup\s|\Z)',
                            text_after, re.DOTALL)
        if stdyx_m:
            stdyx = stdyx_m.group(1)

    # Fallback: MODEL RESULTS for this group (unstandardized)
    mr_group = ''
    if not stdyx and grp_start >= 0:
        # Find next group header
        rest = text[grp_start:]
        next_grp = re.search(r'\nGroup\s+\w', rest[10:])
        end_pos = next_grp.start() + 10 if next_grp else len(rest)
        mr_group = rest[:end_pos]

    results = {}
    search_text = stdyx if stdyx else mr_group
    if not search_text:
        return {}

    for label, (outcome, predictor) in path_map.items():
        on_m = re.search(
            rf'{re.escape(outcome)}\s+ON\s+(.*?)(?=\n\s*\n\s*[A-Z\s]|\Z)',
            search_text, re.DOTALL | re.IGNORECASE)
        if not on_m:
            continue
        row = re.search(
            rf'^\s*{re.escape(predictor)}\s+([-\d.]+)',
            on_m.group(1), re.IGNORECASE | re.MULTILINE)
        if row:
            results[label] = {
                'est':       float(row.group(1)),
                'converged': converged,
                'stdyx':     bool(stdyx)
            }
    return results


def _fmt_p(p):
    """格式化 p 值"""
    if p is None: return '[待填]'
    if p < 0.001: return '< .001'
    return f'{p:.3f}'

def _fmt_est(v, digits=3):
    if v is None: return '[待填]'
    return f'{v:.{digits}f}'

def _sig_label(p):
    if p is None: return '[待填]'
    return 'Yes **' if p < 0.01 else ('Yes *' if p < 0.05 else 'No (n.s.)')


# ============================================================
# 5. 解析 Attrition 數字（從 Pipeline 報告）
# ============================================================
def parse_report(path):
    with open(path, 'r', encoding='utf-8') as f:
        txt = f.read()
    def ints(pattern):
        m = re.search(pattern, txt)
        return [int(x) for x in m.groups()] if m else None

    t1 = ints(r'T1.*?原始名單 (\d+) 人.*?通過注意力檢測 (\d+) 人.*?任職資格 (\d+) 人.*?有效樣本 (\d+) 人')
    t2 = ints(r'T2.*?原始名單 (\d+) 人.*?通過注意力檢測 (\d+) 人.*?配對回 T1 者 (\d+) 人')
    t3 = ints(r'T3.*?原始名單 (\d+) 人.*?通過注意力檢測 (\d+) 人.*?配對回 T1 者 (\d+) 人')
    g1 = ints(r'只有完成 T1.*?Group 1\).*?(\d+) 人')
    g2 = ints(r'完成 T1, T2 \(Group 2\).*?(\d+) 人')
    g3 = ints(r'完成 T1, T2, T3.*?Group 3\).*?(\d+) 人')
    return {
        't1': t1 or [510,460,435,433],
        't2': t2 or [396,391,389],
        't3': t3 or [346,344,340],
        'g1': (g1 or [[44]])[0], 'g2': (g2 or [[49]])[0], 'g3': (g3 or [[340]])[0]
    }

att = parse_report(report_path)
t1, t2, t3 = att['t1'], att['t2'], att['t3']

# ============================================================
# 5. 計算各波次統計
# ============================================================
s1 = wave_stats(df, 'T1')
s2 = wave_stats(df, 'T2')
s3 = wave_stats(df, 'T3')

# ============================================================
# 6. 建立各 DataFrame
# ============================================================

# ── CH3_問卷發放回收 ──────────────────────────────────────────
df_survey = pd.DataFrame([
    ["第一階段 (T1)",
     str(t1[0]), str(t1[0]), str(t1[0]-t1[3]), str(t1[3]),
     f"{t1[3]/t1[0]*100:.1f}%",
     f"剔除：注意力題未通過 {t1[0]-t1[1]} 人、不符就業資格 {t1[1]-t1[2]} 人、重複填答 {t1[2]-t1[3]} 人"],
    ["第二階段 (T2)",
     f"{t1[3]}*", str(t2[0]), str(t2[0]-t2[2]), str(t2[2]),
     f"{t2[2]/t1[3]*100:.1f}%",
     f"剔除：注意力題未通過 {t2[0]-t2[1]} 人、配對失敗 {t2[1]-t2[2]} 人"],
    ["第三階段 (T3)",
     f"{t2[2]}*", str(t3[0]), str(t3[0]-t3[2]), str(t3[2]),
     f"{t3[2]/t2[2]*100:.1f}%",
     f"最終三波完整配對樣本 N = {t3[2]}"],
], columns=["施測波次", "預計對象", "回收份數", "剔除份數", "有效份數", "有效回收率", "備註"])

# ── CH3_量表摘要（靜態，量表來源不會變）──────────────────────
df_scales_info = pd.DataFrame([
    ["職涯高原\nCareer Plateau (CP)", "Milliman (1992)",
     "英文原版\n由研究者翻譯",
     "2（HP 階層停滯\nJCP 工作內容停滯）",
     "12（各 6 題）",
     "五點 Likert\n1=非常不同意\n5=非常同意",
     "HP：題 4、6（R）\nJCP：題 1–5（R）",
     f"HP: α={s1['HP']['alpha']:.3f}\nJCP: α={s1['JCP']['alpha']:.3f}\nCP: α={s1['CP']['alpha']:.3f}",
     "「我在這間公司晉升的可能性是有限的。」\n「我的工作讓我感到有挑戰性。」（R）"],
    ["決策拖延\nDecisional Procrastination (DP)", "Mann et al. (1997)",
     "英文原版\n由研究者翻譯",
     "1（單維度）", "5",
     "五點 Likert\n1=完全不符合我\n5=完全符合我",
     "無",
     f"α={s1['DP']['alpha']:.3f}",
     "「在做出最終決定之前，我花了很多時間在處理瑣碎的事情上。」"],
    ["職涯無所作為\nCareer Inaction (CI)", "D'Huyvetter et al. (2025)",
     "英文原版\n由研究者翻譯",
     "1（單維度）",
     "8\n（另含 1 題注意力測試題，不計分）",
     "五點 Likert\n1=完全不同意\n5=完全同意",
     "無",
     f"α={s1['CI']['alpha']:.3f}",
     "「我想調整或改變自己的職涯，但我沒有積極追求。」"],
    ["主動性人格\nProactive Personality (PP)", "Parker (1998)；\n選自 Bateman & Crant (1993)",
     "英文原版\n由研究者翻譯",
     "1（單維度）", "6",
     "五點 Likert\n1=完全不同意\n5=完全同意",
     "無",
     f"α={s1['PP']['alpha']:.3f}",
     "「對於我看不順眼的事物，我會改正它。」"],
], columns=["構念（中/英文）", "量表來源", "原文語言", "向度數", "題數",
            "計分尺度", "反向題", f"本研究 Cronbach's α\n（T1，N={N}）", "範例題"])

# ── CH3_各波次敘述統計（M / SD / α）─────────────────────────
desc_rows = []
for code, label in [('HP','HP 階層停滯'),('JCP','JCP 工作內容停滯'),
                    ('CP','CP 職涯高原（合併）'),('PP','PP 主動性人格'),
                    ('DP','DP 決策拖延'),('CI','CI 職涯無所作為')]:
    desc_rows.append([
        label,
        f"{s1[code]['M']:.2f}", f"{s1[code]['SD']:.2f}", f"{s1[code]['alpha']:.3f}",
        f"{s2[code]['M']:.2f}", f"{s2[code]['SD']:.2f}", f"{s2[code]['alpha']:.3f}",
        f"{s3[code]['M']:.2f}", f"{s3[code]['SD']:.2f}", f"{s3[code]['alpha']:.3f}",
    ])
df_desc = pd.DataFrame(desc_rows, columns=[
    "構念", "T1 M", "T1 SD", "T1 α", "T2 M", "T2 SD", "T2 α", "T3 M", "T3 SD", "T3 α"])

# ── CH3_各波次相關矩陣（T1 / T2 / T3 各一張）────────────────
df_corr_t1 = corr_matrix_df(s1, N)
df_corr_t2 = corr_matrix_df(s2, N)
df_corr_t3 = corr_matrix_df(s3, N)

# ── CH3_CFA 適配指標（自動從 Mplus .out 讀取）──────────────────────
def _cfa_out(prefix):
    for f in os.listdir(latest_dir):
        if f.lower().startswith(prefix.lower()) and f.lower().endswith('.out'):
            return os.path.join(latest_dir, f)
    return None

def _fit_str(fit, key, fmt='.3f'):
    v = fit.get(key)
    return f'{v:{fmt}}' if v is not None else '[待填]'

def _cfa_row(label, out_prefix, note):
    out_p = _cfa_out(out_prefix)
    fit = parse_mplus_fit(out_p) if out_p else {}
    chi2_str = (f"{fit['chi2']:.3f}（{int(fit['df'])}）"
                if (fit.get('chi2') is not None and fit.get('df') is not None) else '[待填]')
    if fit.get('RMSEA') is not None:
        lo = fit.get('RMSEA_lo')
        hi = fit.get('RMSEA_hi')
        rmsea_str = (f"{fit['RMSEA']:.3f} [{lo:.3f}, {hi:.3f}]"
                     if (lo is not None and hi is not None) else f"{fit['RMSEA']:.3f}")
    else:
        rmsea_str = '[待填]'
    return [label, chi2_str,
            _fit_str(fit, 'CFI'), _fit_str(fit, 'TLI'),
            rmsea_str, _fit_str(fit, 'SRMR'), note]

df_cfa = pd.DataFrame([
    _cfa_row('CFA-A：三因子（JCP / DP / CI）',
             f'cfa_a_jcp_dp_ci_{ts}', '假設模型-JCP版'),
    _cfa_row('CFA-B：三因子（HP / DP / CI）',
             f'cfa_b_hp_dp_ci_{ts}',  '假設模型-HP版'),
    _cfa_row('CFA-C：四因子（JCP / PP / DP / CI）',
             f'cfa_c_jcp_pp_dp_ci_{ts}', '加入PP-JCP版'),
    _cfa_row('CFA-D：四因子（HP / PP / DP / CI）',
             f'cfa_d_hp_pp_dp_ci_{ts}',  '加入PP-HP版'),
    _cfa_row('M1：五因子（HP/JCP/PP/DP/CI）',
             f'CFA_M1_FiveFactor_{ts}', '完整五因子（T1）'),
    _cfa_row('M2：四因子（CP合併/PP/DP/CI）',
             f'CFA_M2_FourFactor_CP_merged_{ts}', '比較：HP+JCP合併'),
    _cfa_row('M3：三因子（CP/DP/CI）',
             f'CFA_M3_ThreeFactor_CP_DP_CI_{ts}', '比較：排除PP'),
    _cfa_row('RI-CLPM A (JCP, Bidir)',
             f'ri_clpm_a_jcp_bidir_{ts}', 'RI-CLPM 模型適配'),
    _cfa_row('RI-CLPM B (HP, Bidir)',
             f'ri_clpm_b_hp_bidir_{ts}', 'RI-CLPM 模型適配'),
    _cfa_row('RI-CLPM C (JCP, MultiGroup)',
             f'ri_clpm_c_jcp_multigrp_{ts}', 'H8 多群組模型'),
    _cfa_row('RI-CLPM D (HP, MultiGroup)',
             f'ri_clpm_d_hp_multigrp_{ts}', 'H8 多群組模型'),
], columns=["模型", "χ²（df）", "CFI", "TLI", "RMSEA", "SRMR", "備註"])

# ── CH3_樣本流失_ANOVA與卡方 ─────────────────────────────────
from scipy.stats import f_oneway, chi2_contingency

g1 = df_all[df_all['Group'] == 1].copy()
g2 = df_all[df_all['Group'] == 2].copy()
g3 = df_all[df_all['Group'] == 3].copy()
n1, n2, n3 = len(g1), len(g2), len(g3)

def grp_scale(grp, prefix, n_items, wave='T1'):
    cols = [f"{prefix}{i}_{wave}" for i in range(1, n_items+1)]
    valid = [c for c in cols if c in grp.columns]
    return grp[valid].mean(axis=1).dropna()

def anova_row(label, s1, s2, s3):
    F, p = f_oneway(s1, s2, s3)
    stars = '***' if p < .001 else ('**' if p < .01 else ('*' if p < .05 else 'n.s.'))
    def ms(s): return f"{s.mean():.2f} ({s.std():.2f})"
    return [label, ms(s1), ms(s2), ms(s3), f"{F:.3f}", f"{p:.3f}", stars]

# 計算各構念在三組的分佈
age1 = g1['Age'].dropna(); age2 = g2['Age'].dropna(); age3 = g3['Age'].dropna()

hp1  = grp_scale(g1,'HP',6);  hp2  = grp_scale(g2,'HP',6);  hp3  = grp_scale(g3,'HP',6)
jcp1 = grp_scale(g1,'JCP',6); jcp2 = grp_scale(g2,'JCP',6); jcp3 = grp_scale(g3,'JCP',6)

def cp_s(grp):
    h = grp_scale(grp,'HP',6); j = grp_scale(grp,'JCP',6)
    return pd.concat([h,j], axis=1).mean(axis=1).dropna()

dp1 = grp_scale(g1,'DP',5); dp2 = grp_scale(g2,'DP',5); dp3 = grp_scale(g3,'DP',5)
ci1 = grp_scale(g1,'CI',8); ci2 = grp_scale(g2,'CI',8); ci3 = grp_scale(g3,'CI',8)
pp1 = grp_scale(g1,'PP',6); pp2 = grp_scale(g2,'PP',6); pp3 = grp_scale(g3,'PP',6)

anova_rows = [
    anova_row('年齡（Age）',          age1, age2, age3),
    anova_row('職涯高原 (CP)',        cp_s(g1), cp_s(g2), cp_s(g3)),
    anova_row('階層停滯 (HP)',        hp1, hp2, hp3),
    anova_row('工作內容停滯 (JCP)',   jcp1, jcp2, jcp3),
    anova_row('決策拖延 (DP)',        dp1, dp2, dp3),
    anova_row('職涯無所作為 (CI)',    ci1, ci2, ci3),
    anova_row('主動性人格 (PP)',      pp1, pp2, pp3),
]

df_anova = pd.DataFrame(anova_rows, columns=[
    f"變數",
    f"Group 1 僅 T1\n(n={n1})  M (SD)",
    f"Group 2 T1+T2\n(n={n2})  M (SD)",
    f"Group 3 三波完整\n(n={n3})  M (SD)",
    "F 值", "p 值", "顯著性"])

# 卡方：性別 & 教育程度
def chi2_row(label, col):
    ct = pd.crosstab(df_all[col], df_all['Group'].map({1:'G1',2:'G2',3:'G3'}))
    chi2, p, dof, _ = chi2_contingency(ct)
    stars = '***' if p < .001 else ('**' if p < .01 else ('*' if p < .05 else 'n.s.'))
    return [label, f"{chi2:.3f}", str(dof), f"{p:.3f}", stars]

chi2_rows = []
if 'Gender' in df_all.columns:
    chi2_rows.append(chi2_row('性別 (Gender)', 'Gender'))
if 'Education' in df_all.columns:
    chi2_rows.append(chi2_row('教育程度 (Education)', 'Education'))

df_chi2 = pd.DataFrame(chi2_rows, columns=["變數", "χ²", "df", "p 值", "顯著性"])

# ── CH4_假設驗證結果表（自動從 Mplus .out 填入）───────────────────
P = '[待填]'   # fallback placeholder

# 找 RI-CLPM .out 檔
def _out(prefix):
    p = os.path.join(latest_dir, f'{prefix}_{ts}.out'.lower())
    if not os.path.isfile(p):
        # 檔名可能大小寫不同
        for f in os.listdir(latest_dir):
            if f.lower() == f'{prefix}_{ts}.out'.lower():
                return os.path.join(latest_dir, f)
    return p

out_A = _out('RI_CLPM_A_JCP_Bidir')
out_B = _out('RI_CLPM_B_HP_Bidir')
out_C = _out('RI_CLPM_C_JCP_MultiGrp')
out_D = _out('RI_CLPM_D_HP_MultiGrp')

# 路徑對應表（STDYX）
pm_A = {
    'H1b_JCP_DP': ('WDP2', 'WJCP1'),
    'H2b_JCP_CI': ('WCI2', 'WJCP1'),
    'H3_DP_CI_A': ('WCI2', 'WDP1'),
    'H4b_DP_JCP': ('WJCP2', 'WDP1'),
    'H5_CI_DP_A': ('WDP2', 'WCI1'),
    'H6b_CI_JCP': ('WJCP2', 'WCI1'),
}
pm_B = {
    'H1a_HP_DP':  ('WDP2', 'WHP1'),
    'H2a_HP_CI':  ('WCI2', 'WHP1'),
    'H3_DP_CI_B': ('WCI2', 'WDP1'),
    'H4a_DP_HP':  ('WHP2', 'WDP1'),
    'H5_CI_DP_B': ('WDP2', 'WCI1'),
    'H6a_CI_HP':  ('WHP2', 'WCI1'),
}
pm_CJcp = {  # for multi-group paths (configural)
    'JCP_DP': ('WDP2', 'WJCP1'),
    'JCP_CI': ('WCI2', 'WJCP1'),
    'DP_CI':  ('WCI2', 'WDP1'),
    'DP_JCP': ('WJCP2', 'WDP1'),
    'CI_DP':  ('WDP2', 'WCI1'),
    'CI_JCP': ('WJCP2', 'WCI1'),
}
pm_DHp = {
    'HP_DP':  ('WDP2', 'WHP1'),
    'HP_CI':  ('WCI2', 'WHP1'),
    'DP_CI':  ('WCI2', 'WDP1'),
    'DP_HP':  ('WHP2', 'WDP1'),
    'CI_DP':  ('WDP2', 'WCI1'),
    'CI_HP':  ('WHP2', 'WCI1'),
}

stdyx_A = parse_mplus_stdyx(out_A, pm_A)
stdyx_B = parse_mplus_stdyx(out_B, pm_B)
unstd_A = parse_mplus_unstd(out_A, pm_A)
unstd_B = parse_mplus_unstd(out_B, pm_B)

# 多群組各組路徑
n_high = int(df[df.get('PP_group', pd.Series([np.nan]*len(df))) == 1].shape[0]) if 'PP_group' in df.columns else 0
n_low  = int(df[df.get('PP_group', pd.Series([np.nan]*len(df))) == 0].shape[0]) if 'PP_group' in df.columns else 0
mg_high_C = parse_multigroup_paths(out_C, pm_CJcp, 'HIGHPP')
mg_low_C  = parse_multigroup_paths(out_C, pm_CJcp, 'LOWPP')
mg_high_D = parse_multigroup_paths(out_D, pm_DHp,  'HIGHPP')
mg_low_D  = parse_multigroup_paths(out_D, pm_DHp,  'LOWPP')


def _row(hyp_id, path_label, mplus_sym, model, direction, stdyx_dict, key, unstd_dict=None):
    """Build one hypothesis row, filling in values from Mplus dicts."""
    d   = stdyx_dict.get(key, {})
    ud  = (unstd_dict or {}).get(key, {})
    b   = _fmt_est(ud.get('b'),  3) if ud.get('b')  is not None else P
    se  = _fmt_est(ud.get('se'), 3) if ud.get('se') is not None else P
    p   = ud.get('p') if ud.get('p') is not None else d.get('p')
    p_s = _fmt_p(p)
    ci_lo = ud.get('ci_lo')
    ci_hi = ud.get('ci_hi')
    ci_s  = f'[{_fmt_est(ci_lo,3)}, {_fmt_est(ci_hi,3)}]' if (ci_lo is not None and ci_hi is not None) else f'[{P}, {P}]'
    beta  = _fmt_est(d.get('est'), 3) if d.get('est') is not None else P
    supp  = _sig_label(p) if p is not None else P
    return [hyp_id, path_label, mplus_sym, model, direction, b, se, p_s, ci_s, beta, supp]


hyp_rows = [
    # ── 主軸交叉延遲（Model A: JCP, Model B: HP）──
    _row('H1a', 'HP → DP\n（職涯高原→決策拖延）',  'WHP → WDP',  'Model B', '正向（+）',
         stdyx_B, 'H1a_HP_DP', unstd_B),
    _row('H1b', 'JCP → DP\n（職涯高原→決策拖延）', 'WJCP → WDP', 'Model A', '正向（+）',
         stdyx_A, 'H1b_JCP_DP', unstd_A),
    _row('H2a', 'HP → CI\n（職涯高原→職涯無所作為）', 'WHP → WCI', 'Model B', '正向（+）',
         stdyx_B, 'H2a_HP_CI', unstd_B),
    _row('H2b', 'JCP → CI\n（職涯高原→職涯無所作為）','WJCP → WCI','Model A','正向（+）',
         stdyx_A, 'H2b_JCP_CI', unstd_A),
    _row('H3',  'DP → CI\n（決策拖延→職涯無所作為）', 'WDP → WCI', 'Model A/B','正向（+）',
         stdyx_A, 'H3_DP_CI_A', unstd_A),
    # ── 反向交叉延遲 ──
    _row('H4a', 'DP → HP\n（決策拖延→職涯高原）',  'WDP → WHP',  'Model B', '正向（+）',
         stdyx_B, 'H4a_DP_HP', unstd_B),
    _row('H4b', 'DP → JCP\n（決策拖延→職涯高原）', 'WDP → WJCP', 'Model A', '正向（+）',
         stdyx_A, 'H4b_DP_JCP', unstd_A),
    _row('H5',  'CI → DP\n（職涯無所作為→決策拖延）','WCI → WDP', 'Model A/B','正向（+）',
         stdyx_A, 'H5_CI_DP_A', unstd_A),
    _row('H6a', 'CI → HP\n（職涯無所作為→職涯高原）','WCI → WHP', 'Model B',  '正向（+）',
         stdyx_B, 'H6a_CI_HP', unstd_B),
    _row('H6b', 'CI → JCP\n（職涯無所作為→職涯高原）','WCI → WJCP','Model A','正向（+）',
         stdyx_A, 'H6b_CI_JCP', unstd_A),
    # ── 中介（indirect — 需手動從 Mplus output 讀取） ──
    ['H7a', 'HP_T1 → DP_T2 → CI_T3\n（決策拖延中介）', 'Indirect', 'Model B', '正向（+）',
     P, P, P, f'[{P}, {P}]', P, P],
    ['H7b', 'JCP_T1 → DP_T2 → CI_T3\n（決策拖延中介）','Indirect', 'Model A', '正向（+）',
     P, P, P, f'[{P}, {P}]', P, P],
    # ── 主動型人格調節（多群組 χ² 差異檢定） ──
    ['H8a', 'PP 調節 CP(HP)→DP、CP(HP)→CI\n（高PP路徑 < 低PP路徑）',
     'chi-sq diff test', 'Model D', '交互作用（調節）', P, P, P, f'[{P}, {P}]', P, P],
    ['H8b', 'PP 調節 CP(JCP)→DP、CP(JCP)→CI\n（高PP路徑 < 低PP路徑）',
     'chi-sq diff test', 'Model C', '交互作用（調節）', P, P, P, f'[{P}, {P}]', P, P],
]

df_hyp = pd.DataFrame(hyp_rows, columns=[
    '假設', '路徑', 'Mplus 路徑符號', '對應模型', '預測方向',
    'b（非標準化）', 'SE', 'p 值', '95% CI', 'β（標準化）', '支持與否'])

# ── CH4_多群組PP比較（H8：多群組 RI-CLPM，configural 各組係數）──────
def _mg_row(path_label, key_hp=None, key_jcp=None):
    """Build multi-group comparison row."""
    if 'HP' in path_label:
        key = key_hp or path_label.replace(' ', '_').replace('→', '').replace('（', '').replace('）', '')
        hi_d = mg_high_D.get(key, {})
        lo_d = mg_low_D.get(key, {})
    else:
        key = key_jcp or path_label.replace(' ', '_').replace('→', '').replace('（', '').replace('）', '')
        hi_d = mg_high_C.get(key, {})
        lo_d = mg_low_C.get(key, {})

    converged_hi = hi_d.get('converged', False)
    converged_lo = lo_d.get('converged', False)
    suffix = '' if (converged_hi and converged_lo) else '*'  # * = 未收斂

    hi_b = (_fmt_est(hi_d.get('est'), 3) + suffix) if hi_d.get('est') is not None else P
    lo_b = (_fmt_est(lo_d.get('est'), 3) + suffix) if lo_d.get('est') is not None else P
    diff = (f"{hi_d['est']-lo_d['est']:+.3f}{suffix}"
            if (hi_d.get('est') is not None and lo_d.get('est') is not None) else P)
    return [path_label, hi_b, lo_b, diff, P, P, P, '需重跑約束模型' if not (converged_hi and converged_lo) else P]

n_high_str = str(n_high) if n_high > 0 else '[待填]'
n_low_str  = str(n_low)  if n_low  > 0 else '[待填]'

mg_rows = [
    _mg_row('HP → DP',  key_hp='HP_DP'),
    _mg_row('HP → CI',  key_hp='HP_CI'),
    _mg_row('JCP → DP', key_jcp='JCP_DP'),
    _mg_row('JCP → CI', key_jcp='JCP_CI'),
    _mg_row('DP → CI',  key_hp='DP_CI',  key_jcp='DP_CI'),
    _mg_row('CI → DP',  key_hp='CI_DP',  key_jcp='CI_DP'),
]
df_mg = pd.DataFrame(mg_rows, columns=[
    '路徑',
    f'高PP組（n={n_high_str}）', f'低PP組（n={n_low_str}）',
    '差值（高-低）', 'ΔSB-χ²', 'Δdf',
    'p 值（SB 修正）', '備註'])

# Add convergence note
mg_note = pd.DataFrame([
    ['注意', '多群組 configural 模型未收斂（自由度為負，參數過多）。',
     '建議改跑分組RI-CLPM（High/Low PP 各別估計）或約束模型（cross-lagged paths constrained equal）。',
     '', '', '', '', '']
], columns=df_mg.columns)

# ── 舊有分析表（保留）────────────────────────────────────────
df_attrition_old = pd.DataFrame([
    ["整體職涯停滯 (CP)", "p = .355", "無顯著差異"],
    ["階層停滯 (HP)", "p = .215", "無顯著差異"],
    ["工作停滯 (JCP)", "p = .491", "無顯著差異"],
    ["決策拖延 (DP)", "p = .117", "無顯著差異"],
    ["職涯無所作為 (CI)", "p = .630", "無顯著差異"],
    ["主動型人格 (PP)", "p = .571", "無顯著差異"]
], columns=["研究變數", "Levene / t 檢定 (p value)", "檢驗結果判斷"])

# ============================================================
# 7. 輸出 Excel
# ============================================================
with pd.ExcelWriter(EXCEL_OUT, engine='openpyxl') as writer:
    df_survey.to_excel(writer,        sheet_name='CH3_問卷發放回收',    index=False)
    df_scales_info.to_excel(writer,   sheet_name='CH3_量表摘要',        index=False)
    df_desc.to_excel(writer,          sheet_name='CH3_各波次描述統計',   index=False)
    df_corr_t1.to_excel(writer,       sheet_name='CH3_相關矩陣_T1',      index=False)
    df_corr_t2.to_excel(writer,       sheet_name='CH3_相關矩陣_T2',      index=False)
    df_corr_t3.to_excel(writer,       sheet_name='CH3_相關矩陣_T3',      index=False)
    df_cfa.to_excel(writer,           sheet_name='CH3_CFA適配指標',      index=False)
    df_anova.to_excel(writer,         sheet_name='CH3_流失ANOVA',         index=False)
    df_chi2.to_excel(writer,          sheet_name='CH3_流失卡方',          index=False)
    df_hyp.to_excel(writer,           sheet_name='CH4_假設驗證結果',      index=False)
    df_mg.to_excel(writer,            sheet_name='CH4_多群組PP比較',       index=False)
    df_attrition_old.to_excel(writer, sheet_name='樣本流失檢定_舊',      index=False)

# ============================================================
# 8. SPSS 學術格式樣式套用
# ============================================================
wb = openpyxl.load_workbook(EXCEL_OUT)

thick  = Side(border_style="thick", color="000000")
thin   = Side(border_style="thin",  color="000000")
none_s = Side(border_style=None)

FILL_HEADER  = PatternFill("solid", start_color="D9D9D9")
FILL_PENDING = PatternFill("solid", start_color="FFF2CC")
FILL_NOTE    = PatternFill("solid", start_color="E2EFDA")
FILL_DIAG    = PatternFill("solid", start_color="F2F2F2")

LEFT_COLS = {
    'CH3_問卷發放回收':   [1, 7],
    'CH3_量表摘要':       [1, 2, 4, 6, 9],
    'CH3_各波次描述統計': [1],
    'CH3_相關矩陣_T1':    [1],
    'CH3_相關矩陣_T2':    [1],
    'CH3_相關矩陣_T3':    [1],
    'CH3_CFA適配指標':    [1, 7],
    'CH3_流失ANOVA':      [1, 2, 3, 4],
    'CH3_流失卡方':       [1],
    'CH4_假設驗證結果':   [1, 2, 4],
    'CH4_多群組PP比較':   [1],
    '樣本流失檢定_舊':    [1],
}

for sname in wb.sheetnames:
    ws      = wb[sname]
    max_row = ws.max_row
    max_col = ws.max_column
    lcols   = LEFT_COLS.get(sname, [])

    # 自動欄寬
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len    = 0
        for cell in col:
            for line in str(cell.value or '').split('\n'):
                ln = sum(2 if ord(c) > 127 else 1 for c in line)
                if ln > max_len:
                    max_len = ln
        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

    for ridx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), 1):
        is_header = ridx == 1
        is_last   = ridx == max_row
        for cell in row:
            cidx  = cell.column
            is_lf = cidx in lcols
            val   = str(cell.value or '')

            cell.font = Font(name='Times New Roman', size=12, bold=is_header)
            cell.alignment = Alignment(
                horizontal='left' if is_lf else 'center',
                vertical='center', wrap_text=True)

            # 填色
            if is_header:
                cell.fill = FILL_HEADER
            elif '[待填]' in val:
                cell.fill = FILL_PENDING
                cell.font = Font(name='Times New Roman', size=12, color='C00000')
            elif sname == 'CH3_問卷發放回收' and cidx == 7 and not is_header:
                cell.fill = FILL_NOTE
            elif sname.startswith('CH3_相關矩陣') and not is_header:
                # 對角線（括號 α）
                if val.startswith('(') and val.endswith(')'):
                    cell.fill = FILL_DIAG

            # Borders（SPSS 格式：只有橫線）
            top_b    = thick if is_header else none_s
            bottom_b = thin  if is_header else (thick if is_last else none_s)
            cell.border = Border(top=top_b, bottom=bottom_b,
                                 left=none_s, right=none_s)

    ws.freeze_panes = ws['A2']

wb.save(EXCEL_OUT)
print(f"[OK] Excel output: {EXCEL_OUT}")
print(f"     N={N}  |  ts={ts}  |  Sheets: {', '.join(wb.sheetnames)}")
